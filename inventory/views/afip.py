"""Utilidades para importar comprobantes AFIP desde 'Mis Comprobantes'."""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import IntegrityError

from ..models import AFIPInvoice


def _parse_dec(val) -> Decimal:
    if val is None or val == "":
        return Decimal("0.00")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def _parse_date(val):
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_tipo_codigo(val) -> int:
    try:
        return int(str(val).split("-")[0].strip().split()[0])
    except (ValueError, IndexError):
        return 0


def _find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and str(row[0]).strip().lower() == "fecha":
            return i
    return None


def _col(row, idx, default=None):
    try:
        v = row[idx]
        return v if v is not None else default
    except IndexError:
        return default


def _parse_afip_xlsx(file_obj):
    """
    Parsea un xlsx de 'Mis Comprobantes Recibidos' de AFIP.
    Devuelve (created, skipped, errors, error_msg) donde error_msg es None si el archivo es válido.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
    except Exception as e:
        return 0, 0, 0, f"No se pudo leer el archivo: {e}"

    header_row = _find_header_row(ws)
    if header_row is None:
        return 0, 0, 0, "No se encontró la fila de encabezados (se espera 'Fecha' en columna A)."

    # Columnas (0-indexed) del formato AFIP "Mis Comprobantes Recibidos":
    # 0=Fecha, 1=Tipo, 2=Punto de Venta, 3=Número Desde, 5=CAE,
    # 7=CUIT Emisor, 8=Razón Social, 18=IVA 10.5%, 19=Neto 10.5%,
    # 20=IVA 21%, 21=Neto 21%, 22=IVA 27%, 23=Neto 27%, 28=Total IVA, 29=Imp Total

    created = skipped = errors = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(v is None or v == "" for v in row[:5]):
            continue

        fecha = _parse_date(_col(row, 0))
        if fecha is None:
            continue

        tipo_codigo = _parse_tipo_codigo(str(_col(row, 1, "")))
        if tipo_codigo not in (AFIPInvoice.FACTURA_A, AFIPInvoice.NOTA_CREDITO_A):
            skipped += 1
            continue

        try:
            punto_venta = int(_col(row, 2, 0) or 0)
            numero = int(_col(row, 3, 0) or 0)
        except (ValueError, TypeError):
            errors += 1
            continue

        try:
            _, was_created = AFIPInvoice.objects.get_or_create(
                cuit_emisor=str(_col(row, 7, "") or "").strip(),
                punto_venta=punto_venta,
                numero=numero,
                tipo_codigo=tipo_codigo,
                defaults=dict(
                    date=fecha,
                    tipo_descripcion=str(_col(row, 1, "") or "").strip(),
                    cae=str(_col(row, 5, "") or "").strip(),
                    razon_social=str(_col(row, 8, "") or "").strip(),
                    iva_105=_parse_dec(_col(row, 18)),
                    neto_105=_parse_dec(_col(row, 19)),
                    iva_21=_parse_dec(_col(row, 20)),
                    neto_21=_parse_dec(_col(row, 21)),
                    iva_27=_parse_dec(_col(row, 22)),
                    neto_27=_parse_dec(_col(row, 23)),
                    total_iva=_parse_dec(_col(row, 28)),
                    imp_total=_parse_dec(_col(row, 29)),
                ),
            )
            if was_created:
                created += 1
            else:
                skipped += 1
        except IntegrityError:
            skipped += 1

    return created, skipped, errors, None
