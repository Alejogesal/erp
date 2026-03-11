"""Excel import/export helpers."""
from decimal import Decimal
import io
import re
import unicodedata
import zipfile

from xml.sax.saxutils import escape


def _col_letter(idx: int) -> str:
    result = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _build_xlsx(
    headers: list[str],
    rows: list[list[str | Decimal]],
    *,
    blue_cols: set[int] | None = None,
    number_cols: set[int] | None = None,
) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Root relations
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
            "</Types>",
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "docProps/core.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
            'xmlns:dc="http://purl.org/dc/elements/1.1/" '
            'xmlns:dcterms="http://purl.org/dc/terms/" '
            'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
            "<dc:title>Precios</dc:title>"
            "</cp:coreProperties>",
        )
        zf.writestr(
            "docProps/app.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
            'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
            "<Application>Django</Application>"
            "</Properties>",
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Precios" sheetId="1" r:id="rId1"/></sheets>'
            "</workbook>",
        )
        zf.writestr(
            "xl/styles.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<fonts count="3">'
            '<font><sz val="11"/><color theme="1"/><name val="Calibri"/></font>'
            '<font><b/><sz val="11"/><color theme="1"/><name val="Calibri"/></font>'
            '<font><sz val="11"/><color rgb="FF1F4EAD"/><name val="Calibri"/></font>'
            "</fonts>"
            '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="4">'
            '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
            '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
            '<xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
            '<xf numFmtId="2" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
            "</cellXfs>"
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            "</styleSheet>",
        )

        cols = len(headers)
        rows_count = len(rows) + 1  # header row
        dimension = f"A1:{_col_letter(cols)}{rows_count}"

        blue_cols = blue_cols or set()
        number_cols = number_cols or set()

        def cell_xml(value, col_idx, row_idx, is_header=False):
            col = _col_letter(col_idx)
            ref = f"{col}{row_idx}"
            if is_header:
                return f'<c r="{ref}" t="inlineStr" s="1"><is><t>{escape(str(value))}</t></is></c>'
            if isinstance(value, str):
                style = "2" if col_idx in blue_cols else "0"
                return f'<c r="{ref}" t="inlineStr" s="{style}"><is><t>{escape(str(value))}</t></is></c>'
            style = "3" if col_idx in number_cols else "0"
            return f'<c r="{ref}" s="{style}"><v>{value}</v></c>'

        max_lengths = [len(str(h)) for h in headers]
        for row in rows:
            for idx, val in enumerate(row):
                if idx >= len(max_lengths):
                    continue
                if isinstance(val, Decimal):
                    text = f"{val:.2f}" if (idx + 1) in number_cols else str(val)
                else:
                    text = str(val)
                if len(text) > max_lengths[idx]:
                    max_lengths[idx] = len(text)
        col_widths = []
        for length in max_lengths:
            col_widths.append(min(60, max(8, round(length * 1.1 + 2, 2))))

        rows_xml = []
        header_cells = "".join(cell_xml(h, i + 1, 1, is_header=True) for i, h in enumerate(headers))
        rows_xml.append(f'<row r="1">{header_cells}</row>')
        for ridx, row in enumerate(rows, start=2):
            cells = "".join(cell_xml(val, cidx + 1, ridx) for cidx, val in enumerate(row))
            rows_xml.append(f'<row r="{ridx}">{cells}</row>')

        cols_xml = "".join(
            f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
            for idx, width in enumerate(col_widths, start=1)
        )
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            f' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<dimension ref="{dimension}"/>'
            f"<cols>{cols_xml}</cols>"
            "<sheetData>"
            f'{"".join(rows_xml)}'
            "</sheetData>"
            "</worksheet>"
        )
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buf.getvalue()


def _decimal_or_zero(value: str | None) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _parse_decimal(value: str | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    cleaned = str(value).strip().replace(".", "").replace(",", ".")
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if cleaned == "":
        return Decimal("0.00")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _abbr(text: str | None, length: int) -> str:
    if not text:
        return "X" * length
    cleaned = re.sub(r"[^A-Za-z0-9]", "", text).upper()
    return cleaned[:length].ljust(length, "X")


def _sku_prefix(group: str, description: str) -> str:
    words = [w for w in re.split(r"\s+", (description or "").strip()) if w]
    word1 = _abbr(words[0], 3) if len(words) > 0 else "XXX"
    word2 = _abbr(words[1], 3) if len(words) > 1 else "XXX"
    return f"{_abbr(group, 4)}{word1}{word2}"


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    without_accents = "".join(c for c in normalized if not unicodedata.combining(c))
    return without_accents.strip().lower()


def _read_costs_xlsx_rows(upload) -> tuple[list[tuple[str, str, Decimal, str]], str | None]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], "Falta la dependencia openpyxl. Instalá openpyxl en el entorno."

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active
    except Exception:
        return [], "No se pudo leer el archivo XLSX."

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    header_map = {_normalize_header(h): idx for idx, h in enumerate(headers)}
    desc_keys = ["descripcion", "descripción", "producto", "descripcion producto", "nombre"]
    cost_keys = ["precio venta", "precio", "costo", "costo unitario", "precio costo", "precio venta unitario"]
    group_keys = ["grupo", "marca", "categoria", "categoría"]
    sku_keys = ["sku", "codigo", "código", "cod", "codigo sku"]

    def _pick_index(keys: list[str]) -> int | None:
        for key in keys:
            normalized = _normalize_header(key)
            if normalized in header_map:
                return header_map[normalized]
        return None

    desc_idx = _pick_index(desc_keys)
    cost_idx = _pick_index(cost_keys)
    group_idx = _pick_index(group_keys)
    sku_idx = _pick_index(sku_keys)

    if desc_idx is None or cost_idx is None:
        return [], "Faltan columnas obligatorias: Descripción/Producto y Precio/Costo."

    rows: list[tuple[str, str, Decimal, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        group = str(row[group_idx] or "").strip() if group_idx is not None else ""
        description = str(row[desc_idx] or "").strip()
        cost = _parse_decimal(row[cost_idx])
        sku = str(row[sku_idx] or "").strip() if sku_idx is not None else ""
        if not description:
            continue
        rows.append((group, description, cost, sku))
    return rows, None


def _read_ml_sales_xlsx_rows(upload) -> tuple[list[dict], str | None]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], "Falta la dependencia openpyxl. Instalá openpyxl en el entorno."

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active
    except Exception:
        return [], "No se pudo leer el archivo XLSX."

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in header_cells
    ]
    header_map = {_normalize_header(h): idx for idx, h in enumerate(headers)}

    def _pick_index(keys: list[str]) -> int | None:
        for key in keys:
            normalized = _normalize_header(key)
            if normalized in header_map:
                return header_map[normalized]
        return None

    date_idx = _pick_index(["fecha", "fecha venta"])
    order_idx = _pick_index(
        [
            "comprobante",
            "nro comprobante",
            "número comprobante",
            "numero comprobante",
            "orden",
            "nro orden",
            "número orden",
            "numero orden",
            "pedido",
            "nro pedido",
            "número pedido",
            "numero pedido",
            "order id",
            "id pedido",
        ]
    )
    product_idx = _pick_index(["producto", "publicacion", "publicación", "titulo", "título", "nombre"])
    qty_idx = _pick_index(["cantidad", "cant"])
    price_total_idx = _pick_index(["precio bruto venta", "precio bruto", "precio"])
    commission_idx = _pick_index(["comision", "comisión"])
    tax_idx = _pick_index(["impuestos", "impuesto", "iibb"])

    required = [date_idx, product_idx, qty_idx, price_total_idx, commission_idx, tax_idx]
    if any(idx is None for idx in required):
        return [], "Faltan columnas obligatorias: Fecha, Producto, Cantidad, Precio Bruto Venta, Comision, Impuestos."

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = str(row[product_idx] or "").strip() if product_idx is not None else ""
        if not title:
            continue
        qty = _parse_decimal(row[qty_idx])
        if qty <= 0:
            continue
        price_total = _parse_decimal(row[price_total_idx])
        commission = _parse_decimal(row[commission_idx])
        taxes = _parse_decimal(row[tax_idx])
        created_at = row[date_idx] if date_idx is not None else None
        order_id = str(row[order_idx] or "").strip() if order_idx is not None else ""
        rows.append(
            {
                "title": title,
                "quantity": qty,
                "price_total": price_total,
                "commission": commission,
                "taxes": taxes,
                "created_at": created_at,
                "order_id": order_id,
            }
        )
    return rows, None


def _process_costs_xlsx(
    upload,
    *,
    group_override: str | None = None,
) -> tuple[int, int, int] | str:
    from ..models import Product

    rows, error = _read_costs_xlsx_rows(upload)
    if error:
        return error

    created = 0
    updated = 0
    skipped = 0
    prefix_counters: dict[str, int] = {}

    effective_group = (group_override or "").strip()

    for group, description, cost, sku in rows:
        if effective_group:
            group = effective_group

        product = None
        if sku:
            product = Product.objects.filter(sku__iexact=sku).first()
        if not product:
            product = Product.objects.filter(name=description, group=group).first()
        if product:
            product.avg_cost = cost
            product.save(update_fields=["avg_cost"])
            updated += 1
            continue

        if effective_group:
            skipped += 1
            continue

        prefix = _sku_prefix(group, description)
        if prefix not in prefix_counters:
            existing = (
                Product.objects.filter(sku__startswith=prefix)
                .values_list("sku", flat=True)
            )
            max_suffix = 0
            for sku in existing:
                suffix = sku[len(prefix):]
                if suffix.isdigit():
                    max_suffix = max(max_suffix, int(suffix))
            prefix_counters[prefix] = max_suffix

        prefix_counters[prefix] += 1
        sku = f"{prefix}{prefix_counters[prefix]:04d}"

        Product.objects.create(
            sku=sku,
            name=description,
            group=group,
            avg_cost=cost,
        )
        created += 1

    return created, updated, skipped
