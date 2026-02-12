from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import os
import base64
import calendar

from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.utils import OperationalError
from django.db.models import Case, DecimalField, Sum, Value, When, Q, Count, Subquery, OuterRef
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce
from django.forms import formset_factory
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from urllib.request import Request, urlopen
import csv
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.text import slugify
from datetime import datetime, time, timedelta
import io
import zipfile
import re
import unicodedata
import json
import secrets
from xml.sax.saxutils import escape
from django.core.files.base import ContentFile

from . import services
from . import mercadolibre as ml
from .models import (
    Customer,
    CustomerGroupDiscount,
    CustomerProductDiscount,
    CustomerPayment,
    MercadoLibreNotification,
    MercadoLibreConnection,
    MercadoLibreItem,
    Product,
    ProductVariant,
    Purchase,
    PurchaseItem,
    Stock,
    StockMovement,
    Supplier,
    SupplierProduct,
    Warehouse,
    Sale,
    SaleItem,
    TaxExpense,
)


def _products_with_last_cost_queryset():
    supplier_cost = (
        SupplierProduct.objects.filter(product=OuterRef("pk"))
        .order_by("-last_purchase_at", "-id")
        .values("last_cost")[:1]
    )
    movement_cost = (
        StockMovement.objects.filter(product=OuterRef("pk"), movement_type=StockMovement.MovementType.ENTRY)
        .order_by("-created_at", "-id")
        .values("unit_cost")[:1]
    )
    return Product.objects.annotate(
        last_purchase_cost_value=Coalesce(
            Subquery(supplier_cost, output_field=DecimalField(max_digits=12, decimal_places=2)),
            Subquery(movement_cost, output_field=DecimalField(max_digits=12, decimal_places=2)),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=12, decimal_places=2),
        )
    )


def _product_label_with_last_cost(obj: Product) -> str:
    last_cost = getattr(obj, "last_purchase_cost_value", None)
    if last_cost is None:
        last_cost = obj.last_purchase_cost()
    return f"{obj.sku or 'Sin SKU'} - {obj.name} (último costo: {last_cost:.2f})"


def _apply_product_queryset_to_formset(formset, products_qs):
    for form in formset.forms:
        if "product" in form.fields:
            form.fields["product"].queryset = products_qs
    if hasattr(formset, "empty_form") and "product" in formset.empty_form.fields:
        formset.empty_form.fields["product"].queryset = Product.objects.none()


def _extract_product_ids_from_payload(raw_payload: str | None) -> set[int]:
    if not raw_payload:
        return set()
    try:
        items_raw = json.loads(raw_payload)
    except Exception:
        return set()
    if not isinstance(items_raw, list):
        return set()
    ids: set[int] = set()
    for entry in items_raw:
        if not isinstance(entry, dict):
            continue
        product_id = entry.get("product_id")
        if not product_id:
            continue
        try:
            ids.add(int(product_id))
        except Exception:
            continue
    return ids


class ProductForm(forms.ModelForm):
    class Meta:
        model = Product
        fields = [
            "sku",
            "name",
            "group",
            "avg_cost",
            "vat_percent",
            "margin_consumer",
            "margin_barber",
            "margin_distributor",
            "default_supplier",
        ]
        labels = {
            "avg_cost": "Costo unitario",
            "vat_percent": "IVA %",
            "sku": "SKU",
            "name": "Nombre",
            "group": "Grupo / Marca",
            "margin_consumer": "Margen % consumidor final",
            "margin_barber": "Margen % peluquerías/barberías",
            "margin_distributor": "Margen % distribuidores",
            "default_supplier": "Proveedor principal",
        }


class PurchaseHeaderForm(forms.Form):
    warehouse = forms.ModelChoiceField(queryset=Warehouse.objects.all())
    purchase_date = forms.DateField(
        label="Fecha",
        required=False,
        widget=forms.DateInput(attrs={"type": "date"}),
    )
    descuento_total = forms.DecimalField(
        label="Descuento %",
        min_value=Decimal("0.00"),
        max_value=Decimal("100.00"),
        decimal_places=2,
        required=False,
        initial=Decimal("0.00"),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if not self.initial.get("warehouse"):
            first_warehouse = Warehouse.objects.order_by("name").first()
            if first_warehouse:
                self.initial["warehouse"] = first_warehouse
        if not self.initial.get("purchase_date"):
            self.initial["purchase_date"] = timezone.localdate()


class PurchaseItemForm(forms.Form):
    product = forms.ModelChoiceField(queryset=Product.objects.all())
    quantity = forms.IntegerField(min_value=1)
    unit_cost = forms.DecimalField(min_value=Decimal("0.00"), decimal_places=2)
    discount_percent = forms.DecimalField(
        label="DTO %",
        min_value=Decimal("0.00"),
        max_value=Decimal("100.00"),
        decimal_places=2,
        required=False,
        initial=Decimal("0.00"),
    )
    supplier = forms.ModelChoiceField(queryset=Supplier.objects.all(), label="Proveedor", required=False)
    vat_percent = forms.DecimalField(
        label="IVA %",
        min_value=Decimal("0.00"),
        max_value=Decimal("100.00"),
        decimal_places=2,
        required=False,
        initial=Decimal("0.00"),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["product"].queryset = _products_with_last_cost_queryset()
        self.fields["product"].empty_label = ""
        self.fields["product"].label_from_instance = _product_label_with_last_cost


class SaleHeaderForm(forms.Form):
    warehouse = forms.ModelChoiceField(queryset=Warehouse.objects.all())
    sale_date = forms.DateField(
        label="Fecha",
        required=False,
        input_formats=["%Y-%m-%d", "%d/%m/%Y"],
        widget=forms.DateInput(format="%Y-%m-%d", attrs={"type": "date"}),
    )
    delivery_status = forms.ChoiceField(
        choices=Sale.DeliveryStatus.choices,
        label="Estado de entrega",
        required=False,
        initial=Sale.DeliveryStatus.NOT_DELIVERED,
    )
    audiencia = forms.ChoiceField(
        choices=Customer.Audience.choices,
        label="Tipo de venta",
        initial=Customer.Audience.CONSUMER,
        required=False,
    )
    cliente = forms.ModelChoiceField(queryset=Customer.objects.all(), required=False)
    total_venta = forms.DecimalField(
        label="Total de venta",
        min_value=Decimal("0.00"),
        decimal_places=2,
        required=False,
    )
    descuento_total = forms.DecimalField(
        label="DTO %",
        min_value=Decimal("0.00"),
        decimal_places=2,
        required=False,
        initial=Decimal("0.00"),
    )
    comision_ml = forms.DecimalField(
        label="Comisión ML",
        min_value=Decimal("0.00"),
        decimal_places=2,
        required=False,
    )
    impuestos_ml = forms.DecimalField(
        label="Impuestos ML",
        min_value=Decimal("0.00"),
        decimal_places=2,
        required=False,
    )


class SaleItemForm(forms.Form):
    product = forms.ModelChoiceField(queryset=Product.objects.all())
    quantity = forms.IntegerField(min_value=1)
    discount_percent = forms.DecimalField(
        label="DTO %",
        min_value=Decimal("0.00"),
        max_value=Decimal("100.00"),
        decimal_places=2,
        required=False,
    )
    vat_percent = forms.DecimalField(
        label="IVA %",
        min_value=Decimal("0.00"),
        max_value=Decimal("100.00"),
        decimal_places=2,
        required=False,
        initial=Decimal("0.00"),
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["product"].queryset = _products_with_last_cost_queryset()
        self.fields["product"].empty_label = ""
        self.fields["product"].label_from_instance = _product_label_with_last_cost


class StockTransferForm(forms.Form):
    product = forms.ModelChoiceField(queryset=Product.objects.all(), label="Producto")
    quantity = forms.DecimalField(min_value=Decimal("0.01"), decimal_places=2, label="Cantidad")


class SupplierForm(forms.ModelForm):
    class Meta:
        model = Supplier
        fields = ["name", "phone"]
        labels = {"name": "Nombre", "phone": "Teléfono"}


class SupplierProductForm(forms.Form):
    supplier = forms.ModelChoiceField(queryset=Supplier.objects.all(), label="Proveedor")
    product = forms.ModelChoiceField(queryset=Product.objects.all(), label="Producto")
    last_cost = forms.DecimalField(
        label="Último costo",
        min_value=Decimal("0.00"),
        decimal_places=2,
        required=False,
    )


class ProductVariantForm(forms.Form):
    name = forms.CharField(label="Variedad")
    quantity = forms.DecimalField(min_value=Decimal("0.00"), decimal_places=2, label="Stock")


class ProductVariantRowForm(forms.Form):
    variant_id = forms.IntegerField(widget=forms.HiddenInput)
    name = forms.CharField(label="Variedad")
    quantity = forms.DecimalField(min_value=Decimal("0.00"), decimal_places=2, label="Stock")
    delete = forms.BooleanField(required=False, label="Eliminar")


class ProductCostRowForm(forms.Form):
    product_id = forms.IntegerField(widget=forms.HiddenInput)
    name = forms.CharField(required=True, label="Producto")
    group = forms.CharField(required=False, label="Marca / Grupo")
    supplier = forms.ModelChoiceField(queryset=Supplier.objects.all(), required=False, label="Proveedor")
    avg_cost = forms.DecimalField(min_value=Decimal("0.00"), decimal_places=2, label="Costo")
    vat_percent = forms.DecimalField(
        min_value=Decimal("0.00"),
        max_value=Decimal("100.00"),
        decimal_places=2,
        required=False,
        label="IVA %",
    )


class ProductBulkUpdateForm(forms.Form):
    group = forms.CharField(required=False, label="Marca / Grupo")
    supplier = forms.ModelChoiceField(queryset=Supplier.objects.all(), required=False, label="Proveedor")
    cost_percent = forms.DecimalField(
        required=False,
        decimal_places=2,
        label="Ajuste % costo",
        help_text="Ej: 10 para subir 10%, -5 para bajar 5%",
    )
    margin_consumer = forms.DecimalField(
        required=False,
        decimal_places=2,
        label="Margen % consumidor final",
    )
    margin_barber = forms.DecimalField(
        required=False,
        decimal_places=2,
        label="Margen % peluquerías/barberías",
    )
    margin_distributor = forms.DecimalField(
        required=False,
        decimal_places=2,
        label="Margen % distribuidores",
    )


@login_required
def dashboard(request):
    start_date = request.GET.get("start_date") or ""
    end_date = request.GET.get("end_date") or ""
    start_dt = None
    end_dt = None
    start_date_obj = None
    end_date_obj = None
    if start_date:
        try:
            parsed = datetime.strptime(start_date, "%Y-%m-%d")
            start_dt = timezone.make_aware(datetime.combine(parsed, time.min))
            start_date_obj = parsed.date()
        except ValueError:
            start_dt = None
    if end_date:
        try:
            parsed = datetime.strptime(end_date, "%Y-%m-%d")
            end_dt = timezone.make_aware(datetime.combine(parsed, time.max))
            end_date_obj = parsed.date()
        except ValueError:
            end_dt = None

    purchase_qs = Purchase.objects.all()
    sale_item_qs = SaleItem.objects.all()
    sales_qs = Sale.objects.select_related("warehouse").prefetch_related("items__product", "items__variant")
    tax_qs = TaxExpense.objects.all()
    if start_dt:
        purchase_qs = purchase_qs.filter(created_at__gte=start_dt)
        sale_item_qs = sale_item_qs.filter(sale__created_at__gte=start_dt)
        sales_qs = sales_qs.filter(created_at__gte=start_dt)
    if start_date_obj:
        tax_qs = tax_qs.filter(paid_at__gte=start_date_obj)
    if end_dt:
        purchase_qs = purchase_qs.filter(created_at__lte=end_dt)
        sale_item_qs = sale_item_qs.filter(sale__created_at__lte=end_dt)
        sales_qs = sales_qs.filter(created_at__lte=end_dt)
    if end_date_obj:
        tax_qs = tax_qs.filter(paid_at__lte=end_date_obj)

    purchase_total = purchase_qs.aggregate(total=Sum("total")).get("total") or Decimal("0.00")
    sale_total = sale_item_qs.aggregate(total=Sum("line_total")).get("total") or Decimal("0.00")
    tax_total = tax_qs.aggregate(total=Sum("amount")).get("total") or Decimal("0.00")

    margin_ml = Decimal("0.00")
    margin_comun = Decimal("0.00")
    for sale in sales_qs:
        cost_total = Decimal("0.00")
        for item in sale.items.all():
            cost_unit = item.cost_unit
            if cost_unit is None or cost_unit <= 0:
                cost_unit = item.product.cost_with_vat()
            cost_total += item.quantity * cost_unit
        if sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
            net_total = (sale.total or Decimal("0.00")) - (sale.ml_commission_total or Decimal("0.00")) - (
                sale.ml_tax_total or Decimal("0.00")
            )
            margin_ml += net_total - cost_total
        else:
            margin_comun += (sale.total or Decimal("0.00")) - cost_total

    net_margin = (margin_ml + margin_comun) - tax_total

    ranking_map = {}
    for sale in sales_qs:
        items = list(sale.items.all())
        if not items:
            continue
        items_total = sum((item.line_total or Decimal("0.00")) for item in items)
        if items_total <= 0:
            continue
        if sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
            revenue_total = (sale.total or Decimal("0.00")) - (sale.ml_commission_total or Decimal("0.00")) - (
                sale.ml_tax_total or Decimal("0.00")
            )
        else:
            revenue_total = sale.total or Decimal("0.00")
        for item in items:
            line_total = item.line_total or Decimal("0.00")
            revenue_share = (revenue_total * line_total / items_total) if items_total else Decimal("0.00")
            cost_unit = item.cost_unit if item.cost_unit and item.cost_unit > 0 else item.product.cost_with_vat()
            cost_total = item.quantity * cost_unit
            profit = (revenue_share - cost_total).quantize(Decimal("0.01"))
            key = (item.product_id, item.variant_id)
            if key not in ranking_map:
                ranking_map[key] = {
                    "product_id": item.product_id,
                    "sku": item.product.sku,
                    "name": item.product.name,
                    "variant": item.variant.name if item.variant else None,
                    "quantity": Decimal("0.00"),
                    "profit": Decimal("0.00"),
                }
            ranking_map[key]["quantity"] += item.quantity
            ranking_map[key]["profit"] += profit

    ranking = sorted(ranking_map.values(), key=lambda item: item["profit"], reverse=True)

    context = {
        "purchase_total": purchase_total,
        "sale_total": sale_total,
        "gross_margin": net_margin,
        "gross_margin_pct": (net_margin / sale_total * Decimal("100.00")) if sale_total else None,
        "margin_ml": margin_ml,
        "margin_comun": margin_comun,
        "ranking": ranking,
        "start_date": start_date,
        "end_date": end_date,
        "tax_total": tax_total,
    }
    return render(request, "inventory/dashboard.html", context)


@login_required
def create_product(request):
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save(commit=False)
            if not product.sku:
                prefix = _sku_prefix(product.group or "", product.name or "")
                existing = Product.objects.filter(sku__startswith=prefix).values_list("sku", flat=True)
                max_suffix = 0
                for sku in existing:
                    suffix = sku[len(prefix):]
                    if suffix.isdigit():
                        max_suffix = max(max_suffix, int(suffix))
                product.sku = f"{prefix}{max_suffix + 1:04d}"
            product.save()
            messages.success(request, "Producto creado.")
            return redirect("inventory_dashboard")
    else:
        form = ProductForm()
    return render(request, "inventory/product_form.html", {"form": form, "title": "Nuevo producto"})


@login_required
def edit_product(request, pk: int):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            product = form.save(commit=False)
            if not product.sku:
                prefix = _sku_prefix(product.group or "", product.name or "")
                existing = Product.objects.filter(sku__startswith=prefix).values_list("sku", flat=True)
                max_suffix = 0
                for sku in existing:
                    suffix = sku[len(prefix):]
                    if suffix.isdigit():
                        max_suffix = max(max_suffix, int(suffix))
                product.sku = f"{prefix}{max_suffix + 1:04d}"
            product.save()
            messages.success(request, "Producto actualizado.")
            return redirect("inventory_product_prices")
        messages.error(request, "Revisá los datos del producto.")
    else:
        form = ProductForm(instance=product)
    return render(request, "inventory/product_form.html", {"form": form, "title": "Editar producto"})


@login_required
def register_purchase(request):
    return purchases_list(request)


@login_required
def register_sale(request):
    return sales_list(request)


@login_required
def sale_receipt(request, sale_id: int):
    sale = get_object_or_404(
        Sale.objects.select_related("customer", "warehouse").prefetch_related("items__product", "items__variant"),
        pk=sale_id,
    )
    items = list(sale.items.all())
    subtotal = sum((item.unit_price * item.quantity for item in items), Decimal("0.00"))
    discount_total = sale.discount_total or Decimal("0.00")
    per_item_discount = sum(((item.unit_price * item.quantity) - item.line_total for item in items), Decimal("0.00"))
    extra_discount_amount = discount_total - per_item_discount
    if extra_discount_amount < 0:
        extra_discount_amount = Decimal("0.00")
    extra_discount_percent = (
        (extra_discount_amount / subtotal * Decimal("100.00")).quantize(Decimal("0.01"))
        if subtotal > 0
        else Decimal("0.00")
    )
    is_ml = sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE
    subtotal_display = subtotal
    total_display = sale.total if is_ml and sale.total is not None else (subtotal - discount_total)
    commission_total = sale.ml_commission_total or Decimal("0.00")
    tax_total = sale.ml_tax_total or Decimal("0.00")
    net_total = (total_display - commission_total - tax_total) if is_ml else None
    invoice_number = sale.ml_order_id or sale.invoice_number
    context = {
        "sale": sale,
        "items": items,
        "subtotal": subtotal_display,
        "discount_total": discount_total,
        "extra_discount_percent": extra_discount_percent,
        "total": total_display,
        "is_ml": is_ml,
        "commission_total": commission_total,
        "tax_total": tax_total,
        "net_total": net_total,
        "invoice_number": invoice_number,
    }
    return render(request, "inventory/sale_receipt.html", context)


@login_required
def sale_receipt_pdf(request, sale_id: int):
    sale = get_object_or_404(
        Sale.objects.select_related("customer", "warehouse").prefetch_related("items__product", "items__variant"),
        pk=sale_id,
    )
    items = list(sale.items.all())
    subtotal = sum((item.unit_price * item.quantity for item in items), Decimal("0.00"))
    discount_total = sale.discount_total or Decimal("0.00")
    per_item_discount = sum(((item.unit_price * item.quantity) - item.line_total for item in items), Decimal("0.00"))
    extra_discount_amount = discount_total - per_item_discount
    if extra_discount_amount < 0:
        extra_discount_amount = Decimal("0.00")
    extra_discount_percent = (
        (extra_discount_amount / subtotal * Decimal("100.00")).quantize(Decimal("0.01"))
        if subtotal > 0
        else Decimal("0.00")
    )
    is_ml = sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE
    subtotal_display = subtotal
    total_display = sale.total if is_ml and sale.total is not None else (subtotal - discount_total)
    commission_total = sale.ml_commission_total or Decimal("0.00")
    tax_total = sale.ml_tax_total or Decimal("0.00")
    net_total = (total_display - commission_total - tax_total) if is_ml else None
    from django.template.loader import render_to_string

    html = render_to_string(
        "inventory/sale_receipt_pdf.html",
        {
            "sale": sale,
            "items": items,
            "subtotal": subtotal_display,
            "discount_total": discount_total,
            "extra_discount_percent": extra_discount_percent,
            "total": total_display,
            "is_ml": is_ml,
            "commission_total": commission_total,
            "tax_total": tax_total,
            "net_total": net_total,
            "invoice_number": sale.ml_order_id or sale.invoice_number,
        },
        request=request,
    )
    try:
        from weasyprint import HTML
    except Exception:
        return HttpResponse(
            "WeasyPrint no está instalado. Instalalo con 'pip install weasyprint' para generar PDF.",
            status=500,
        )

    pdf_bytes = HTML(string=html, base_url=request.build_absolute_uri("/")).write_pdf()
    invoice_raw = sale.ml_order_id or sale.invoice_number
    if isinstance(invoice_raw, str) and "-" in invoice_raw:
        invoice_raw = invoice_raw.split("-", 1)[1]
    invoice_number = str(invoice_raw).lstrip("0") or str(sale.id)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="FACT-{invoice_number}.pdf"'
    return response


@login_required
def sale_edit(request, sale_id: int):
    sale = get_object_or_404(
        Sale.objects.select_related("customer", "warehouse").prefetch_related("items__product", "items__variant", "movements"),
        pk=sale_id,
    )
    SaleItemFormSet = formset_factory(SaleItemForm, extra=0, can_delete=True)
    customer_audiences = {str(customer.id): customer.audience for customer in Customer.objects.only("id", "audience")}

    if request.method == "POST":
        post_data = request.POST.copy()
        for key in list(post_data.keys()):
            if not key.endswith("-product_text"):
                continue
            prefix = key[:-len("product_text")]
            product_key = f"{prefix}product"
            if post_data.get(product_key):
                continue
            product_text = (post_data.get(key) or "").strip()
            if not product_text:
                continue
            label = product_text.split(" (", 1)[0].strip()
            sku_candidate = ""
            name_candidate = label
            if " - " in label:
                sku_candidate, name_candidate = [part.strip() for part in label.split(" - ", 1)]
            product = None
            if sku_candidate and sku_candidate.lower() != "sin sku":
                product = Product.objects.filter(sku__iexact=sku_candidate).first()
            if not product and name_candidate:
                product = (
                    Product.objects.filter(name__iexact=name_candidate).first()
                    or Product.objects.filter(name__icontains=name_candidate).first()
                )
            if product:
                post_data[product_key] = str(product.id)
        header_form = SaleHeaderForm(post_data)
        formset = SaleItemFormSet(post_data)
        for form in formset.forms:
            prefix = form.prefix
            product_raw = (post_data.get(f"{prefix}-product") or "").strip()
            product_text = (post_data.get(f"{prefix}-product_text") or "").strip()
            qty_raw = (post_data.get(f"{prefix}-quantity") or "").strip()
            qty_value = None
            if qty_raw:
                try:
                    qty_value = Decimal(str(qty_raw).replace(",", "."))
                except Exception:
                    qty_value = None
            qty_zeroish = qty_raw in {"", "0", "0.0", "0.00", "0,00"} or (qty_value is not None and qty_value <= 0)
            if not product_raw and not product_text and qty_zeroish:
                form.empty_permitted = True
            elif not form.has_changed():
                form.empty_permitted = True
        if header_form.is_valid() and formset.is_valid():
            warehouse = header_form.cleaned_data["warehouse"]
            require_variants = warehouse.type == Warehouse.WarehouseType.COMUN
            items = []
            for form in formset:
                if form.cleaned_data.get("DELETE"):
                    continue
                product = form.cleaned_data.get("product")
                if not product:
                    continue
                variant_id_raw = (post_data.get(f"{form.prefix}-variant_id") or "").strip()
                variant = None
                if variant_id_raw:
                    variant = ProductVariant.objects.filter(id=variant_id_raw, product=product).first()
                    if not variant:
                        form.add_error(None, "Variedad inválida.")
                if require_variants and ProductVariant.objects.filter(product=product).exists() and not variant:
                    form.add_error(None, "Seleccioná una variedad.")
                items.append({**form.cleaned_data, "variant": variant})
            if any(form.errors for form in formset):
                messages.error(request, "Revisá los campos de la venta.")
                return redirect("inventory_sale_edit", sale_id=sale.id)
            if not items:
                messages.error(request, "Agregá al menos un producto.")
                return redirect("inventory_sale_edit", sale_id=sale.id)
            comun_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.COMUN).first()
            audience = header_form.cleaned_data.get("audiencia") or Customer.Audience.CONSUMER
            customer = header_form.cleaned_data.get("cliente")
            total_venta = header_form.cleaned_data.get("total_venta")
            sale_date = header_form.cleaned_data.get("sale_date")
            extra_discount_percent = header_form.cleaned_data.get("descuento_total") or Decimal("0.00")
            comision_ml = header_form.cleaned_data.get("comision_ml") or Decimal("0.00")
            impuestos_ml = header_form.cleaned_data.get("impuestos_ml") or Decimal("0.00")
            delivery_status = header_form.cleaned_data.get("delivery_status") or Sale.DeliveryStatus.NOT_DELIVERED
            if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
                customer = None
                audience = Customer.Audience.CONSUMER
                delivery_status = Sale.DeliveryStatus.NOT_DELIVERED
            elif customer:
                audience = customer.audience
            try:
                with transaction.atomic():
                    is_ml_sale = (
                        sale.ml_order_id
                        or sale.reference.startswith("ML ORDER")
                        or sale.reference.startswith("GS ORDER")
                        or warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE
                    )
                    previous_items = list(sale.items.select_related("variant", "product"))
                    if sale.warehouse.type == Warehouse.WarehouseType.COMUN:
                        for prev_item in previous_items:
                            if prev_item.variant_id:
                                variant = (
                                    ProductVariant.objects.select_for_update()
                                    .filter(id=prev_item.variant_id, product=prev_item.product)
                                    .first()
                                )
                                if variant:
                                    variant.quantity = (variant.quantity + prev_item.quantity).quantize(Decimal("0.01"))
                                    variant.save(update_fields=["quantity"])
                                    if comun_wh:
                                        _koda_sync_common_with_variants(prev_item.product, comun_wh)
                    for movement in sale.movements.select_for_update():
                        if not is_ml_sale and movement.movement_type == StockMovement.MovementType.EXIT and movement.from_warehouse:
                            stock, _ = Stock.objects.select_for_update().get_or_create(
                                product=movement.product,
                                warehouse=movement.from_warehouse,
                                defaults={"quantity": Decimal("0.00")},
                            )
                            stock.quantity = (stock.quantity + movement.quantity).quantize(Decimal("0.01"))
                            stock.save(update_fields=["quantity"])
                        movement.delete()
                    sale.items.all().delete()

                    sale.customer = customer
                    sale.warehouse = warehouse
                    sale.audience = audience
                    sale.delivery_status = delivery_status
                    sale.ml_commission_total = (
                        comision_ml if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE else Decimal("0.00")
                    )
                    sale.ml_tax_total = (
                        impuestos_ml if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE else Decimal("0.00")
                    )
                    update_fields = [
                        "customer",
                        "warehouse",
                        "audience",
                        "delivery_status",
                        "ml_commission_total",
                        "ml_tax_total",
                    ]
                    if sale_date:
                        created_at = datetime.combine(sale_date, time(12, 0))
                        if timezone.is_naive(created_at):
                            created_at = timezone.make_aware(created_at)
                        sale.created_at = created_at
                        update_fields.append("created_at")
                    sale.save(update_fields=update_fields)

                    total = Decimal("0.00")
                    discount_total = Decimal("0.00")
                    base_subtotal = Decimal("0.00")
                    for data in items:
                        base_price = {
                            Customer.Audience.CONSUMER: data["product"].consumer_price,
                            Customer.Audience.BARBER: data["product"].barber_price,
                            Customer.Audience.DISTRIBUTOR: data["product"].distributor_price,
                        }.get(audience, data["product"].consumer_price)
                        discount = data.get("discount_percent")
                        if discount is None:
                        discount = data.get("discount_percent")
                        if discount is None:
                            discount = Decimal("0.00")
                            if customer:
                                discount_obj = CustomerProductDiscount.objects.filter(
                                    customer=customer, product=data["product"]
                                ).first()
                                if discount_obj:
                                    discount = discount_obj.discount_percent
                                else:
                                    group_key = (data["product"].group or "").strip()
                                    if group_key:
                                        group_discount = CustomerGroupDiscount.objects.filter(
                                            customer=customer, group__iexact=group_key
                                        ).first()
                                        if group_discount:
                                            discount = group_discount.discount_percent
                        final_price = base_price * (Decimal("1.00") - discount / Decimal("100.00"))
                        qty = Decimal(data["quantity"])
                        line_total = (qty * final_price).quantize(Decimal("0.01"))
                        discount_amount = (qty * (base_price - final_price)).quantize(Decimal("0.01"))
                        base_subtotal += (qty * base_price).quantize(Decimal("0.01"))
                        variant = data.get("variant")
                        if warehouse.type == Warehouse.WarehouseType.COMUN and variant:
                            variant = (
                                ProductVariant.objects.select_for_update()
                                .filter(id=variant.id, product=data["product"])
                                .first()
                            )
                            if variant:
                                if variant.quantity - qty < 0:
                                    raise services.NegativeStockError("Stock insuficiente en variedad.")
                                variant.quantity = (variant.quantity - qty).quantize(Decimal("0.01"))
                                variant.save(update_fields=["quantity"])
                                if comun_wh:
                                    _koda_sync_common_with_variants(data["product"], comun_wh)
                        SaleItem.objects.create(
                            sale=sale,
                            product=data["product"],
                            variant=variant,
                            quantity=qty,
                            unit_price=base_price,
                            cost_unit=data["product"].last_purchase_cost(),
                            discount_percent=discount,
                            final_unit_price=final_price,
                            line_total=line_total,
                            vat_percent=data.get("vat_percent") or Decimal("0.00"),
                        )
                        total += line_total
                        discount_total += discount_amount
                        if warehouse.type != Warehouse.WarehouseType.MERCADOLIBRE:
                            services.register_exit(
                                product=data["product"],
                                warehouse=warehouse,
                                quantity=data["quantity"],
                                user=request.user,
                                reference=f"Venta {audience} #{sale.id}",
                                sale_price=final_price,
                                vat_percent=data.get("vat_percent") or Decimal("0.00"),
                                sale=sale,
                            )
                    gross_total = (
                        total_venta
                        if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE and total_venta is not None
                        else total
                    )
                    discount_base = total_venta if total_venta is not None else base_subtotal
                    extra_discount_amount = (discount_base * extra_discount_percent / Decimal("100.00")).quantize(Decimal("0.01"))
                    sale.total = (gross_total - extra_discount_amount).quantize(Decimal("0.01"))
                    sale.discount_total = (discount_total + extra_discount_amount).quantize(Decimal("0.01"))
                    sale.save(update_fields=["total", "discount_total"])
                messages.success(request, "Venta actualizada.")
                if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
                    return redirect("inventory_sales_list")
                return redirect("inventory_sale_receipt", sale_id=sale.id)
            except services.NegativeStockError:
                messages.error(request, "No se puede actualizar: el stock quedaría negativo.")
            except services.InvalidMovementError as exc:
                messages.error(request, str(exc))
            except Exception as exc:
                messages.error(request, f"No se pudo actualizar la venta: {exc}")
        else:
            messages.error(request, "Revisá los campos de la venta.")
        return redirect("inventory_sale_edit", sale_id=sale.id)
    else:
        items_for_discount = list(sale.items.all())
        subtotal_base = sum((item.unit_price * item.quantity for item in items_for_discount), Decimal("0.00"))
        per_item_discount = sum(
            ((item.unit_price * item.quantity) - item.line_total for item in items_for_discount),
            Decimal("0.00"),
        )
        extra_discount_amount = (sale.discount_total or Decimal("0.00")) - per_item_discount
        if extra_discount_amount < 0:
            extra_discount_amount = Decimal("0.00")
        discount_percent = (
            (extra_discount_amount / subtotal_base * Decimal("100.00")).quantize(Decimal("0.01"))
            if subtotal_base > 0
            else Decimal("0.00")
        )
        header_form = SaleHeaderForm(
            initial={
                "warehouse": sale.warehouse,
                "sale_date": timezone.localtime(sale.created_at).date() if sale.created_at else None,
                "audiencia": sale.audience,
                "cliente": sale.customer,
                "delivery_status": sale.delivery_status,
                "total_venta": sale.total if sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE else None,
                "descuento_total": discount_percent,
                "comision_ml": sale.ml_commission_total or Decimal("0.00"),
                "impuestos_ml": sale.ml_tax_total or Decimal("0.00"),
            }
        )
        items = list(sale.items.select_related("variant", "product"))
        initial = [
            {
                "product": item.product,
                "quantity": int(item.quantity),
                "discount_percent": item.discount_percent,
                "vat_percent": item.vat_percent,
            }
            for item in items
        ]
        formset = SaleItemFormSet(initial=initial)
        item_rows = [
            {"form": form, "variant_id": item.variant_id}
            for form, item in zip(formset.forms, items)
        ]

    variant_data = {}
    for row in ProductVariant.objects.values("id", "product_id", "name").order_by("name", "id"):
        variant_data.setdefault(str(row["product_id"]), []).append({"id": row["id"], "name": row["name"]})

    return render(
        request,
        "inventory/sale_edit.html",
        {
            "sale": sale,
            "form": header_form,
            "formset": formset,
            "item_rows": item_rows if request.method != "POST" else None,
            "customer_audiences": customer_audiences,
            "variant_data": variant_data,
        },
    )


@login_required
def purchase_receipt(request, purchase_id: int):
    purchase = get_object_or_404(
        Purchase.objects.select_related("supplier", "warehouse").prefetch_related("items__product"), pk=purchase_id
    )
    items = list(purchase.items.all())
    has_vat = False
    subtotal_no_vat = Decimal("0.00")
    vat_total = Decimal("0.00")
    for item in items:
        discount_percent = item.discount_percent or Decimal("0.00")
        effective_unit_cost = (item.unit_cost * (Decimal("1.00") - (discount_percent / Decimal("100.00")))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        item.discount_percent = discount_percent
        item.unit_cost_effective = effective_unit_cost
        item.line_total = (item.quantity * effective_unit_cost).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        product_vat = item.product.vat_percent or Decimal("0.00")
        item_vat = item.vat_percent or Decimal("0.00")
        vat_percent = item_vat if item_vat > 0 else product_vat
        if vat_percent > 0:
            has_vat = True
            vat_factor = Decimal("1.00") + (vat_percent / Decimal("100.00"))
            item.unit_cost_no_vat = (effective_unit_cost / vat_factor).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
            item.unit_vat = (effective_unit_cost - item.unit_cost_no_vat).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            item.unit_cost_no_vat = effective_unit_cost
            item.unit_vat = Decimal("0.00")
        item.vat_percent = vat_percent
        item.line_vat = (item.unit_vat * item.quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        subtotal_no_vat += (item.unit_cost_no_vat * item.quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        vat_total += item.line_vat
    subtotal = sum((item.line_total for item in items), Decimal("0.00"))
    discount_percent = purchase.discount_percent or Decimal("0.00")
    discount_total = (subtotal * discount_percent / Decimal("100.00")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    total = (subtotal - discount_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if purchase.total != total:
        purchase.total = total
        purchase.save(update_fields=["total"])
    context = {
        "purchase": purchase,
        "items": items,
        "subtotal": subtotal,
        "subtotal_no_vat": subtotal_no_vat,
        "vat_total": vat_total,
        "discount_total": discount_total,
        "discount_percent": discount_percent,
        "total": total,
        "invoice_number": purchase.invoice_number,
        "has_vat": has_vat,
    }
    return render(request, "inventory/purchase_receipt.html", context)


@login_required
def sales_list(request):
    SaleItemFormSet = formset_factory(SaleItemForm, extra=1, can_delete=False)
    default_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.COMUN).first()
    if not default_wh:
        default_wh = Warehouse.objects.first()
    customer_audiences = {
        str(customer.id): customer.audience
        for customer in Customer.objects.only("id", "audience")
    }
    search_query = (request.GET.get("q") or "").strip()
    start_date_raw = (request.GET.get("start_date") or "").strip()
    end_date_raw = (request.GET.get("end_date") or "").strip()
    include_comun = request.GET.get("wh_comun") == "1"
    include_ml = request.GET.get("wh_ml") == "1"
    show_history = request.GET.get("show_history") == "1"
    customers = Customer.objects.order_by("name")
    action = request.POST.get("action") if request.method == "POST" else ""
    if action == "import_ml_sales_xlsx":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Subí el archivo XLSX con las ventas.")
            return redirect("inventory_sales_list")
        rows, error = _read_ml_sales_xlsx_rows(upload)
        if error:
            messages.error(request, error)
            return redirect("inventory_sales_list")
        if not rows:
            messages.error(request, "El archivo no tiene filas válidas.")
            return redirect("inventory_sales_list")

        def parse_datetime(value: object) -> datetime | None:
            if value is None or value == "":
                return None
            if isinstance(value, datetime):
                parsed = value
            elif hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
                parsed = datetime(value.year, value.month, value.day)
            else:
                raw = str(value).strip()
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
                    try:
                        parsed = datetime.strptime(raw, fmt)
                        break
                    except ValueError:
                        parsed = None
                if parsed is None:
                    return None
            if timezone.is_naive(parsed):
                return timezone.make_aware(parsed)
            return parsed

        ml_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.MERCADOLIBRE).first()
        if not ml_wh:
            messages.error(request, "Falta el depósito MercadoLibre.")
            return redirect("inventory_sales_list")

        ml_items = list(MercadoLibreItem.objects.select_related("product"))
        ml_title_index = {}
        for ml_item in ml_items:
            key = ml._normalize(ml_item.title or "")
            if key:
                ml_title_index.setdefault(key, []).append(ml_item)

        created_sales = 0
        skipped = 0
        unmatched = 0
        unmatched_refs: list[str] = []
        for idx, row in enumerate(rows, start=1):
            title = row["title"]
            qty = row["quantity"]
            price_total = row["price_total"]
            commission = row["commission"]
            taxes = row["taxes"]
            created_at = parse_datetime(row.get("created_at"))
            order_id = (row.get("order_id") or "").strip()

            product = None
            title_norm = ml._normalize(title)
            if title_norm in ml_title_index:
                product = ml_title_index[title_norm][0].product
            if not product and title_norm:
                for ml_item in ml_items:
                    item_norm = ml._normalize(ml_item.title or "")
                    if not item_norm:
                        continue
                    if title_norm in item_norm or item_norm in title_norm:
                        product = ml_item.product
                        if product:
                            break
            if not product:
                unmatched += 1
                ref_label = order_id or f"fila {idx}"
                unmatched_refs.append(f"{ref_label}: {title}")
                product = (
                    Product.objects.filter(name=title, group="MercadoLibre (sin match)").first()
                    or Product.objects.create(
                        name=title,
                        group="MercadoLibre (sin match)",
                        avg_cost=Decimal("0.00"),
                        vat_percent=Decimal("0.00"),
                    )
                )

            if order_id:
                reference = f"XLSX ML {order_id}"
                if Sale.objects.filter(ml_order_id=order_id).exists():
                    skipped += 1
                    continue
            else:
                reference = f"XLSX ML {created_at.date() if created_at else 'SIN-FECHA'} #{idx}"
                if Sale.objects.filter(reference=reference).exists():
                    skipped += 1
                    continue

            with transaction.atomic():
                sale = Sale.objects.create(
                    warehouse=ml_wh,
                    audience=Customer.Audience.CONSUMER,
                    total=price_total.quantize(Decimal("0.01")),
                    reference=reference,
                    ml_order_id=order_id,
                    ml_commission_total=commission.quantize(Decimal("0.01")),
                    ml_tax_total=taxes.quantize(Decimal("0.01")),
                    user=request.user,
                )
                if created_at:
                    Sale.objects.filter(pk=sale.pk).update(created_at=created_at)
                unit_price = (price_total / qty).quantize(Decimal("0.01"))
                line_total = (unit_price * qty).quantize(Decimal("0.01"))
                SaleItem.objects.create(
                    sale=sale,
                    product=product,
                    quantity=qty,
                    unit_price=unit_price,
                    cost_unit=product.last_purchase_cost(),
                    discount_percent=Decimal("0.00"),
                    final_unit_price=unit_price,
                    line_total=line_total,
                    vat_percent=product.vat_percent or Decimal("0.00"),
                )
                created_sales += 1

        messages.success(
            request,
            f"Ventas importadas: {created_sales}, omitidas: {skipped}, sin match: {unmatched}.",
        )
        if unmatched_refs:
            preview = ", ".join(unmatched_refs[:10])
            more = "" if len(unmatched_refs) <= 10 else f" (+{len(unmatched_refs) - 10} más)"
            messages.warning(
                request,
                "Sin match en comprobantes: "
                f"{preview}{more}. Se registraron con el nombre del XLSX. Editá la venta para asignar el producto.",
            )
        return redirect("inventory_sales_list")
    if action in {"sync_google_sales", "reset_google_sales"}:
        if action == "reset_google_sales":
            ml_sales = Sale.objects.filter(
                Q(reference__startswith="ML ORDER ")
                | Q(reference__startswith="GS ORDER ")
                | Q(ml_order_id__gt="")
            )
            StockMovement.objects.filter(sale__in=ml_sales).delete()
            deleted_count = ml_sales.count()
            ml_sales.delete()
            messages.info(request, f"Ventas ML eliminadas: {deleted_count}.")
        sheet_url = (request.POST.get("sheet_url") or "").strip()
        if not sheet_url:
            sheet_url = os.environ.get("GOOGLE_SHEETS_SALES_URL", "")
        if not sheet_url:
            messages.error(request, "Falta GOOGLE_SHEETS_SALES_URL en el entorno.")
            return redirect("inventory_sales_list")

        def normalize_header(value: str) -> str:
            value = (value or "").strip().lower()
            value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
            return re.sub(r"\s+", "", value)

        def parse_decimal(value: str) -> Decimal:
            raw = (value or "").strip()
            raw = raw.replace("$", "").replace(" ", "")
            if not raw:
                return Decimal("0.00")
            if "," in raw and "." in raw:
                raw = raw.replace(".", "").replace(",", ".")
            elif "," in raw:
                raw = raw.replace(",", ".")
            try:
                return Decimal(raw)
            except Exception:
                return Decimal("0.00")

        def parse_datetime(value: str | None) -> datetime | None:
            if not value:
                return None
            raw = value.strip()
            match = re.match(r"^(.*)([+-]\d{2})(\d{2})$", raw)
            if match:
                raw = f"{match.group(1)}{match.group(2)}:{match.group(3)}"
            raw = raw.replace("Z", "+00:00")
            try:
                parsed = datetime.fromisoformat(raw)
            except ValueError:
                return None
            if timezone.is_naive(parsed):
                return timezone.make_aware(parsed)
            return parsed

        match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
        gid_match = re.search(r"[?&]gid=(\d+)", sheet_url)
        if not match or not gid_match:
            messages.error(request, "URL de Google Sheets inválida (falta ID o gid).")
            return redirect("inventory_sales_list")

        sheet_id = match.group(1)
        gid = gid_match.group(1)
        csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

        try:
            with urlopen(csv_url, timeout=30) as resp:
                content = resp.read().decode("utf-8", errors="ignore")
        except Exception as exc:
            messages.error(request, f"No se pudo leer la hoja: {exc}")
            return redirect("inventory_sales_list")

        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            messages.error(request, "La hoja no tiene datos.")
            return redirect("inventory_sales_list")

        headers = [normalize_header(h) for h in rows[0]]

        def idx(*names: str) -> int | None:
            for name in names:
                key = normalize_header(name)
                if key in headers:
                    return headers.index(key)
            return None

        fecha_idx = idx("Fecha")
        idventa_idx = idx("IDVenta", "ID Vent", "IDVenta")
        idorder_idx = idx("IDOrder", "ID Order", "IDOrden", "ID Or", "IDor")
        producto_idx = idx("Producto")
        cantidad_idx = idx("Cantidad")
        precio_idx = idx("Precio Bruto Venta", "Precio Bruto", "Precio")
        comision_idx = idx("Comision", "Comisión")
        iibb_idx = idx("IIBB", "Impuestos", "Impuesto")

        required = [fecha_idx, idventa_idx, producto_idx, cantidad_idx, precio_idx, comision_idx, iibb_idx]
        if any(i is None for i in required):
            messages.error(
                request,
                "Faltan columnas requeridas en la hoja (Fecha, IDVenta, Producto, Cantidad, Precio Bruto Venta, Comision, Impuestos).",
            )
            return redirect("inventory_sales_list")

        ml_items = list(MercadoLibreItem.objects.select_related("product"))
        ml_title_index = {}
        for ml_item in ml_items:
            key = ml._normalize(ml_item.title or "")
            if key:
                ml_title_index.setdefault(key, []).append(ml_item)
        orders: dict[str, dict[str, object]] = {}
        unmatched = 0

        for row in rows[1:]:
            if not row:
                continue
            if len(row) < len(headers):
                row = row + [""] * (len(headers) - len(row))
            order_id = (row[idorder_idx] if idorder_idx is not None else "") or row[idventa_idx]
            order_id = str(order_id).strip()
            if not order_id:
                continue
            title = str(row[producto_idx]).strip()
            qty = parse_decimal(row[cantidad_idx])
            if qty <= 0:
                continue
            price_total = parse_decimal(row[precio_idx])
            commission = parse_decimal(row[comision_idx])
            iibb = parse_decimal(row[iibb_idx])
            created_at = parse_datetime(row[fecha_idx])

            product = None
            title_norm = ml._normalize(title)
            if title_norm in ml_title_index:
                product = ml_title_index[title_norm][0].product
            if not product and title_norm:
                for ml_item in ml_items:
                    item_norm = ml._normalize(ml_item.title or "")
                    if not item_norm:
                        continue
                    if title_norm in item_norm or item_norm in title_norm:
                        product = ml_item.product
                        if product:
                            break
            if not product:
                unmatched += 1
                continue

            entry = orders.setdefault(
                order_id,
                {
                    "created_at": created_at,
                    "total": Decimal("0.00"),
                    "commission": Decimal("0.00"),
                    "iibb": Decimal("0.00"),
                    "items": [],
                },
            )
            entry["total"] += price_total
            entry["commission"] += commission
            entry["iibb"] += iibb
            unit_price = (price_total / qty).quantize(Decimal("0.01"))
            entry["items"].append(
                {
                    "product": product,
                    "quantity": qty,
                    "unit_price": unit_price,
                }
            )
            if created_at and (entry["created_at"] is None or created_at < entry["created_at"]):
                entry["created_at"] = created_at

        created_sales = 0
        skipped = 0
        ml_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.MERCADOLIBRE).first()
        for order_id, data in orders.items():
            reference = f"GS ORDER {order_id}"
            if Sale.objects.filter(reference=reference).exists():
                skipped += 1
                continue
            if not data["items"]:
                continue
            if not ml_wh:
                messages.error(request, "Falta el depósito MercadoLibre.")
                return redirect("inventory_sales_list")
            with transaction.atomic():
                sale = Sale.objects.create(
                    warehouse=ml_wh,
                    audience=Customer.Audience.CONSUMER,
                    total=data["total"].quantize(Decimal("0.01")),
                    reference=reference,
                    ml_order_id=str(order_id),
                    ml_commission_total=data["commission"].quantize(Decimal("0.01")),
                    ml_tax_total=data["iibb"].quantize(Decimal("0.01")),
                    user=request.user,
                )
                if data["created_at"]:
                    Sale.objects.filter(pk=sale.pk).update(created_at=data["created_at"])
                for item in data["items"]:
                    qty = item["quantity"]
                    unit_price = item["unit_price"]
                    line_total = (unit_price * qty).quantize(Decimal("0.01"))
                    SaleItem.objects.create(
                        sale=sale,
                        product=item["product"],
                        quantity=qty,
                        unit_price=unit_price,
                        cost_unit=item["product"].last_purchase_cost(),
                        discount_percent=Decimal("0.00"),
                        final_unit_price=unit_price,
                        line_total=line_total,
                        vat_percent=item["product"].vat_percent or Decimal("0.00"),
                    )
                created_sales += 1

        messages.success(
            request,
            f"Ventas importadas: {created_sales}, omitidas: {skipped}, sin match: {unmatched}.",
        )
        return redirect("inventory_sales_list")
    if action == "bulk_delete_sales":
        start_raw = (request.POST.get("bulk_start") or "").strip()
        end_raw = (request.POST.get("bulk_end") or "").strip()
        only_ml = request.POST.get("bulk_only_ml") == "1"
        if not start_raw and not end_raw and not only_ml:
            messages.error(request, "Indicá un rango de fechas o seleccioná solo MercadoLibre.")
            return redirect("inventory_sales_list")

        sales_qs = Sale.objects.all()
        if only_ml:
            sales_qs = sales_qs.filter(
                Q(reference__startswith="ML ORDER ")
                | Q(reference__startswith="GS ORDER ")
                | Q(ml_order_id__gt="")
            )
        if start_raw:
            try:
                start_dt = datetime.fromisoformat(start_raw)
                if timezone.is_naive(start_dt):
                    start_dt = timezone.make_aware(start_dt)
                sales_qs = sales_qs.filter(created_at__gte=start_dt)
            except ValueError:
                messages.error(request, "Fecha desde inválida.")
                return redirect("inventory_sales_list")
        if end_raw:
            try:
                end_dt = datetime.fromisoformat(end_raw)
                if timezone.is_naive(end_dt):
                    end_dt = timezone.make_aware(end_dt)
                sales_qs = sales_qs.filter(created_at__lte=end_dt)
            except ValueError:
                messages.error(request, "Fecha hasta inválida.")
                return redirect("inventory_sales_list")

        deleted_ids = list(sales_qs.values_list("id", flat=True))
        StockMovement.objects.filter(sale_id__in=deleted_ids).delete()
        deleted = len(deleted_ids)
        Sale.objects.filter(id__in=deleted_ids).delete()
        messages.success(request, f"Ventas eliminadas: {deleted}.")
        return redirect("inventory_sales_list")
    if action == "bulk_delete_selected":
        ids = request.POST.getlist("sale_ids")
        if not ids:
            messages.error(request, "No seleccionaste ventas para eliminar.")
            return redirect("inventory_sales_list")
        sales = list(
            Sale.objects.filter(id__in=ids).prefetch_related("items", "movements")
        )
        deleted = 0
        with transaction.atomic():
            for sale in sales:
                is_ml_sale = sale.ml_order_id or sale.reference.startswith("ML ORDER") or sale.reference.startswith("GS ORDER")
                comun_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.COMUN).first()
                if sale.warehouse.type == Warehouse.WarehouseType.COMUN:
                    for item in sale.items.select_related("variant", "product"):
                        if item.variant_id:
                            variant = (
                                ProductVariant.objects.select_for_update()
                                .filter(id=item.variant_id, product=item.product)
                                .first()
                            )
                            if variant:
                                variant.quantity = (variant.quantity + item.quantity).quantize(Decimal("0.01"))
                                variant.save(update_fields=["quantity"])
                                if comun_wh:
                                    _koda_sync_common_with_variants(item.product, comun_wh)
                for movement in sale.movements.select_for_update():
                    if not is_ml_sale and movement.movement_type == StockMovement.MovementType.EXIT and movement.from_warehouse:
                        stock, _ = Stock.objects.select_for_update().get_or_create(
                            product=movement.product,
                            warehouse=movement.from_warehouse,
                            defaults={"quantity": Decimal("0.00")},
                        )
                        stock.quantity = (stock.quantity + movement.quantity).quantize(Decimal("0.01"))
                        stock.save(update_fields=["quantity"])
                    movement.delete()
                sale.delete()
                deleted += 1
        messages.success(request, f"Ventas eliminadas: {deleted}.")
        return redirect("inventory_sales_list")
    if request.method == "POST":
        post_data = request.POST.copy()
        for key in list(post_data.keys()):
            if not key.endswith("-product_text"):
                continue
            prefix = key[:-len("product_text")]
            product_key = f"{prefix}product"
            if post_data.get(product_key):
                continue
            product_text = (post_data.get(key) or "").strip()
            if not product_text:
                continue
            label = product_text.split(" (", 1)[0].strip()
            sku_candidate = ""
            name_candidate = label
            if " - " in label:
                sku_candidate, name_candidate = [part.strip() for part in label.split(" - ", 1)]
            product = None
            if sku_candidate and sku_candidate.lower() != "sin sku":
                product = Product.objects.filter(sku__iexact=sku_candidate).first()
            if not product and name_candidate:
                product = (
                    Product.objects.filter(name__iexact=name_candidate).first()
                    or Product.objects.filter(name__icontains=name_candidate).first()
                )
            if product:
                post_data[product_key] = str(product.id)
        header_form = SaleHeaderForm(post_data)
        formset = SaleItemFormSet(post_data)
        for form in formset.forms:
            prefix = form.prefix
            product_raw = (post_data.get(f"{prefix}-product") or "").strip()
            product_text = (post_data.get(f"{prefix}-product_text") or "").strip()
            qty_raw = (post_data.get(f"{prefix}-quantity") or "").strip()
            qty_value = None
            if qty_raw:
                try:
                    qty_value = Decimal(str(qty_raw).replace(",", "."))
                except Exception:
                    qty_value = None
            qty_zeroish = qty_raw in {"", "0", "0.0", "0.00", "0,00"} or (qty_value is not None and qty_value <= 0)
            if not product_raw and not product_text and qty_zeroish:
                form.empty_permitted = True
            elif not form.has_changed():
                form.empty_permitted = True
        if header_form.is_valid() and formset.is_valid():
            warehouse = header_form.cleaned_data["warehouse"]
            require_variants = warehouse.type == Warehouse.WarehouseType.COMUN
            audience = header_form.cleaned_data.get("audiencia") or Customer.Audience.CONSUMER
            customer = header_form.cleaned_data.get("cliente")
            total_venta = header_form.cleaned_data.get("total_venta")
            sale_date = header_form.cleaned_data.get("sale_date")
            extra_discount_percent = header_form.cleaned_data.get("descuento_total") or Decimal("0.00")
            comision_ml = header_form.cleaned_data.get("comision_ml") or Decimal("0.00")
            impuestos_ml = header_form.cleaned_data.get("impuestos_ml") or Decimal("0.00")
            delivery_status = header_form.cleaned_data.get("delivery_status") or Sale.DeliveryStatus.NOT_DELIVERED
            if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
                customer = None
                audience = Customer.Audience.CONSUMER
                delivery_status = Sale.DeliveryStatus.NOT_DELIVERED
            elif customer:
                audience = customer.audience
            items = []
            for form in formset.forms:
                if form.cleaned_data and form.cleaned_data.get("DELETE"):
                    continue
                product = form.cleaned_data.get("product") if form.cleaned_data else None
                if not product:
                    continue
                variant_id_raw = (post_data.get(f"{form.prefix}-variant_id") or "").strip()
                variant = None
                if variant_id_raw:
                    variant = ProductVariant.objects.filter(id=variant_id_raw, product=product).first()
                    if not variant:
                        form.add_error(None, "Variedad inválida.")
                if require_variants and ProductVariant.objects.filter(product=product).exists() and not variant:
                    form.add_error(None, "Seleccioná una variedad.")
                items.append({**form.cleaned_data, "variant": variant})
            if any(form.errors for form in formset):
                messages.error(request, "Revisá los campos de la venta.")
                return redirect("inventory_sales_list")
            if not items:
                messages.error(request, "Agregá al menos un producto.")
            else:
                try:
                    with transaction.atomic():
                        comun_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.COMUN).first()
                        sale = Sale.objects.create(
                            customer=customer,
                            warehouse=warehouse,
                            audience=audience,
                            delivery_status=delivery_status,
                            reference=f"Venta {audience}",
                            user=request.user,
                            ml_commission_total=comision_ml if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE else Decimal("0.00"),
                            ml_tax_total=impuestos_ml if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE else Decimal("0.00"),
                        )
                        if sale_date:
                            created_at = datetime.combine(sale_date, time(12, 0))
                            if timezone.is_naive(created_at):
                                created_at = timezone.make_aware(created_at)
                            Sale.objects.filter(pk=sale.pk).update(created_at=created_at)
                        if sale_date:
                            created_at = datetime.combine(sale_date, time(12, 0))
                            if timezone.is_naive(created_at):
                                created_at = timezone.make_aware(created_at)
                            Sale.objects.filter(pk=sale.pk).update(created_at=created_at)
                        total = Decimal("0.00")
                        discount_total = Decimal("0.00")
                        base_subtotal = Decimal("0.00")
                        for data in items:
                            base_price = {
                                Customer.Audience.CONSUMER: data["product"].consumer_price,
                                Customer.Audience.BARBER: data["product"].barber_price,
                                Customer.Audience.DISTRIBUTOR: data["product"].distributor_price,
                            }.get(audience, data["product"].consumer_price)
                            discount = Decimal("0.00")
                            if customer:
                                discount_obj = CustomerProductDiscount.objects.filter(
                                    customer=customer, product=data["product"]
                                ).first()
                                if discount_obj:
                                    discount = discount_obj.discount_percent
                                else:
                                    group_key = (data["product"].group or "").strip()
                                    if group_key:
                                        group_discount = CustomerGroupDiscount.objects.filter(
                                            customer=customer, group__iexact=group_key
                                        ).first()
                                        if group_discount:
                                            discount = group_discount.discount_percent
                            final_price = base_price * (Decimal("1.00") - discount / Decimal("100.00"))
                            qty = Decimal(data["quantity"])
                            line_total = (qty * final_price).quantize(Decimal("0.01"))
                            discount_amount = (qty * (base_price - final_price)).quantize(Decimal("0.01"))
                            base_subtotal += (qty * base_price).quantize(Decimal("0.01"))
                            variant = data.get("variant")
                            if warehouse.type == Warehouse.WarehouseType.COMUN and variant:
                                variant = (
                                    ProductVariant.objects.select_for_update()
                                    .filter(id=variant.id, product=data["product"])
                                    .first()
                                )
                                if variant:
                                    if variant.quantity - qty < 0:
                                        raise services.NegativeStockError("Stock insuficiente en variedad.")
                                    variant.quantity = (variant.quantity - qty).quantize(Decimal("0.01"))
                                    variant.save(update_fields=["quantity"])
                                    if comun_wh:
                                        _koda_sync_common_with_variants(data["product"], comun_wh)
                            SaleItem.objects.create(
                                sale=sale,
                                product=data["product"],
                                variant=variant,
                                quantity=qty,
                                unit_price=base_price,
                                cost_unit=data["product"].last_purchase_cost(),
                                discount_percent=discount,
                                final_unit_price=final_price,
                                line_total=line_total,
                                vat_percent=data.get("vat_percent") or Decimal("0.00"),
                            )
                            total += line_total
                            discount_total += discount_amount
                            if warehouse.type != Warehouse.WarehouseType.MERCADOLIBRE:
                                services.register_exit(
                                    product=data["product"],
                                    warehouse=warehouse,
                                    quantity=data["quantity"],
                                    user=request.user,
                                    reference=f"Venta {audience} #{sale.id}",
                                    sale_price=final_price,
                                    vat_percent=data.get("vat_percent") or Decimal("0.00"),
                                    sale=sale,
                                )
                        gross_total = (
                            total_venta
                            if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE and total_venta is not None
                            else total
                        )
                        discount_base = total_venta if total_venta is not None else base_subtotal
                        extra_discount_amount = (discount_base * extra_discount_percent / Decimal("100.00")).quantize(Decimal("0.01"))
                        sale.total = (gross_total - extra_discount_amount).quantize(Decimal("0.01"))
                        sale.discount_total = (discount_total + extra_discount_amount).quantize(Decimal("0.01"))
                        sale.save(update_fields=["total", "discount_total"])
                    messages.success(request, "Venta registrada.")
                    if warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
                        return redirect("inventory_sales_list")
                    return redirect("inventory_sale_receipt", sale_id=sale.id)
                except services.NegativeStockError:
                    messages.error(request, "No hay stock suficiente para completar la venta.")
                except services.InvalidMovementError as exc:
                    messages.error(request, str(exc))
        else:
            messages.error(request, "Revisá los campos de la venta.")
    else:
        header_initial = {"sale_date": timezone.localdate()}
        if default_wh:
            header_initial["warehouse"] = default_wh
        header_form = SaleHeaderForm(initial=header_initial)
        formset = SaleItemFormSet()
    if show_history:
        sales = (
            Sale.objects.select_related("customer", "warehouse", "user")
            .prefetch_related("items__product")
            .order_by("-created_at", "-id")
        )
        if start_date_raw:
            try:
                start_dt = datetime.fromisoformat(start_date_raw)
                if timezone.is_naive(start_dt):
                    start_dt = timezone.make_aware(datetime.combine(start_dt.date(), time.min))
                else:
                    start_dt = datetime.combine(start_dt.date(), time.min, tzinfo=start_dt.tzinfo)
                sales = sales.filter(created_at__gte=start_dt)
            except ValueError:
                messages.error(request, "Fecha desde inválida.")
                return redirect("inventory_sales_list")
        if end_date_raw:
            try:
                end_dt = datetime.fromisoformat(end_date_raw)
                if timezone.is_naive(end_dt):
                    end_dt = timezone.make_aware(datetime.combine(end_dt.date(), time.max))
                else:
                    end_dt = datetime.combine(end_dt.date(), time.max, tzinfo=end_dt.tzinfo)
                sales = sales.filter(created_at__lte=end_dt)
            except ValueError:
                messages.error(request, "Fecha hasta inválida.")
                return redirect("inventory_sales_list")
        if search_query:
            sales = sales.filter(
                Q(customer__name__icontains=search_query)
                | Q(items__product__name__icontains=search_query)
                | Q(items__product__sku__icontains=search_query)
            ).distinct()
        if include_comun and not include_ml:
            sales = sales.filter(warehouse__type=Warehouse.WarehouseType.COMUN)
        elif include_ml and not include_comun:
            sales = sales.filter(warehouse__type=Warehouse.WarehouseType.MERCADOLIBRE)
        page_number = request.GET.get("page")
        paginator = Paginator(sales, 25)
        page_obj = paginator.get_page(page_number)
        sales_list = list(page_obj.object_list)
        for sale in sales_list:
            cost_total = Decimal("0.00")
            for item in sale.items.all():
                cost_unit = item.cost_unit
                if cost_unit is None or cost_unit <= 0:
                    cost_unit = item.product.cost_with_vat()
                cost_total += item.quantity * cost_unit
            commission_total = sale.ml_commission_total or Decimal("0.00")
            tax_total = sale.ml_tax_total or Decimal("0.00")
            is_ml_sale = (
                sale.ml_order_id
                or sale.reference.startswith("ML ORDER")
                or sale.reference.startswith("GS ORDER")
                or sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE
            )
            if is_ml_sale:
                net_total = (sale.total or Decimal("0.00")) - commission_total - tax_total
                sale.margin_total = net_total - cost_total
            else:
                sale.margin_total = (sale.total or Decimal("0.00")) - cost_total
        sales_comun = [sale for sale in sales_list if sale.warehouse.type == Warehouse.WarehouseType.COMUN]
        sales_ml = [sale for sale in sales_list if sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE]
    else:
        sales_list = []
        sales_comun = []
        sales_ml = []
        page_obj = None
    show_comun = show_history and (include_comun or (not include_ml and not include_comun))
    show_ml = show_history and (include_ml or (not include_ml and not include_comun))
    delivery_status_choices = [
        (Sale.DeliveryStatus.NOT_DELIVERED, Sale.DeliveryStatus.NOT_DELIVERED.label),
        (Sale.DeliveryStatus.IN_TRANSIT, Sale.DeliveryStatus.IN_TRANSIT.label),
        (Sale.DeliveryStatus.DELIVERED, Sale.DeliveryStatus.DELIVERED.label),
    ]
    if request.GET.get("ajax") == "1":
        from django.template.loader import render_to_string

        html = render_to_string(
            "inventory/_sales_history.html",
            {
                "sales_comun": sales_comun,
                "sales_ml": sales_ml,
                "page_obj": page_obj,
                "search_query": search_query,
                "start_date": start_date_raw,
                "end_date": end_date_raw,
                "include_comun": include_comun,
                "include_ml": include_ml,
                "show_history": show_history,
                "show_comun": show_comun,
                "show_ml": show_ml,
                "delivery_status_choices": delivery_status_choices,
            },
            request=request,
        )
        return JsonResponse({"html": html})
    variant_data = {}
    for row in ProductVariant.objects.values("id", "product_id", "name").order_by("name", "id"):
        variant_data.setdefault(str(row["product_id"]), []).append({"id": row["id"], "name": row["name"]})
    return render(
        request,
        "inventory/sales_list.html",
        {
            "sales": sales_list,
            "sales_comun": sales_comun,
            "sales_ml": sales_ml,
            "page_obj": page_obj,
            "search_query": search_query,
            "start_date": start_date_raw,
            "end_date": end_date_raw,
            "include_comun": include_comun,
            "include_ml": include_ml,
            "show_history": show_history,
            "show_comun": show_comun,
            "show_ml": show_ml,
            "delivery_status_choices": delivery_status_choices,
            "sales_sheet_url": os.environ.get("GOOGLE_SHEETS_SALES_URL", ""),
            "customers": customers,
            "form": header_form,
            "formset": formset,
            "customer_audiences": customer_audiences,
            "variant_data": variant_data,
            "clear_sale_form": request.method != "POST",
            "default_warehouse_id": str(default_wh.id) if default_wh else "",
        },
    )


@login_required
@require_http_methods(["POST"])
def sale_delivery_status_update(request, sale_id: int):
    sale = get_object_or_404(Sale, pk=sale_id)
    delivery_status = (request.POST.get("delivery_status") or "").strip()
    allowed_statuses = {
        Sale.DeliveryStatus.NOT_DELIVERED,
        Sale.DeliveryStatus.IN_TRANSIT,
        Sale.DeliveryStatus.DELIVERED,
    }
    if delivery_status not in allowed_statuses:
        return JsonResponse({"ok": False, "error": "Estado inválido."}, status=400)
    if sale.delivery_status != delivery_status:
        sale.delivery_status = delivery_status
        sale.save(update_fields=["delivery_status"])
    return JsonResponse(
        {
            "ok": True,
            "delivery_status": sale.delivery_status,
            "label": sale.get_delivery_status_display(),
        }
    )


@login_required
@require_http_methods(["POST"])
def sale_delete(request, sale_id: int):
    sale = get_object_or_404(Sale.objects.prefetch_related("items", "movements"), pk=sale_id)
    try:
        with transaction.atomic():
            is_ml_sale = sale.ml_order_id or sale.reference.startswith("ML ORDER") or sale.reference.startswith("GS ORDER")
            comun_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.COMUN).first()
            if sale.warehouse.type == Warehouse.WarehouseType.COMUN:
                for item in sale.items.select_related("variant", "product"):
                    if item.variant_id:
                        variant = (
                            ProductVariant.objects.select_for_update()
                            .filter(id=item.variant_id, product=item.product)
                            .first()
                        )
                        if variant:
                            variant.quantity = (variant.quantity + item.quantity).quantize(Decimal("0.01"))
                            variant.save(update_fields=["quantity"])
                            if comun_wh:
                                _koda_sync_common_with_variants(item.product, comun_wh)
            # Revert stock only for non-ML sales
            for movement in sale.movements.select_for_update():
                if not is_ml_sale and movement.movement_type == StockMovement.MovementType.EXIT and movement.from_warehouse:
                    stock, _ = Stock.objects.select_for_update().get_or_create(
                        product=movement.product,
                        warehouse=movement.from_warehouse,
                        defaults={"quantity": Decimal("0.00")},
                    )
                    stock.quantity = (stock.quantity + movement.quantity).quantize(Decimal("0.01"))
                    stock.save(update_fields=["quantity"])
                movement.delete()
            sale.delete()
        if is_ml_sale:
            messages.success(request, "Venta eliminada.")
        else:
            messages.success(request, "Venta eliminada y stock ajustado.")
    except Exception as exc:
        messages.error(request, f"No se pudo eliminar la venta: {exc}")
    return redirect("inventory_sales_list")


@login_required
def purchases_list(request):
    PurchaseItemFormSet = formset_factory(PurchaseItemForm, extra=1, can_delete=True)
    show_history = request.GET.get("show_history") == "1"
    if request.method == "POST":
        header_form = PurchaseHeaderForm(request.POST)
        formset = PurchaseItemFormSet(request.POST)
        payload_ids = _extract_product_ids_from_payload(request.POST.get("items_payload"))
        if payload_ids:
            products_qs = _products_with_last_cost_queryset().filter(id__in=payload_ids)
        else:
            products_qs = Product.objects.none()
        _apply_product_queryset_to_formset(formset, products_qs)
        for form in formset.forms:
            prefix = form.prefix
            product_raw = (request.POST.get(f"{prefix}-product") or "").strip()
            qty_raw = (request.POST.get(f"{prefix}-quantity") or "").strip()
            cost_raw = (request.POST.get(f"{prefix}-unit_cost") or "").strip()
            qty_zeroish = qty_raw in {"", "0", "0.0", "0.00", "0,00"}
            cost_zeroish = cost_raw in {"", "0", "0.0", "0.00", "0,00"}
            if not product_raw and qty_zeroish and cost_zeroish:
                form.empty_permitted = True
            elif not form.has_changed():
                form.empty_permitted = True
        if header_form.is_valid():
            def parse_decimal(raw: str | None) -> Decimal | None:
                if raw is None:
                    return None
                value = str(raw).strip().replace(",", ".")
                if not value:
                    return None
                try:
                    return Decimal(value)
                except Exception:
                    return None

            def resolve_product(product_id: str, product_text: str):
                product = None
                if product_id:
                    product = Product.objects.filter(id=product_id).first()
                if not product and product_text:
                    label = product_text.split(" (", 1)[0].strip()
                    sku_candidate = ""
                    name_candidate = label
                    if " - " in label:
                        sku_candidate, name_candidate = [part.strip() for part in label.split(" - ", 1)]
                    if sku_candidate and sku_candidate.lower() != "sin sku":
                        product = Product.objects.filter(sku__iexact=sku_candidate).first()
                    if not product and name_candidate:
                        product = (
                            Product.objects.filter(name__iexact=name_candidate).first()
                            or Product.objects.filter(name__icontains=name_candidate).first()
                        )
                return product

            def parse_items_from_payload() -> tuple[list[dict], list[str]]:
                raw = request.POST.get("items_payload")
                if not raw:
                    return [], []
                try:
                    items_raw = json.loads(raw)
                except Exception:
                    return [], ["No se pudo leer el detalle de la compra."]
                if not isinstance(items_raw, list):
                    return [], ["Detalle de compra inválido."]
                items_data: list[dict] = []
                errors: list[str] = []
                for idx, entry in enumerate(items_raw, start=1):
                    if not isinstance(entry, dict):
                        continue
                    product_id = str(entry.get("product_id") or "").strip()
                    product_text = str(entry.get("product_text") or "").strip()
                    qty_raw = entry.get("quantity")
                    cost_raw = entry.get("unit_cost")
                    discount_raw = entry.get("discount_percent")
                    supplier_id = str(entry.get("supplier_id") or "").strip()
                    variant_id = str(entry.get("variant_id") or "").strip()
                    vat_raw = entry.get("vat_percent")
                    if not product_id and not str(qty_raw or "").strip() and not str(cost_raw or "").strip():
                        continue
                    if not product_id and not product_text:
                        errors.append(f"Fila {idx}: producto inválido.")
                        continue
                    if not str(qty_raw or "").strip() or not str(cost_raw or "").strip():
                        errors.append(f"Fila {idx}: completá cantidad y costo.")
                        continue
                    product = resolve_product(product_id, product_text)
                    if not product:
                        errors.append(f"Fila {idx}: producto inválido.")
                        continue
                    variant = None
                    if variant_id:
                        variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                        if not variant:
                            errors.append(f"Fila {idx}: variedad inválida.")
                            continue
                    if ProductVariant.objects.filter(product=product).exists() and not variant:
                        errors.append(f"Fila {idx}: elegí variedad.")
                        continue
                    qty = parse_decimal(str(qty_raw))
                    unit_cost = parse_decimal(str(cost_raw))
                    discount_percent = parse_decimal(str(discount_raw)) or Decimal("0.00")
                    vat_percent = parse_decimal(str(vat_raw)) or Decimal("0.00")
                    if qty is None or qty <= 0:
                        errors.append(f"Fila {idx}: cantidad inválida.")
                        continue
                    if unit_cost is None or unit_cost < 0:
                        errors.append(f"Fila {idx}: costo inválido.")
                        continue
                    if discount_percent < 0 or discount_percent > 100:
                        errors.append(f"Fila {idx}: descuento inválido.")
                        continue
                    supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id else None
                    items_data.append(
                        {
                            "product": product,
                            "quantity": qty,
                            "unit_cost": unit_cost,
                            "discount_percent": discount_percent,
                            "vat_percent": vat_percent,
                            "supplier": supplier,
                            "variant": variant,
                        }
                    )
                return items_data, errors

            def parse_items_from_post() -> tuple[list[dict], list[str]]:
                indices: set[int] = set()
                for key in request.POST.keys():
                    if key.startswith("form-") and (key.endswith("-product") or key.endswith("-product_text")):
                        parts = key.split("-")
                        if len(parts) >= 3 and parts[1].isdigit():
                            indices.add(int(parts[1]))
                items_data: list[dict] = []
                errors: list[str] = []
                for idx in sorted(indices):
                    prefix = f"form-{idx}-"
                    product_id = (request.POST.get(f"{prefix}product") or "").strip()
                    product_text = (request.POST.get(f"{prefix}product_text") or "").strip()
                    qty_raw = request.POST.get(f"{prefix}quantity")
                    cost_raw = request.POST.get(f"{prefix}unit_cost")
                    discount_raw = request.POST.get(f"{prefix}discount_percent")
                    supplier_id = (request.POST.get(f"{prefix}supplier") or "").strip()
                    variant_id = (request.POST.get(f"{prefix}variant_id") or "").strip()
                    vat_raw = request.POST.get(f"{prefix}vat_percent")
                    if not product_id and not (qty_raw or "").strip() and not (cost_raw or "").strip():
                        continue
                    if not product_id and product_text:
                        label = product_text.split(" (", 1)[0].strip()
                        sku_candidate = ""
                        name_candidate = label
                        if " - " in label:
                            sku_candidate, name_candidate = [part.strip() for part in label.split(" - ", 1)]
                        if sku_candidate and sku_candidate.lower() != "sin sku":
                            product = Product.objects.filter(sku__iexact=sku_candidate).first()
                        else:
                            product = None
                        if not product and name_candidate:
                            product = (
                                Product.objects.filter(name__iexact=name_candidate).first()
                                or Product.objects.filter(name__icontains=name_candidate).first()
                            )
                        if product:
                            product_id = str(product.id)
                    if not product_id or not (qty_raw or "").strip() or not (cost_raw or "").strip():
                        errors.append(f"Fila {idx + 1}: completá producto, cantidad y costo.")
                        continue
                    product = resolve_product(product_id, product_text)
                    if not product:
                        errors.append(f"Fila {idx + 1}: producto inválido.")
                        continue
                    variant = None
                    if variant_id:
                        variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                        if not variant:
                            errors.append(f"Fila {idx + 1}: variedad inválida.")
                            continue
                    if ProductVariant.objects.filter(product=product).exists() and not variant:
                        errors.append(f"Fila {idx + 1}: elegí variedad.")
                        continue
                    qty = parse_decimal(qty_raw)
                    unit_cost = parse_decimal(cost_raw)
                    discount_percent = parse_decimal(discount_raw) or Decimal("0.00")
                    vat_percent = parse_decimal(vat_raw) or Decimal("0.00")
                    if qty is None or qty <= 0:
                        errors.append(f"Fila {idx + 1}: cantidad inválida.")
                        continue
                    if unit_cost is None or unit_cost < 0:
                        errors.append(f"Fila {idx + 1}: costo inválido.")
                        continue
                    if discount_percent < 0 or discount_percent > 100:
                        errors.append(f"Fila {idx + 1}: descuento inválido.")
                        continue
                    supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id else None
                    items_data.append(
                        {
                            "product": product,
                            "quantity": qty,
                            "unit_cost": unit_cost,
                            "discount_percent": discount_percent,
                            "vat_percent": vat_percent,
                            "supplier": supplier,
                            "variant": variant,
                        }
                    )
                return items_data, errors

            items, parse_errors = parse_items_from_payload()
            if not items and not parse_errors:
                items, parse_errors = parse_items_from_post()
            if parse_errors:
                for err in parse_errors[:3]:
                    messages.error(request, err)
                if len(parse_errors) > 3:
                    messages.error(request, f"Hay {len(parse_errors)} filas con errores.")
                items = []
            warehouse = header_form.cleaned_data["warehouse"]
            purchase_date = header_form.cleaned_data.get("purchase_date")
            discount_percent = header_form.cleaned_data.get("descuento_total") or Decimal("0.00")
            if not items:
                messages.error(request, "Agregá al menos un producto.")
                return redirect("inventory_purchases_list")
            try:
                with transaction.atomic():
                    purchase_supplier = items[0].get("supplier") if items else None
                    purchase = Purchase.objects.create(
                        supplier=purchase_supplier,
                        warehouse=warehouse,
                        reference="Compra",
                        discount_percent=discount_percent,
                        user=request.user,
                    )
                    if purchase_date:
                        created_at = datetime.combine(purchase_date, time(12, 0))
                        if timezone.is_naive(created_at):
                            created_at = timezone.make_aware(created_at)
                        Purchase.objects.filter(pk=purchase.pk).update(created_at=created_at)
                    subtotal = Decimal("0.00")
                    for data in items:
                        qty = Decimal(data["quantity"])
                        unit_cost = data["unit_cost"]
                        discount_percent = data.get("discount_percent") or Decimal("0.00")
                        effective_unit_cost = (unit_cost * (Decimal("1.00") - (discount_percent / Decimal("100.00")))).quantize(
                            Decimal("0.01"), rounding=ROUND_HALF_UP
                        )
                        subtotal += qty * effective_unit_cost
                        PurchaseItem.objects.create(
                            purchase=purchase,
                            product=data["product"],
                            variant=data.get("variant"),
                            quantity=qty,
                            unit_cost=unit_cost,
                            discount_percent=discount_percent,
                            vat_percent=data.get("vat_percent") or Decimal("0.00"),
                        )
                        if warehouse.type == Warehouse.WarehouseType.COMUN and data.get("variant") is not None:
                            variant = (
                                ProductVariant.objects.select_for_update()
                                .filter(id=data["variant"].id, product=data["product"])
                                .first()
                            )
                            if variant:
                                variant.quantity = (variant.quantity + qty).quantize(Decimal("0.01"))
                                variant.save(update_fields=["quantity"])
                                _koda_sync_common_with_variants(data["product"], warehouse)
                        services.register_entry(
                            product=data["product"],
                            warehouse=warehouse,
                            quantity=qty,
                            unit_cost=effective_unit_cost,
                            supplier=data["supplier"],
                            vat_percent=data.get("vat_percent") or Decimal("0.00"),
                            user=request.user,
                            reference=f"Compra #{purchase.id}",
                            purchase=purchase,
                        )
                    discount_total = (subtotal * discount_percent / Decimal("100.00")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                    total = (subtotal - discount_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    purchase.total = total
                    purchase.save(update_fields=["total", "discount_percent"])
                messages.success(request, "Compra registrada.")
                return redirect("inventory_purchases_list")
            except Exception as exc:
                messages.error(request, f"No se pudo registrar la compra: {exc}")
                return redirect("inventory_purchases_list")
        messages.error(request, "Revisá los campos de la compra.")
    else:
        header_form = PurchaseHeaderForm()
        formset = PurchaseItemFormSet()
        _apply_product_queryset_to_formset(formset, Product.objects.none())

    if show_history:
        purchases = (
            Purchase.objects.select_related("supplier", "warehouse", "user")
            .order_by("-created_at", "-id")
        )
        paginator = Paginator(purchases, 25)
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)
        purchase_list = page_obj.object_list
    else:
        page_obj = None
        purchase_list = []
    variant_data = {}
    for row in ProductVariant.objects.values("id", "product_id", "name").order_by("name", "id"):
        variant_data.setdefault(str(row["product_id"]), []).append({"id": row["id"], "name": row["name"]})
    return render(
        request,
        "inventory/purchases_list.html",
        {
            "purchases": purchase_list,
            "page_obj": page_obj,
            "form": header_form,
            "formset": formset,
            "show_history": show_history,
            "variant_data": variant_data,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def purchase_edit(request, purchase_id: int):
    purchase = get_object_or_404(
        Purchase.objects.select_related("supplier", "warehouse").prefetch_related("items__product", "movements"),
        pk=purchase_id,
    )
    PurchaseItemFormSet = formset_factory(PurchaseItemForm, extra=0, can_delete=True)
    if request.method == "POST":
        header_form = PurchaseHeaderForm(request.POST)
        formset = PurchaseItemFormSet(request.POST)
        payload_ids = _extract_product_ids_from_payload(request.POST.get("items_payload"))
        if payload_ids:
            products_qs = _products_with_last_cost_queryset().filter(id__in=payload_ids)
        else:
            products_qs = Product.objects.none()
        _apply_product_queryset_to_formset(formset, products_qs)
        items_payload_raw = request.POST.get("items_payload")
        for form in formset.forms:
            prefix = form.prefix
            product_raw = (request.POST.get(f"{prefix}-product") or "").strip()
            qty_raw = (request.POST.get(f"{prefix}-quantity") or "").strip()
            cost_raw = (request.POST.get(f"{prefix}-unit_cost") or "").strip()
            qty_zeroish = qty_raw in {"", "0", "0.0", "0.00", "0,00"}
            cost_zeroish = cost_raw in {"", "0", "0.0", "0.00", "0,00"}
            if not product_raw and qty_zeroish and cost_zeroish:
                form.empty_permitted = True
            elif not form.has_changed():
                form.empty_permitted = True
        if header_form.is_valid():
            items = []
            errors: list[str] = []
            if items_payload_raw:
                def parse_decimal(raw: str | None) -> Decimal | None:
                    if raw is None:
                        return None
                    value = str(raw).strip().replace(",", ".")
                    if not value:
                        return None
                    try:
                        return Decimal(value)
                    except Exception:
                        return None

                def resolve_product(product_id: str, product_text: str):
                    product = None
                    if product_id:
                        product = Product.objects.filter(id=product_id).first()
                    if not product and product_text:
                        label = product_text.split(" (", 1)[0].strip()
                        sku_candidate = ""
                        name_candidate = label
                        if " - " in label:
                            sku_candidate, name_candidate = [part.strip() for part in label.split(" - ", 1)]
                        if sku_candidate and sku_candidate.lower() != "sin sku":
                            product = Product.objects.filter(sku__iexact=sku_candidate).first()
                        if not product and name_candidate:
                            product = (
                                Product.objects.filter(name__iexact=name_candidate).first()
                                or Product.objects.filter(name__icontains=name_candidate).first()
                            )
                    return product

                try:
                    items_raw = json.loads(items_payload_raw)
                except Exception:
                    items_raw = []
                if isinstance(items_raw, list):
                    for idx, entry in enumerate(items_raw, start=1):
                        if not isinstance(entry, dict):
                            continue
                        product_id = str(entry.get("product_id") or "").strip()
                        product_text = str(entry.get("product_text") or "").strip()
                        qty_raw = entry.get("quantity")
                        cost_raw = entry.get("unit_cost")
                        discount_raw = entry.get("discount_percent")
                        supplier_id = str(entry.get("supplier_id") or "").strip()
                        variant_id = str(entry.get("variant_id") or "").strip()
                        vat_raw = entry.get("vat_percent")
                        if not product_id and not product_text:
                            continue
                        product = resolve_product(product_id, product_text)
                        qty = parse_decimal(str(qty_raw))
                        unit_cost = parse_decimal(str(cost_raw))
                        discount_percent = parse_decimal(str(discount_raw)) or Decimal("0.00")
                        if not product or qty is None or qty <= 0 or unit_cost is None:
                            errors.append(f"Fila {idx}: datos inválidos.")
                            continue
                        variant = None
                        if variant_id:
                            variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                            if not variant:
                                errors.append(f"Fila {idx}: variedad inválida.")
                                continue
                        if ProductVariant.objects.filter(product=product).exists() and not variant:
                            errors.append(f"Fila {idx}: elegí variedad.")
                            continue
                        if discount_percent < 0 or discount_percent > 100:
                            errors.append(f"Fila {idx}: descuento inválido.")
                            continue
                        supplier = Supplier.objects.filter(id=supplier_id).first() if supplier_id else None
                        items.append(
                            {
                                "product": product,
                                "quantity": qty,
                                "unit_cost": unit_cost,
                                "discount_percent": discount_percent,
                                "vat_percent": parse_decimal(str(vat_raw)) or Decimal("0.00"),
                                "supplier": supplier,
                                "variant": variant,
                            }
                        )
            elif formset.is_valid():
                for form in formset:
                    if form.cleaned_data.get("DELETE"):
                        continue
                    if not form.cleaned_data.get("product"):
                        continue
                    prefix = form.prefix
                    product = form.cleaned_data.get("product")
                    variant_id = (request.POST.get(f"{prefix}-variant_id") or "").strip()
                    variant = None
                    if variant_id:
                        variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                        if not variant:
                            errors.append("Variedad inválida.")
                            continue
                    if product and ProductVariant.objects.filter(product=product).exists() and not variant:
                        errors.append("Elegí variedad.")
                        continue
                    items.append({**form.cleaned_data, "variant": variant})
            if errors:
                for err in errors[:3]:
                    messages.error(request, err)
            if not items:
                messages.error(request, "Agregá al menos un producto.")
                return redirect("inventory_purchase_edit", purchase_id=purchase.id)
            try:
                with transaction.atomic():
                    new_warehouse = header_form.cleaned_data["warehouse"]
                    current_items = list(purchase.items.all())
                    current_signature = sorted(
                        [
                            (item.product_id, item.variant_id or 0, Decimal(item.quantity))
                            for item in current_items
                        ],
                        key=lambda row: (row[0], row[1], row[2]),
                    )
                    new_signature = sorted(
                        [
                            (item["product"].id, item.get("variant").id if item.get("variant") else 0, Decimal(item["quantity"]))
                            for item in items
                        ],
                        key=lambda row: (row[0], row[1], row[2]),
                    )
                    stock_changed = (
                        purchase.warehouse_id != new_warehouse.id
                        or current_signature != new_signature
                    )
                    if stock_changed:
                        # Revert previous stock movements
                        for movement in purchase.movements.select_for_update():
                            if movement.movement_type == StockMovement.MovementType.ENTRY and movement.to_warehouse:
                                stock, _ = Stock.objects.select_for_update().get_or_create(
                                    product=movement.product,
                                    warehouse=movement.to_warehouse,
                                    defaults={"quantity": Decimal("0.00")},
                                )
                                stock.quantity = (stock.quantity - movement.quantity).quantize(Decimal("0.01"))
                                if stock.quantity < 0:
                                    raise services.NegativeStockError("Stock cannot go negative")
                                stock.save(update_fields=["quantity"])
                            movement.delete()
                        if purchase.warehouse.type == Warehouse.WarehouseType.COMUN:
                            for item in current_items:
                                if not item.variant_id:
                                    continue
                                variant = (
                                    ProductVariant.objects.select_for_update()
                                    .filter(id=item.variant_id, product=item.product)
                                    .first()
                                )
                                if not variant:
                                    continue
                                if variant.quantity - item.quantity < 0:
                                    raise services.NegativeStockError("Stock cannot go negative")
                                variant.quantity = (variant.quantity - item.quantity).quantize(Decimal("0.01"))
                                variant.save(update_fields=["quantity"])
                                _koda_sync_common_with_variants(item.product, purchase.warehouse)
                    purchase.items.all().delete()

                    purchase.warehouse = new_warehouse
                    purchase.supplier = items[0].get("supplier")
                    purchase_date = header_form.cleaned_data.get("purchase_date")
                    discount_percent = header_form.cleaned_data.get("descuento_total") or Decimal("0.00")
                    purchase.discount_percent = discount_percent
                    update_fields = ["warehouse", "supplier", "discount_percent"]
                    if purchase_date:
                        created_at = datetime.combine(purchase_date, time(12, 0))
                        if timezone.is_naive(created_at):
                            created_at = timezone.make_aware(created_at)
                        purchase.created_at = created_at
                        update_fields.append("created_at")
                    purchase.save(update_fields=update_fields)

                    subtotal = Decimal("0.00")
                    for item in items:
                        product = item["product"]
                        qty = item["quantity"]
                        unit_cost = item["unit_cost"]
                        discount_percent = item.get("discount_percent") or Decimal("0.00")
                        effective_unit_cost = (unit_cost * (Decimal("1.00") - (discount_percent / Decimal("100.00")))).quantize(
                            Decimal("0.01"), rounding=ROUND_HALF_UP
                        )
                        vat = item.get("vat_percent") or Decimal("0.00")
                        subtotal += effective_unit_cost * qty
                        PurchaseItem.objects.create(
                            purchase=purchase,
                            product=product,
                            variant=item.get("variant"),
                            quantity=qty,
                            unit_cost=unit_cost,
                            discount_percent=discount_percent,
                            vat_percent=vat,
                        )
                        if stock_changed and purchase.warehouse.type == Warehouse.WarehouseType.COMUN and item.get("variant") is not None:
                            variant = (
                                ProductVariant.objects.select_for_update()
                                .filter(id=item["variant"].id, product=product)
                                .first()
                            )
                            if variant:
                                variant.quantity = (variant.quantity + qty).quantize(Decimal("0.01"))
                                variant.save(update_fields=["quantity"])
                                _koda_sync_common_with_variants(product, purchase.warehouse)
                        if stock_changed:
                            services.register_entry(
                                product=product,
                                warehouse=purchase.warehouse,
                                quantity=qty,
                                unit_cost=effective_unit_cost,
                                vat_percent=vat,
                                user=request.user,
                                reference=f"Compra #{purchase.id}",
                                supplier=purchase.supplier,
                                purchase=purchase,
                            )
                    discount_total = (subtotal * discount_percent / Decimal("100.00")).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                    total = (subtotal - discount_total).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    purchase.total = total
                    purchase.save(update_fields=["total"])
                messages.success(request, "Compra actualizada.")
                return redirect("inventory_purchase_receipt", purchase_id=purchase.id)
            except services.NegativeStockError:
                messages.error(request, "No se puede actualizar: el stock quedaría negativo.")
                return redirect("inventory_purchase_edit", purchase_id=purchase.id)
            except Exception as exc:
                messages.error(request, f"No se pudo actualizar la compra: {exc}")
                return redirect("inventory_purchase_edit", purchase_id=purchase.id)
        messages.error(request, "Revisá los campos de la compra.")
        return redirect("inventory_purchase_edit", purchase_id=purchase.id)
    else:
        header_form = PurchaseHeaderForm(
            initial={
                "warehouse": purchase.warehouse,
                "purchase_date": timezone.localtime(purchase.created_at).date(),
                "descuento_total": purchase.discount_percent,
            }
        )
        purchase_items = list(purchase.items.select_related("variant", "product"))
        initial = [
            {
                "product": item.product,
                "quantity": int(item.quantity),
                "unit_cost": item.unit_cost,
                "discount_percent": item.discount_percent,
                "supplier": purchase.supplier,
                "vat_percent": item.vat_percent,
            }
            for item in purchase_items
        ]
        formset = PurchaseItemFormSet(initial=initial)
        product_ids = {item["product"].id for item in initial}
        products_qs = _products_with_last_cost_queryset().filter(id__in=product_ids) if product_ids else Product.objects.none()
        _apply_product_queryset_to_formset(formset, products_qs)
        item_rows = [
            {"form": form, "variant_id": item.variant_id}
            for form, item in zip(formset.forms, purchase_items)
        ]
    variant_data = {}
    for row in ProductVariant.objects.values("id", "product_id", "name").order_by("name", "id"):
        variant_data.setdefault(str(row["product_id"]), []).append({"id": row["id"], "name": row["name"]})
    return render(
        request,
        "inventory/purchase_edit.html",
        {
            "purchase": purchase,
            "form": header_form,
            "formset": formset,
            "item_rows": item_rows if request.method != "POST" else None,
            "variant_data": variant_data,
        },
    )


@login_required
@require_http_methods(["POST"])
def purchase_delete(request, purchase_id: int):
    purchase = get_object_or_404(Purchase.objects.prefetch_related("items", "movements"), pk=purchase_id)
    try:
        with transaction.atomic():
            for movement in purchase.movements.select_for_update():
                if movement.movement_type == StockMovement.MovementType.ENTRY and movement.to_warehouse:
                    stock, _ = Stock.objects.select_for_update().get_or_create(
                        product=movement.product,
                        warehouse=movement.to_warehouse,
                        defaults={"quantity": Decimal("0.00")},
                    )
                    stock.quantity = (stock.quantity - movement.quantity).quantize(Decimal("0.01"))
                    if stock.quantity < 0:
                        raise services.NegativeStockError("Stock cannot go negative")
                    stock.save(update_fields=["quantity"])
                movement.delete()
            purchase.delete()
        messages.success(request, "Compra eliminada y stock ajustado.")
    except services.NegativeStockError:
        messages.error(request, "No se puede eliminar: el stock quedaría negativo.")
    except Exception as exc:
        messages.error(request, f"No se pudo eliminar la compra: {exc}")
    return redirect("inventory_purchases_list")


@login_required
@require_http_methods(["GET", "POST"])
def suppliers(request):
    supplier_form = SupplierForm()
    link_form = SupplierProductForm()
    suppliers_qs = Supplier.objects.prefetch_related("supplier_products__product")

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_supplier":
            supplier_form = SupplierForm(request.POST)
            if supplier_form.is_valid():
                supplier_form.save()
                messages.success(request, "Proveedor creado.")
                return redirect("inventory_suppliers")
        elif action == "link_supplier":
            link_form = SupplierProductForm(request.POST)
            if link_form.is_valid():
                supplier = link_form.cleaned_data["supplier"]
                product = link_form.cleaned_data["product"]
                last_cost = link_form.cleaned_data.get("last_cost") or product.avg_cost
                SupplierProduct.objects.update_or_create(
                    supplier=supplier,
                    product=product,
                    defaults={"last_cost": last_cost, "last_purchase_at": timezone.now()},
                )
                if product.default_supplier_id is None:
                    product.default_supplier = supplier
                    product.save(update_fields=["default_supplier"])
                messages.success(request, "Proveedor vinculado al producto.")
                return redirect("inventory_suppliers")
        elif action == "delete_supplier":
            supplier_id = request.POST.get("supplier_id")
            Supplier.objects.filter(pk=supplier_id).delete()
            messages.success(request, "Proveedor eliminado.")
            return redirect("inventory_suppliers")
        elif action == "remove_link":
            link_id = request.POST.get("link_id")
            SupplierProduct.objects.filter(pk=link_id).delete()
            messages.success(request, "Vínculo eliminado.")
            return redirect("inventory_suppliers")

    context = {
        "supplier_form": supplier_form,
        "link_form": link_form,
        "suppliers": suppliers_qs,
    }
    return render(request, "inventory/suppliers.html", context)


@login_required
def stock_list(request):
    decimal_field = DecimalField(max_digits=12, decimal_places=2)
    comun_code = Warehouse.WarehouseType.COMUN
    ml_code = Warehouse.WarehouseType.MERCADOLIBRE
    comun_wh = Warehouse.objects.filter(type=comun_code).first()
    ml_wh = Warehouse.objects.filter(type=ml_code).first()
    transfer_form = StockTransferForm()
    query = (request.GET.get("q") or "").strip()
    show_history = (request.GET.get("show_history") or "").strip() == "1"
    aris_supplier = Supplier.objects.filter(name__iexact="Aris Norma").first()
    aurill_supplier = Supplier.objects.filter(name__iexact="Aurill- Dario").first()

    def parse_decimal(value: str) -> Decimal:
        raw = (value or "").strip().replace(" ", "")
        if not raw:
            return Decimal("0.00")
        if "," in raw and "." in raw:
            raw = raw.replace(".", "").replace(",", ".")
        elif "," in raw:
            raw = raw.replace(",", ".")
        try:
            return Decimal(raw)
        except Exception:
            return Decimal("0.00")

    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "set_comun_stock":
            if not comun_wh:
                messages.error(request, "Falta el depósito común.")
                return redirect("inventory_stock_list")
            product_id = request.POST.get("product_id")
            desired_raw = request.POST.get("quantity")
            product = Product.objects.filter(pk=product_id).first()
            if not product:
                messages.error(request, "Producto no encontrado.")
                return redirect("inventory_stock_list")
            desired = parse_decimal(desired_raw)
            stock = Stock.objects.filter(product=product, warehouse=comun_wh).first()
            current = stock.quantity if stock else Decimal("0.00")
            diff = (desired - current).quantize(Decimal("0.01"))
            try:
                if diff != 0:
                    services.register_adjustment(
                        product=product,
                        warehouse=comun_wh,
                        quantity=diff,
                        user=request.user,
                        reference="Ajuste manual depósito común",
                        allow_negative=True,
                    )
                messages.success(request, "Stock actualizado.")
                return redirect("inventory_stock_list")
            except services.InvalidMovementError as exc:
                messages.error(request, str(exc))
        elif action == "create_ml_purchase":
            if not ml_wh:
                messages.error(request, "Falta el depósito MercadoLibre.")
                return redirect("inventory_stock_list")

            products = (
                Product.objects.order_by("sku")
                .annotate(
                    ml_qty=Coalesce(
                        Sum(
                            Case(
                                When(stocks__warehouse__type=ml_code, then="stocks__quantity"),
                                output_field=decimal_field,
                            )
                        ),
                        Value(0, output_field=decimal_field),
                        output_field=decimal_field,
                    ),
                )
                .filter(ml_qty__gt=0)
            )
            if not products.exists():
                messages.error(request, "No hay stock en MercadoLibre para importar.")
                return redirect("inventory_stock_list")

            purchases = {
                "aurill": {"supplier": aurill_supplier, "items": []},
                "aris": {"supplier": aris_supplier, "items": []},
            }
            for product in products:
                qty = product.ml_qty or Decimal("0.00")
                if qty <= 0:
                    continue
                vat = product.vat_percent or Decimal("0.00")
                unit_cost = product.cost_with_vat()
                target = "aurill" if "aurill" in (product.group or "").lower() else "aris"
                purchases[target]["items"].append(
                    {
                        "product": product,
                        "quantity": qty,
                        "unit_cost": unit_cost.quantize(Decimal("0.01")),
                        "vat_percent": vat,
                    }
                )

            created = 0
            with transaction.atomic():
                for data in purchases.values():
                    if not data["items"]:
                        continue
                    purchase = Purchase.objects.create(
                        supplier=data["supplier"],
                        warehouse=ml_wh,
                        reference="Compra ML stock inicial",
                        user=request.user,
                    )
                    total = Decimal("0.00")
                    for item in data["items"]:
                        qty = item["quantity"]
                        unit_cost = item["unit_cost"]
                        total += qty * unit_cost
                        PurchaseItem.objects.create(
                            purchase=purchase,
                            product=item["product"],
                            quantity=qty,
                            unit_cost=unit_cost,
                            vat_percent=item["vat_percent"],
                        )
                    purchase.total = total.quantize(Decimal("0.01"))
                    purchase.save(update_fields=["total"])
                    created += 1
            messages.success(request, f"Compras creadas: {created}.")
            return redirect("inventory_stock_list")
        else:
            transfer_form = StockTransferForm(request.POST)
            if not comun_wh or not ml_wh:
                messages.error(request, "Faltan depósitos configurados para transferir stock.")
                return redirect("inventory_stock_list")
            def sync_common_with_variants(product):
                total = (
                    ProductVariant.objects.filter(product=product)
                    .aggregate(total=Sum("quantity"))
                    .get("total")
                )
                total = total if total is not None else Decimal("0.00")
                stock = Stock.objects.select_for_update().get_or_create(
                    product=product,
                    warehouse=comun_wh,
                    defaults={"quantity": total},
                )[0]
                stock.quantity = total
                stock.save(update_fields=["quantity"])
            bulk_product_ids = request.POST.getlist("bulk_product")
            bulk_quantities = request.POST.getlist("bulk_quantity")
            bulk_variant_ids = request.POST.getlist("bulk_variant")
            if bulk_product_ids or bulk_quantities:
                if len(bulk_product_ids) != len(bulk_quantities) or len(bulk_product_ids) != len(bulk_variant_ids):
                    messages.error(request, "La lista de productos está incompleta.")
                    return redirect("inventory_stock_list")

                bulk_items = []
                errors = []
                for index, (product_id, qty_raw, variant_id) in enumerate(
                    zip(bulk_product_ids, bulk_quantities, bulk_variant_ids), start=1
                ):
                    try:
                        quantity = Decimal(qty_raw.replace(",", "."))
                    except (InvalidOperation, ValueError):
                        errors.append(f"Línea {index}: cantidad inválida.")
                        continue
                    if quantity <= 0:
                        errors.append(f"Línea {index}: la cantidad debe ser mayor a 0.")
                        continue
                    product = Product.objects.filter(id=product_id).first()
                    if not product:
                        errors.append(f"Línea {index}: producto no encontrado.")
                        continue
                    variant = None
                    if variant_id:
                        variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                        if not variant:
                            errors.append(f"Línea {index}: variedad no encontrada.")
                            continue
                    else:
                        if ProductVariant.objects.filter(product=product).exists():
                            errors.append(f"Línea {index}: seleccioná una variedad.")
                            continue
                    bulk_items.append({"product": product, "quantity": quantity, "variant": variant})

                if errors:
                    messages.error(request, " ".join(errors))
                    return redirect("inventory_stock_list")

                try:
                    with transaction.atomic():
                        for item in bulk_items:
                            if item["variant"] is not None:
                                sync_common_with_variants(item["product"])
                                if item["variant"].quantity - item["quantity"] < 0:
                                    raise services.NegativeStockError(
                                        "No hay stock suficiente en depósito común."
                                    )
                                item["variant"].quantity = (item["variant"].quantity - item["quantity"]).quantize(
                                    Decimal("0.01"), rounding=ROUND_HALF_UP
                                )
                                item["variant"].save(update_fields=["quantity"])
                            services.register_transfer(
                                product=item["product"],
                                from_warehouse=comun_wh,
                                to_warehouse=ml_wh,
                                quantity=item["quantity"],
                                user=request.user,
                                reference="Transferencia Comun -> MercadoLibre",
                            )
                    messages.success(request, f"Transferencias registradas: {len(bulk_items)}.")
                    return redirect("inventory_stock_list")
                except services.NegativeStockError:
                    messages.error(request, "No hay stock suficiente en depósito común.")
                except services.InvalidMovementError as exc:
                    messages.error(request, str(exc))
            elif transfer_form.is_valid():
                try:
                    product = transfer_form.cleaned_data["product"]
                    variant_id = (request.POST.get("variant_id") or "").strip()
                    if variant_id:
                        variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                        if not variant:
                            messages.error(request, "Variedad no encontrada.")
                            return redirect("inventory_stock_list")
                        with transaction.atomic():
                            sync_common_with_variants(product)
                            if variant.quantity - transfer_form.cleaned_data["quantity"] < 0:
                                raise services.NegativeStockError(
                                    "No hay stock suficiente en depósito común."
                                )
                            variant.quantity = (variant.quantity - transfer_form.cleaned_data["quantity"]).quantize(
                                Decimal("0.01"), rounding=ROUND_HALF_UP
                            )
                            variant.save(update_fields=["quantity"])
                            services.register_transfer(
                                product=product,
                                from_warehouse=comun_wh,
                                to_warehouse=ml_wh,
                                quantity=transfer_form.cleaned_data["quantity"],
                                user=request.user,
                                reference="Transferencia Comun -> MercadoLibre",
                            )
                        messages.success(request, "Transferencia registrada.")
                        return redirect("inventory_stock_list")
                    else:
                        if ProductVariant.objects.filter(product=product).exists():
                            messages.error(request, "Seleccioná una variedad.")
                            return redirect("inventory_stock_list")
                    services.register_transfer(
                        product=transfer_form.cleaned_data["product"],
                        from_warehouse=comun_wh,
                        to_warehouse=ml_wh,
                        quantity=transfer_form.cleaned_data["quantity"],
                        user=request.user,
                        reference="Transferencia Comun -> MercadoLibre",
                    )
                    messages.success(request, "Transferencia registrada.")
                    return redirect("inventory_stock_list")
                except services.NegativeStockError:
                    messages.error(request, "No hay stock suficiente en depósito común.")
                except services.InvalidMovementError as exc:
                    messages.error(request, str(exc))
            else:
                error_list = []
                for field_errors in transfer_form.errors.values():
                    error_list.extend(field_errors)
                if error_list:
                    messages.error(request, " ".join(error_list))
                else:
                    messages.error(request, "Revisá los datos de la transferencia.")

    products = (
        Product.objects.order_by("sku")
        .annotate(
            comun_qty=Coalesce(
                Sum(
                    Case(
                        When(stocks__warehouse__type=comun_code, then="stocks__quantity"),
                        output_field=decimal_field,
                    )
                ),
                Value(0, output_field=decimal_field),
                output_field=decimal_field,
            ),
        )
    )
    variant_qty_map = {
        row["product_id"]: Decimal(str(row["total_qty"]))
        for row in ProductVariant.objects.values("product_id").annotate(
            total_qty=Coalesce(Sum("quantity"), Value(0, output_field=decimal_field), output_field=decimal_field)
        )
    }
    variant_data = {}
    for row in ProductVariant.objects.values("id", "product_id", "name").order_by("name", "id"):
        variant_data.setdefault(str(row["product_id"]), []).append({"id": row["id"], "name": row["name"]})
    if query:
        products = products.filter(
            Q(sku__icontains=query) | Q(name__icontains=query) | Q(group__icontains=query)
        )
    for product in products:
        if product.id in variant_qty_map:
            product.comun_qty = variant_qty_map[product.id]
            product.has_variants = True
        else:
            product.has_variants = False
        product.total_qty = product.comun_qty
    if ml_wh and show_history:
        transfer_movements = (
            StockMovement.objects.filter(
                movement_type=StockMovement.MovementType.TRANSFER,
                to_warehouse=ml_wh,
            )
            .select_related("product", "user", "from_warehouse", "to_warehouse")
            .order_by("-created_at", "-id")[:200]
        )
        grouped: dict[str, dict] = {}
        for movement in transfer_movements:
            local_date = timezone.localtime(movement.created_at).date()
            key = local_date.isoformat()
            group = grouped.get(key)
            if not group:
                group = {
                    "date": local_date,
                    "items": [],
                    "total_qty": Decimal("0.00"),
                    "users": set(),
                    "origins": set(),
                    "references": set(),
                }
                grouped[key] = group
            group["items"].append(
                {
                    "sku": movement.product.sku,
                    "name": movement.product.name,
                    "quantity": movement.quantity,
                    "origin": movement.from_warehouse.name if movement.from_warehouse else "",
                    "user": movement.user.get_full_name() or movement.user.username,
                    "reference": movement.reference,
                }
            )
            group["total_qty"] += movement.quantity
            group["users"].add(movement.user.get_full_name() or movement.user.username)
            group["origins"].add(movement.from_warehouse.name if movement.from_warehouse else "")
            group["references"].add(movement.reference or "")
        transfer_history = []
        for key, group in grouped.items():
            users = sorted(u for u in group["users"] if u)
            origins = sorted(o for o in group["origins"] if o)
            references = sorted(r for r in group["references"] if r)
            transfer_history.append(
                {
                    "date": group["date"],
                    "items": group["items"],
                    "total_qty": group["total_qty"],
                    "user_label": users[0] if len(users) == 1 else "Varios",
                    "origin_label": origins[0] if len(origins) == 1 else "Varios",
                    "reference_label": references[0] if len(references) == 1 else "Varios",
                }
            )
    else:
        transfer_history = []
    return render(
        request,
        "inventory/stock_list.html",
        {
            "products": products,
            "transfer_form": transfer_form,
            "can_transfer": comun_wh is not None and ml_wh is not None,
            "query": query,
            "variant_data": variant_data,
            "transfer_history": transfer_history,
            "show_history": show_history,
        },
    )


@login_required
def product_variants(request, product_id: int):
    product = get_object_or_404(Product, pk=product_id)
    VariantFormSet = formset_factory(ProductVariantRowForm, extra=0)
    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "add_variant":
            add_form = ProductVariantForm(request.POST)
            if add_form.is_valid():
                ProductVariant.objects.create(
                    product=product,
                    name=add_form.cleaned_data["name"].strip(),
                    quantity=add_form.cleaned_data["quantity"],
                )
                messages.success(request, "Variedad agregada.")
                return redirect("inventory_product_variants", product_id=product.id)
            messages.error(request, "Revisá los datos de la variedad.")
        elif action == "update_variants":
            formset = VariantFormSet(request.POST)
            if formset.is_valid():
                for form in formset:
                    variant_id = form.cleaned_data["variant_id"]
                    variant = ProductVariant.objects.filter(id=variant_id, product=product).first()
                    if not variant:
                        continue
                    if form.cleaned_data.get("delete"):
                        variant.delete()
                        continue
                    variant.name = form.cleaned_data["name"].strip()
                    variant.quantity = form.cleaned_data["quantity"]
                    variant.save(update_fields=["name", "quantity"])
                messages.success(request, "Variedades actualizadas.")
                return redirect("inventory_product_variants", product_id=product.id)
            messages.error(request, "Revisá los datos de las variedades.")
    variants = ProductVariant.objects.filter(product=product).order_by("name", "id")
    formset = VariantFormSet(
        initial=[
            {"variant_id": v.id, "name": v.name, "quantity": v.quantity, "delete": False}
            for v in variants
        ]
    )
    add_form = ProductVariantForm()
    return render(
        request,
        "inventory/product_variants.html",
        {
            "product": product,
            "formset": formset,
            "add_form": add_form,
        },
    )


def _koda_allowed(user) -> bool:
    if not user or not user.is_authenticated:
        return False
    full_name = (user.get_full_name() or "").strip().lower()
    username = (getattr(user, "username", "") or "").strip().lower()
    email = (getattr(user, "email", "") or "").strip().lower()
    if full_name == "alejo salmeron":
        return True
    if username in {"alejo", "alejosalmeron", "alejo.salmeron"}:
        return True
    if email in {"alejo@stylmoda.site", "alejosalmeron@gmail.com"}:
        return True
    return bool(getattr(user, "is_superuser", False))


def _koda_system_prompt() -> str:
    return (
        "Sos Koda, el asistente del ERP. Respondé SOLO en JSON válido con las claves: "
        "reply (string), actions (array), needs_confirmation (boolean). "
        "Cada acción debe ser un objeto con {type, data}. "
        "Tu objetivo es ayudar con precisión, evitando suposiciones. "
        "Nunca inventes cifras o resultados. Si el usuario pide reportes o totales "
        "y no hay datos explícitos en el mensaje, pedí rango de fechas. "
        "Solo preguntá por filtros (depósito, canal, cliente) si el usuario los menciona "
        "o si explícitamente pide un filtro. Dejá actions vacío y needs_confirmation=false. "
        "Para consultas informativas (márgenes, ventas, stock), NO uses actions. "
        "Siempre respondé algo útil en reply, aunque no haya acciones. "
        "Si falta información para ejecutar una acción, hacé preguntas concretas. "
        "Si hay una imagen adjunta, extraé los datos relevantes de la imagen. "
        "Nunca digas que ejecutaste una acción si no envías actions. "
        "Si el usuario pide ejecutar/registrar/crear/transferir, devolvé actions y needs_confirmation=true. "
        "Si el usuario pide confirmación de lo que entendiste, resumí y pedí OK. "
        "Respondé en español rioplatense, directo y claro. "
        "Acciones permitidas:\n"
        "- create_product: {name, sku?, group?, avg_cost?, vat_percent?, price_consumer?, price_barber?, price_distributor?}\n"
        "- add_stock_comun: {items:[{product, quantity, unit_cost?, variant?}]}\n"
        "- transfer_to_ml: {items:[{product, quantity, variant?}]}\n"
        "- register_sale: {warehouse, audience?, customer?, items:[{product, quantity, unit_price?, vat_percent?}]}\n"
        "- update_sale: {sale_id? or invoice_number?, sale_date?, total?, ml_commission_total?, ml_tax_total?}\n"
        "- register_purchase: {warehouse?, supplier, reference?, items:[{product, quantity, unit_cost?, vat_percent?}], has_invoice_image?}\n"
        "Para compras a proveedor, extraé items con {product: descripcion, quantity} y dejá unit_cost vacío.\n"
        "El backend completará el costo con el precio de lista/costo del ERP.\n"
        "Usá identificadores de producto por SKU si es posible, sino por nombre exacto. "
        "Si el producto tiene variedades, pedí o incluí la variedad."
    )


def _koda_format_amount(value: Decimal, currency: bool = True) -> str:
    try:
        number = value if isinstance(value, Decimal) else Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        number = Decimal("0.00")
    number = number.quantize(Decimal("1.00"), rounding=ROUND_HALF_UP)
    formatted = f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"${formatted}" if currency else formatted


def _koda_extract_percent(message: str) -> Decimal | None:
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*(?:%|por\s*ciento)", message, re.IGNORECASE)
    if not match:
        return None
    raw = match.group(1).replace(",", ".")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _koda_extract_warehouse(message: str) -> str | None:
    lowered = message.lower()
    if "mercadolibre" in lowered or "mercado libre" in lowered or re.search(r"\bml\b", lowered):
        return Warehouse.WarehouseType.MERCADOLIBRE
    if "comun" in lowered or "común" in lowered:
        return Warehouse.WarehouseType.COMUN
    return None


def _koda_parse_date_range(message: str) -> tuple[datetime, datetime, datetime.date, datetime.date] | None:
    lowered = message.lower()
    today = timezone.localdate()
    if "hoy" in lowered:
        start_date = end_date = today
    elif "ayer" in lowered:
        day = today - timedelta(days=1)
        start_date = end_date = day
    elif "este mes" in lowered or "mes actual" in lowered:
        start_date = today.replace(day=1)
        end_date = today
    elif "mes pasado" in lowered or "mes anterior" in lowered:
        first_this_month = today.replace(day=1)
        last_prev_month = first_this_month - timedelta(days=1)
        start_date = last_prev_month.replace(day=1)
        end_date = last_prev_month
    else:
        month_map = {
            "enero": 1,
            "febrero": 2,
            "marzo": 3,
            "abril": 4,
            "mayo": 5,
            "junio": 6,
            "julio": 7,
            "agosto": 8,
            "septiembre": 9,
            "setiembre": 9,
            "octubre": 10,
            "noviembre": 11,
            "diciembre": 12,
        }
        month_match = re.search(
            r"\b(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|setiembre|octubre|noviembre|diciembre)\b(?:\s+de)?\s*(\d{4})?",
            lowered,
        )
        if month_match:
            month = month_map[month_match.group(1)]
            year_raw = month_match.group(2)
            if year_raw:
                year = int(year_raw)
            else:
                year = today.year
                if month > today.month:
                    year -= 1
            last_day = calendar.monthrange(year, month)[1]
            start_date = datetime(year, month, 1).date()
            end_date = datetime(year, month, last_day).date()
        else:
            matches = list(re.finditer(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?", lowered))
            if not matches:
                return None
            def parse_match(match, fallback_year: int) -> datetime.date | None:
                day = int(match.group(1))
                month = int(match.group(2))
                year_raw = match.group(3)
                if year_raw:
                    year = int(year_raw)
                    if year < 100:
                        year += 2000
                else:
                    year = fallback_year
                try:
                    return datetime(year, month, day).date()
                except ValueError:
                    return None

            first = matches[0]
            fallback_year = today.year
            first_date = parse_match(first, fallback_year)
            if not first_date:
                return None
            if len(matches) >= 2:
                second = matches[1]
                second_date = parse_match(second, first_date.year)
                if not second_date:
                    return None
                start_date, end_date = first_date, second_date
            else:
                start_date = end_date = first_date
    if end_date < start_date:
        start_date, end_date = end_date, start_date
    start_dt = timezone.make_aware(datetime.combine(start_date, time.min))
    end_dt = timezone.make_aware(datetime.combine(end_date, time.max))
    return start_dt, end_dt, start_date, end_date


def _koda_safe_eval(expression: str) -> Decimal:
    import ast
    import operator

    bin_ops = {
        ast.Add: operator.add,
        ast.Sub: operator.sub,
        ast.Mult: operator.mul,
        ast.Div: operator.truediv,
    }
    unary_ops = {ast.UAdd: operator.pos, ast.USub: operator.neg}

    def eval_node(node):
        if isinstance(node, ast.Expression):
            return eval_node(node.body)
        if isinstance(node, ast.BinOp) and type(node.op) in bin_ops:
            return bin_ops[type(node.op)](eval_node(node.left), eval_node(node.right))
        if isinstance(node, ast.UnaryOp) and type(node.op) in unary_ops:
            return unary_ops[type(node.op)](eval_node(node.operand))
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return Decimal(str(node.value))
        raise ValueError("Expresión inválida")

    tree = ast.parse(expression, mode="eval")
    return eval_node(tree)


def _koda_try_math(message: str) -> str | None:
    lowered = message.lower()
    if any(word in lowered for word in ("venta", "ventas", "compra", "compras", "ganancia", "margen", "stock", "producto", "cliente", "proveedor", "iva", "impuesto")):
        return None
    if not re.search(r"\d", message):
        return None
    expr = re.sub(r"[^0-9\.\,\+\-\*\/\(\)\s]", "", message)
    if not re.search(r"[\+\-\*\/]", expr):
        return None
    expr = expr.replace(",", ".")
    try:
        result = _koda_safe_eval(expr)
        return f"Resultado: {_koda_format_amount(result, currency=False)}"
    except Exception:
        return None


def _koda_sales_profit(sales_qs) -> Decimal:
    profit_total = Decimal("0.00")
    for sale in sales_qs:
        cost_total = Decimal("0.00")
        for item in sale.items.all():
            cost_unit = item.cost_unit if item.cost_unit and item.cost_unit > 0 else item.product.cost_with_vat()
            cost_total += item.quantity * cost_unit
        if sale.warehouse.type == Warehouse.WarehouseType.MERCADOLIBRE:
            net_total = (sale.total or Decimal("0.00")) - (sale.ml_commission_total or Decimal("0.00")) - (
                sale.ml_tax_total or Decimal("0.00")
            )
            profit_total += net_total - cost_total
        else:
            profit_total += (sale.total or Decimal("0.00")) - cost_total
    return profit_total


def _koda_try_local_response(message: str) -> str | None:
    if not message:
        return None
    lowered = message.lower()
    percent = _koda_extract_percent(lowered)
    wants_profit = any(word in lowered for word in ("ganancia", "margen", "utilidad", "gané", "gane", "ganado", "ganar")) or bool(re.search(r"\bgan", lowered))
    wants_sales = (
        "venta" in lowered
        or "ventas" in lowered
        or "monto de venta" in lowered
        or "monto venta" in lowered
        or "ingreso" in lowered
        or "ingresos" in lowered
        or bool(re.search(r"\bfactur", lowered))
    )
    wants_purchases = "compra" in lowered or "compras" in lowered
    wants_tax = "impuesto" in lowered or "iva" in lowered

    if percent and not (wants_sales or wants_purchases):
        return "¿Querés ese % sobre compras, ventas o ambos?"

    if wants_profit or wants_sales or wants_purchases or percent or wants_tax:
        date_range = _koda_parse_date_range(lowered)
        if not date_range:
            return "¿De qué fechas? Indicame desde y hasta (ej: 01/01/2026 al 11/01/2026)."
        start_dt, end_dt, start_date, end_date = date_range
        warehouse = _koda_extract_warehouse(lowered)
        range_label = f"{start_date.strftime('%d/%m/%Y')} al {end_date.strftime('%d/%m/%Y')}"
        reply_lines = [f"Rango: {range_label}."]
        if warehouse == Warehouse.WarehouseType.MERCADOLIBRE:
            reply_lines.append("Depósito: MercadoLibre.")
        elif warehouse == Warehouse.WarehouseType.COMUN:
            reply_lines.append("Depósito: Común.")

        sales_qs = Sale.objects.filter(created_at__gte=start_dt, created_at__lte=end_dt).select_related("warehouse").prefetch_related("items__product")
        sale_item_qs = SaleItem.objects.filter(sale__created_at__gte=start_dt, sale__created_at__lte=end_dt)
        purchase_qs = Purchase.objects.filter(created_at__gte=start_dt, created_at__lte=end_dt)
        tax_qs = TaxExpense.objects.filter(paid_at__gte=start_date, paid_at__lte=end_date)
        if warehouse:
            sales_qs = sales_qs.filter(warehouse__type=warehouse)
            sale_item_qs = sale_item_qs.filter(sale__warehouse__type=warehouse)
            purchase_qs = purchase_qs.filter(warehouse__type=warehouse)

        sales_total = sale_item_qs.aggregate(total=Sum("line_total")).get("total") or Decimal("0.00")
        purchases_total = purchase_qs.aggregate(total=Sum("total")).get("total") or Decimal("0.00")
        tax_total = tax_qs.aggregate(total=Sum("amount")).get("total") or Decimal("0.00")

        if wants_sales:
            reply_lines.append(f"Total ventas: {_koda_format_amount(sales_total)}.")
        if wants_purchases:
            reply_lines.append(f"Total compras: {_koda_format_amount(purchases_total)}.")
        if wants_profit:
            profit_sales = _koda_sales_profit(sales_qs)
            reply_lines.append(f"Ganancia por ventas (neto ML - costo): {_koda_format_amount(profit_sales)}.")
            if tax_total:
                net_profit = profit_sales - tax_total
                reply_lines.append(f"Impuestos registrados: {_koda_format_amount(tax_total)}.")
                reply_lines.append(f"Ganancia neta después de impuestos: {_koda_format_amount(net_profit)}.")
            else:
                reply_lines.append("Impuestos registrados: $0,00.")
        if wants_tax and not wants_profit:
            reply_lines.append(f"Impuestos registrados: {_koda_format_amount(tax_total)}.")
        if percent is not None and (wants_sales or wants_purchases):
            if wants_sales:
                sales_pct = (sales_total * percent / Decimal("100.00")).quantize(Decimal("1.00"))
                reply_lines.append(f"{percent}% de ventas: {_koda_format_amount(sales_pct)}.")
            if wants_purchases:
                purchases_pct = (purchases_total * percent / Decimal("100.00")).quantize(Decimal("1.00"))
                reply_lines.append(f"{percent}% de compras: {_koda_format_amount(purchases_pct)}.")
        return " ".join(reply_lines)

    return _koda_try_math(message)


def _koda_call_openai(messages, image_data_url: str | None = None) -> dict:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return {"reply": "Falta configurar OPENAI_API_KEY.", "actions": [], "needs_confirmation": False}

    user_content = [{"type": "text", "text": messages[-1]["content"]}]
    if image_data_url:
        user_content.append({"type": "image_url", "image_url": {"url": image_data_url}})

    payload_messages = [{"role": "system", "content": _koda_system_prompt()}]
    payload_messages.extend(messages[:-1])
    payload_messages.append({"role": "user", "content": user_content})

    payload = {
        "model": "gpt-4.1",
        "temperature": 0.2,
        "messages": payload_messages,
    }
    request = Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=30) as response:
            data = json.loads(response.read().decode("utf-8"))
        content = data["choices"][0]["message"]["content"]
    except Exception as exc:
        return {"reply": f"No pude contactar a OpenAI: {exc}", "actions": [], "needs_confirmation": False}

    try:
        parsed = json.loads(content)
        return {
            "reply": parsed.get("reply", ""),
            "actions": parsed.get("actions", []) or [],
            "needs_confirmation": bool(parsed.get("needs_confirmation")),
        }
    except Exception:
        return {"reply": content, "actions": [], "needs_confirmation": False}


def _koda_extract_pdf_text(path: str) -> tuple[str | None, str | None]:
    try:
        from pdfminer.high_level import extract_text
    except Exception:
        return None, "Falta pdfminer.six para leer PDFs."
    try:
        text = extract_text(path) or ""
        text = text.strip()
        return text, None
    except Exception as exc:
        return None, f"No pude leer el PDF: {exc}"


def _koda_actions_summary(actions: list[dict]) -> str:
    summaries = []
    for action in actions:
        action_type = action.get("type")
        data = action.get("data", {})
        if action_type == "create_product":
            summaries.append(f"Crear producto: {data.get('sku') or 'Sin SKU'} - {data.get('name')}")
        elif action_type == "add_stock_comun":
            items = data.get("items") or []
            summaries.append(f"Agregar stock Común: {len(items)} ítems")
        elif action_type == "transfer_to_ml":
            items = data.get("items") or []
            summaries.append(f"Transferir a MercadoLibre: {len(items)} ítems")
        elif action_type == "register_sale":
            items = data.get("items") or []
            summaries.append(f"Registrar venta: {len(items)} ítems")
        elif action_type == "update_sale":
            summaries.append("Actualizar venta")
        elif action_type == "register_purchase":
            items = data.get("items") or []
            summaries.append(f"Registrar compra: {len(items)} ítems")
        else:
            summaries.append(f"Acción: {action_type}")
    return " | ".join(summaries)


def _koda_resolve_product(identifier: str) -> Product | None:
    if not identifier:
        return None
    raw = identifier.strip()
    product = Product.objects.filter(sku__iexact=raw).first()
    if product:
        return product
    product = Product.objects.filter(name__iexact=raw).first()
    if product:
        return product
    normalized = _normalize_header(raw)
    tokens = [token for token in normalized.split() if len(token) > 2]
    if not tokens:
        return None
    candidates = Product.objects.filter(name__icontains=tokens[0])[:50]
    best = None
    best_score = 0
    for candidate in candidates:
        name_norm = _normalize_header(candidate.name)
        score = sum(1 for token in tokens if token in name_norm)
        if score > best_score:
            best_score = score
            best = candidate
    return best


def _koda_resolve_variant(product: Product, variant_name: str | None) -> ProductVariant | None:
    if not variant_name:
        return None
    return ProductVariant.objects.filter(product=product, name__iexact=variant_name.strip()).first()


def _koda_resolve_supplier(name: str | None) -> Supplier | None:
    if not name:
        return None
    raw = name.strip()
    supplier = Supplier.objects.filter(name__iexact=raw).first()
    if supplier:
        return supplier
    return Supplier.objects.filter(name__icontains=raw).first()


def _koda_resolve_sale(data: dict) -> Sale | None:
    sale_id = data.get("sale_id")
    if sale_id:
        return Sale.objects.filter(id=sale_id).first()
    invoice_number = (data.get("invoice_number") or "").strip()
    if invoice_number:
        parts = invoice_number.split("-")
        tail = parts[-1] if parts else invoice_number
        if tail.isdigit():
            return Sale.objects.filter(id=int(tail)).first()
    return None


def _koda_sync_common_with_variants(product: Product, comun_wh: Warehouse):
    total = (
        ProductVariant.objects.filter(product=product)
        .aggregate(total=Sum("quantity"))
        .get("total")
    )
    total = total if total is not None else Decimal("0.00")
    stock = Stock.objects.select_for_update().get_or_create(
        product=product,
        warehouse=comun_wh,
        defaults={"quantity": total},
    )[0]
    stock.quantity = total
    stock.save(update_fields=["quantity"])


def _koda_execute_actions(actions: list[dict], user, invoice_path: str | None) -> list[str]:
    results = []
    comun_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.COMUN).first()
    ml_wh = Warehouse.objects.filter(type=Warehouse.WarehouseType.MERCADOLIBRE).first()
    for action in actions:
        action_type = action.get("type")
        data = action.get("data", {})
        if action_type == "create_product":
            name = (data.get("name") or "").strip()
            if not name:
                raise ValueError("Falta el nombre del producto.")
            sku = (data.get("sku") or "").strip()
            group = (data.get("group") or "").strip()
            if not sku:
                prefix = _sku_prefix(group or "", name)
                existing = Product.objects.filter(sku__startswith=prefix).values_list("sku", flat=True)
                max_suffix = 0
                for existing_sku in existing:
                    suffix = existing_sku[len(prefix):]
                    if suffix.isdigit():
                        max_suffix = max(max_suffix, int(suffix))
                sku = f"{prefix}{max_suffix + 1:04d}"
            product = Product.objects.create(
                sku=sku,
                name=name,
                group=group,
                avg_cost=Decimal(str(data.get("avg_cost") or "0.00")),
                vat_percent=Decimal(str(data.get("vat_percent") or "0.00")),
                price_consumer=Decimal(str(data.get("price_consumer") or "0.00")),
                price_barber=Decimal(str(data.get("price_barber") or "0.00")),
                price_distributor=Decimal(str(data.get("price_distributor") or "0.00")),
            )
            results.append(f"Producto creado: {product.sku} - {product.name}")
        elif action_type == "add_stock_comun":
            if not comun_wh:
                raise ValueError("No está configurado el depósito Común.")
            items = data.get("items") or []
            if not items:
                raise ValueError("No hay ítems para agregar stock.")
            with transaction.atomic():
                for item in items:
                    product = _koda_resolve_product(item.get("product", ""))
                    if not product:
                        raise ValueError("Producto no encontrado.")
                    quantity = Decimal(str(item.get("quantity") or "0"))
                    if quantity <= 0:
                        raise ValueError("Cantidad inválida.")
                    variant = _koda_resolve_variant(product, item.get("variant"))
                    if not variant and ProductVariant.objects.filter(product=product).exists():
                        raise ValueError(f"El producto {product.sku} requiere variedad.")
                    unit_cost = item.get("unit_cost")
                    services.register_adjustment(
                        product=product,
                        warehouse=comun_wh,
                        quantity=quantity,
                        user=user,
                        unit_cost=Decimal(str(unit_cost)) if unit_cost is not None else None,
                        reference="Koda: ajuste de stock Común",
                    )
                    if variant:
                        variant.quantity = (variant.quantity + quantity).quantize(Decimal("0.01"))
                        variant.save(update_fields=["quantity"])
                        _koda_sync_common_with_variants(product, comun_wh)
            results.append(f"Stock Común actualizado: {len(items)} ítems.")
        elif action_type == "transfer_to_ml":
            if not comun_wh or not ml_wh:
                raise ValueError("Faltan depósitos configurados para transferencias.")
            items = data.get("items") or []
            if not items:
                raise ValueError("No hay ítems para transferir.")
            with transaction.atomic():
                for item in items:
                    product = _koda_resolve_product(item.get("product", ""))
                    if not product:
                        raise ValueError("Producto no encontrado.")
                    quantity = Decimal(str(item.get("quantity") or "0"))
                    if quantity <= 0:
                        raise ValueError("Cantidad inválida.")
                    variant = _koda_resolve_variant(product, item.get("variant"))
                    if not variant and ProductVariant.objects.filter(product=product).exists():
                        raise ValueError(f"El producto {product.sku} requiere variedad.")
                    if variant:
                        _koda_sync_common_with_variants(product, comun_wh)
                        if variant.quantity - quantity < 0:
                            raise services.NegativeStockError("Stock insuficiente en variedad.")
                        variant.quantity = (variant.quantity - quantity).quantize(Decimal("0.01"))
                        variant.save(update_fields=["quantity"])
                        _koda_sync_common_with_variants(product, comun_wh)
                    services.register_transfer(
                        product=product,
                        from_warehouse=comun_wh,
                        to_warehouse=ml_wh,
                        quantity=quantity,
                        user=user,
                        reference="Koda: transferencia Común -> MercadoLibre",
                    )
            results.append(f"Transferencias a ML: {len(items)} ítems.")
        elif action_type == "register_sale":
            warehouse_label = (data.get("warehouse") or "COMUN").strip().lower()
            warehouse_code = {
                "comun": Warehouse.WarehouseType.COMUN,
                "común": Warehouse.WarehouseType.COMUN,
                "mercadolibre": Warehouse.WarehouseType.MERCADOLIBRE,
                "mercado libre": Warehouse.WarehouseType.MERCADOLIBRE,
                "ml": Warehouse.WarehouseType.MERCADOLIBRE,
            }.get(warehouse_label, warehouse_label.upper())
            warehouse = Warehouse.objects.filter(type=warehouse_code).first()
            if not warehouse:
                raise ValueError("Depósito no encontrado.")
            customer = None
            if data.get("customer"):
                customer = Customer.objects.filter(name__iexact=data.get("customer")).first()
            audience = data.get("audience") or Customer.Audience.CONSUMER
            items = data.get("items") or []
            if not items:
                raise ValueError("No hay ítems para la venta.")
            with transaction.atomic():
                sale = Sale.objects.create(
                    customer=customer,
                    warehouse=warehouse,
                    audience=audience,
                    total=Decimal("0.00"),
                    discount_total=Decimal("0.00"),
                    reference="Koda: venta",
                    user=user,
                )
                total = Decimal("0.00")
                for item in items:
                    product = _koda_resolve_product(item.get("product", ""))
                    if not product:
                        raise ValueError("Producto no encontrado.")
                    quantity = Decimal(str(item.get("quantity") or "0"))
                    if quantity <= 0:
                        raise ValueError("Cantidad inválida.")
                    unit_price = item.get("unit_price")
                    if unit_price is None:
                        if audience == Customer.Audience.BARBER:
                            unit_price = product.barber_price
                        elif audience == Customer.Audience.DISTRIBUTOR:
                            unit_price = product.distributor_price
                        else:
                            unit_price = product.consumer_price
                    unit_price = Decimal(str(unit_price))
                    line_total = (unit_price * quantity).quantize(Decimal("0.01"))
                    SaleItem.objects.create(
                        sale=sale,
                        product=product,
                        quantity=quantity,
                        unit_price=unit_price,
                        cost_unit=product.last_purchase_cost(),
                        discount_percent=Decimal("0.00"),
                        final_unit_price=unit_price,
                        line_total=line_total,
                        vat_percent=Decimal(str(item.get("vat_percent") or "0.00")),
                    )
                    total += line_total
                    if warehouse.type != Warehouse.WarehouseType.MERCADOLIBRE:
                        services.register_exit(
                            product=product,
                            warehouse=warehouse,
                            quantity=quantity,
                            user=user,
                            reference=f"Koda: venta #{sale.id}",
                            sale_price=unit_price,
                            vat_percent=Decimal(str(item.get("vat_percent") or "0.00")),
                            sale=sale,
                        )
                sale.total = total
                sale.save(update_fields=["total"])
            results.append(f"Venta registrada #{sale.id}.")
        elif action_type == "update_sale":
            sale = _koda_resolve_sale(data)
            if not sale:
                raise ValueError("Venta no encontrada.")
            update_fields = []
            sale_date = data.get("sale_date")
            if sale_date:
                try:
                    created_at = datetime.fromisoformat(str(sale_date))
                except Exception:
                    created_at = datetime.combine(str(sale_date), time(12, 0))
                if timezone.is_naive(created_at):
                    created_at = timezone.make_aware(created_at)
                sale.created_at = created_at
                update_fields.append("created_at")
            if data.get("total") is not None:
                sale.total = Decimal(str(data.get("total")))
                update_fields.append("total")
            if data.get("ml_commission_total") is not None:
                sale.ml_commission_total = Decimal(str(data.get("ml_commission_total")))
                update_fields.append("ml_commission_total")
            if data.get("ml_tax_total") is not None:
                sale.ml_tax_total = Decimal(str(data.get("ml_tax_total")))
                update_fields.append("ml_tax_total")
            if update_fields:
                sale.save(update_fields=update_fields)
            results.append(f"Venta actualizada #{sale.id}.")
        elif action_type == "register_purchase":
            warehouse_label = (data.get("warehouse") or "COMUN").strip().lower()
            warehouse_code = {
                "comun": Warehouse.WarehouseType.COMUN,
                "común": Warehouse.WarehouseType.COMUN,
                "mercadolibre": Warehouse.WarehouseType.MERCADOLIBRE,
                "mercado libre": Warehouse.WarehouseType.MERCADOLIBRE,
                "ml": Warehouse.WarehouseType.MERCADOLIBRE,
            }.get(warehouse_label, warehouse_label.upper())
            warehouse = Warehouse.objects.filter(type=warehouse_code).first()
            if not warehouse:
                raise ValueError("Depósito no encontrado.")
            supplier_name = (data.get("supplier") or "").strip()
            if not supplier_name:
                raise ValueError("Falta proveedor.")
            supplier = _koda_resolve_supplier(supplier_name)
            if not supplier:
                supplier = Supplier.objects.create(name=supplier_name)
            items = data.get("items") or []
            if not items:
                raise ValueError("No hay ítems para la compra.")
            with transaction.atomic():
                purchase = Purchase.objects.create(
                    supplier=supplier,
                    warehouse=warehouse,
                    reference=data.get("reference") or "Koda: compra",
                    user=user,
                )
                total = Decimal("0.00")
                for item in items:
                    product = _koda_resolve_product(item.get("product", ""))
                    if not product:
                        raise ValueError("Producto no encontrado.")
                    quantity = Decimal(str(item.get("quantity") or "0"))
                    if quantity <= 0:
                        raise ValueError("Cantidad inválida.")
                    unit_cost_raw = item.get("unit_cost")
                    if unit_cost_raw is None:
                        unit_cost = product.cost_with_vat()
                    else:
                        unit_cost = Decimal(str(unit_cost_raw))
                    vat_percent = Decimal(str(item.get("vat_percent") or "0.00"))
                    PurchaseItem.objects.create(
                        purchase=purchase,
                        product=product,
                        quantity=quantity,
                        unit_cost=unit_cost,
                        vat_percent=vat_percent,
                    )
                    services.register_entry(
                        product=product,
                        warehouse=warehouse,
                        quantity=quantity,
                        unit_cost=unit_cost,
                        user=user,
                        reference=f"Koda: compra #{purchase.id}",
                    )
                    total += (quantity * unit_cost).quantize(Decimal("0.01"))
                purchase.total = total
                if invoice_path:
                    with open(invoice_path, "rb") as handle:
                        purchase.invoice_image.save(
                            os.path.basename(invoice_path),
                            ContentFile(handle.read()),
                            save=False,
                        )
                purchase.save()
            results.append(f"Compra registrada #{purchase.id}.")
        else:
            raise ValueError(f"Acción desconocida: {action_type}")
    return results


@login_required
@require_http_methods(["POST"])
def koda_chat(request):
    if not _koda_allowed(request.user):
        return JsonResponse({"reply": "No tenés permisos para usar Koda.", "actions": []})

    message = ""
    image_file = None
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        message = (request.POST.get("message") or "").strip()
        image_file = request.FILES.get("image")
    else:
        try:
            payload = json.loads(request.body.decode("utf-8"))
            message = (payload.get("message") or "").strip()
        except Exception:
            message = ""

    if not message and not image_file:
        return JsonResponse({"reply": "Decime qué necesitás.", "actions": []})

    if not image_file:
        local_reply = _koda_try_local_response(message)
        if local_reply:
            history = request.session.get("koda_history", [])
            history = history[-8:]
            history.append({"role": "user", "content": message})
            history.append({"role": "assistant", "content": local_reply})
            request.session["koda_history"] = history[-10:]
            return JsonResponse({"reply": local_reply, "needs_confirmation": False, "summary": ""})

    image_data_url = None
    pending_file_path = None
    if image_file:
        raw = image_file.read()
        pending_dir = os.path.join(str(settings.MEDIA_ROOT), "koda_pending")
        os.makedirs(pending_dir, exist_ok=True)
        filename = f"{secrets.token_hex(8)}_{image_file.name}"
        pending_file_path = os.path.join(pending_dir, filename)
        with open(pending_file_path, "wb") as handle:
            handle.write(raw)

        if image_file.content_type and image_file.content_type.startswith("image/"):
            encoded = base64.b64encode(raw).decode("utf-8")
            image_data_url = f"data:{image_file.content_type};base64,{encoded}"
            if not message:
                message = "Analizá la imagen adjunta y respondé."
            elif "imagen" not in message.lower():
                message = f"{message}\n\nUsá la imagen adjunta para extraer los datos."
        elif image_file.content_type == "application/pdf":
            pdf_text, pdf_error = _koda_extract_pdf_text(pending_file_path)
            if pdf_error:
                return JsonResponse({"reply": pdf_error, "actions": []})
            if not pdf_text:
                return JsonResponse(
                    {"reply": "El PDF no tiene texto legible. Subí una imagen o pegá el contenido.", "actions": []}
                )
            message = f"{message}\n\nContenido del PDF:\n{pdf_text}".strip()

    history = request.session.get("koda_history", [])
    history = history[-8:]
    messages = history + [{"role": "user", "content": message}]
    result = _koda_call_openai(messages, image_data_url=image_data_url)

    actions = result.get("actions") or []
    needs_confirmation = bool(result.get("needs_confirmation")) and bool(actions)
    reply = result.get("reply") or ""
    lowered = reply.lower()
    if "no tengo acceso" in lowered or "no tengo acceso directo" in lowered or "no puedo acceder" in lowered:
        reply = "Puedo gestionarlo. Confirmame los datos exactos para proceder."
    if actions and not needs_confirmation:
        needs_confirmation = True
    if not actions and any(word in lowered for word in ("registr", "cre", "transfer", "actualic", "listo", "confirm")):
        reply = "Puedo hacerlo, pero necesito confirmación y los datos exactos para ejecutar."

    if needs_confirmation:
        request.session["koda_pending"] = {
            "actions": actions,
            "image_path": pending_file_path,
        }
    else:
        if pending_file_path:
            os.remove(pending_file_path)

    history.append({"role": "user", "content": message})
    history.append({"role": "assistant", "content": reply})
    request.session["koda_history"] = history[-10:]

    return JsonResponse(
        {
            "reply": reply,
            "needs_confirmation": needs_confirmation,
            "summary": _koda_actions_summary(actions) if needs_confirmation else "",
        }
    )


@login_required
@require_http_methods(["POST"])
def koda_confirm(request):
    if not _koda_allowed(request.user):
        return JsonResponse({"reply": "No tenés permisos para usar Koda."}, status=403)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}
    decision = payload.get("decision")
    pending = request.session.get("koda_pending")
    if not pending:
        return JsonResponse({"reply": "No hay acciones pendientes."})

    image_path = pending.get("image_path")
    if decision == "cancel":
        if image_path and os.path.exists(image_path):
            os.remove(image_path)
        request.session.pop("koda_pending", None)
        return JsonResponse({"reply": "Acción cancelada."})

    try:
        results = _koda_execute_actions(pending.get("actions") or [], request.user, image_path)
        reply = "Listo. " + " ".join(results)
    except Exception as exc:
        reply = f"No pude ejecutar la acción: {exc}"
    finally:
        if image_path and os.path.exists(image_path):
            os.remove(image_path)
        request.session.pop("koda_pending", None)

    return JsonResponse({"reply": reply})


@login_required
def product_prices(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_prices":
            product_id = request.POST.get("product_id")
            product = Product.objects.filter(id=product_id).first()
            if not product:
                messages.error(request, "Producto no encontrado.")
                return redirect("inventory_product_prices")
            price_consumer = _parse_decimal(request.POST.get("price_consumer"))
            price_barber = _parse_decimal(request.POST.get("price_barber"))
            price_distributor = _parse_decimal(request.POST.get("price_distributor"))
            product.price_consumer = price_consumer
            product.price_barber = price_barber
            product.price_distributor = price_distributor
            product.save(update_fields=["price_consumer", "price_barber", "price_distributor"])
            messages.success(request, "Precios actualizados.")
            return redirect("inventory_product_prices")
    products = Product.objects.order_by("sku")
    group_options = (
        Product.objects.exclude(group="")
        .exclude(group__isnull=True)
        .values_list("group", flat=True)
        .distinct()
        .order_by("group")
    )
    return render(
        request,
        "inventory/product_prices.html",
        {"products": products, "group_options": group_options},
    )


@login_required
@require_http_methods(["GET", "POST"])
def product_costs(request):
    ProductCostFormSet = formset_factory(ProductCostRowForm, extra=0)
    products_qs = Product.objects.select_related("default_supplier").order_by("sku")
    products = list(products_qs)
    group_options = (
        Product.objects.exclude(group="")
        .exclude(group__isnull=True)
        .values_list("group", flat=True)
        .distinct()
        .order_by("group")
    )
    product_map = {product.id: product for product in products}
    product_form = ProductForm()
    bulk_form = ProductBulkUpdateForm()
    initial = [
        {
            "product_id": product.id,
            "name": product.name,
            "group": product.group,
            "supplier": product.default_supplier,
            "avg_cost": product.avg_cost,
            "vat_percent": product.vat_percent,
        }
        for product in products
    ]
    formset = ProductCostFormSet(initial=initial)
    if request.method == "POST":
        action = request.POST.get("action", "update_costs")
        if action == "create_product":
            product_form = ProductForm(request.POST)
            if product_form.is_valid():
                product = product_form.save(commit=False)
                if not product.sku:
                    prefix = _sku_prefix(product.group or "", product.name or "")
                    existing = Product.objects.filter(sku__startswith=prefix).values_list("sku", flat=True)
                    max_suffix = 0
                    for sku in existing:
                        suffix = sku[len(prefix):]
                        if suffix.isdigit():
                            max_suffix = max(max_suffix, int(suffix))
                    product.sku = f"{prefix}{max_suffix + 1:04d}"
                product.save()
                messages.success(request, "Producto creado.")
                return redirect("inventory_product_costs")
            messages.error(request, "Revisá los datos del producto.")
        elif action == "bulk_update_margins":
            def parse_optional_decimal(raw: str | None) -> Decimal | None:
                if raw is None:
                    return None
                value = str(raw).strip().replace(",", ".")
                if value == "":
                    return None
                try:
                    return Decimal(value).quantize(Decimal("0.01"))
                except Exception:
                    return None

            margin_consumer = parse_optional_decimal(request.POST.get("margin_consumer"))
            margin_barber = parse_optional_decimal(request.POST.get("margin_barber"))
            margin_distributor = parse_optional_decimal(request.POST.get("margin_distributor"))
            if margin_consumer is None and margin_barber is None and margin_distributor is None:
                messages.error(request, "Completá al menos un margen para aplicar cambios.")
                return redirect("inventory_product_costs")

            target_products = list(Product.objects.order_by("sku"))
            if not target_products:
                messages.error(request, "No se encontraron productos para aplicar los cambios.")
                return redirect("inventory_product_costs")

            updated = 0
            for product in target_products:
                update_fields = []
                if margin_consumer is not None:
                    product.margin_consumer = margin_consumer
                    update_fields.append("margin_consumer")
                if margin_barber is not None:
                    product.margin_barber = margin_barber
                    update_fields.append("margin_barber")
                if margin_distributor is not None:
                    product.margin_distributor = margin_distributor
                    update_fields.append("margin_distributor")
                if update_fields:
                    product.save(update_fields=update_fields)
                    updated += 1
            messages.success(request, f"Márgenes actualizados en {updated} productos.")
            return redirect("inventory_product_costs")
        elif action == "import_costs":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Subí un archivo XLSX.")
            else:
                group_override = (request.POST.get("group_override") or "").strip()
                result = _process_costs_xlsx(upload, group_override=group_override or None)
                if isinstance(result, str):
                    messages.error(request, result)
                else:
                    created, updated, skipped = result
                    if created == 0 and updated == 0 and skipped == 0:
                        messages.warning(
                            request,
                            "No se encontraron filas válidas para importar. Revisá las columnas y datos.",
                        )
                        return redirect("inventory_product_costs")
                    messages.success(
                        request,
                        f"Importación completa. Nuevos: {created}, Actualizados: {updated}, Omitidos: {skipped}.",
                    )
                    return redirect("inventory_product_costs")
        elif action == "delete_import":
            upload = request.FILES.get("file")
            if not upload:
                messages.error(request, "Subí un archivo XLSX.")
            else:
                rows, error = _read_costs_xlsx_rows(upload)
                if error:
                    messages.error(request, error)
                else:
                    deleted = 0
                    skipped = 0
                    not_found = 0
                    for group, description, _cost, sku in rows:
                        product = None
                        if sku:
                            product = Product.objects.filter(sku__iexact=sku).first()
                        if not product:
                            product = Product.objects.filter(name=description, group=group).first()
                        if not product:
                            not_found += 1
                            continue
                        if (
                            product.sale_items.exists()
                            or product.purchase_items.exists()
                            or product.movements.exists()
                            or product.stocks.exists()
                        ):
                            skipped += 1
                            continue
                        product.delete()
                        deleted += 1
                    messages.success(
                        request,
                        f"Eliminación completa. Borrados: {deleted}, En uso: {skipped}, No encontrados: {not_found}.",
                    )
                    return redirect("inventory_product_costs")
        elif action == "delete_product":
            product_id = request.POST.get("product_id")
            if not product_id:
                messages.error(request, "No se pudo identificar el producto.")
                return redirect("inventory_product_costs")
            product = Product.objects.filter(id=product_id).first()
            if not product:
                messages.error(request, "Producto no encontrado.")
                return redirect("inventory_product_costs")
            try:
                product.delete()
                messages.success(request, "Producto eliminado.")
            except ProtectedError:
                messages.error(
                    request,
                    "No se puede eliminar el producto porque está usado en ventas o movimientos. Podés desactivarlo retirando stock o duplicarlo.",
                )
            return redirect("inventory_product_costs")
        elif action == "quick_update_cost":
            product_id = request.POST.get("product_id")
            avg_cost_raw = request.POST.get("avg_cost")
            if not product_id:
                return JsonResponse({"ok": False, "error": "missing_product_id"}, status=400)
            product = Product.objects.filter(id=product_id).first()
            if not product:
                return JsonResponse({"ok": False, "error": "product_not_found"}, status=404)
            try:
                avg_cost = _parse_decimal(avg_cost_raw)
            except Exception:
                avg_cost = Decimal("0.00")
            product.avg_cost = avg_cost
            product.save(update_fields=["avg_cost"])
            return JsonResponse({"ok": True})
        elif action == "bulk_update":
            bulk_form = ProductBulkUpdateForm(request.POST)
            if bulk_form.is_valid():
                group = (bulk_form.cleaned_data.get("group") or "").strip()
                supplier = bulk_form.cleaned_data.get("supplier")
                cost_percent = bulk_form.cleaned_data.get("cost_percent")
                margin_consumer = bulk_form.cleaned_data.get("margin_consumer")
                margin_barber = bulk_form.cleaned_data.get("margin_barber")
                margin_distributor = bulk_form.cleaned_data.get("margin_distributor")
                if (
                    not group
                    and not supplier
                    and cost_percent is None
                    and margin_consumer is None
                    and margin_barber is None
                    and margin_distributor is None
                ):
                    messages.error(request, "Completá al menos un campo para aplicar cambios.")
                else:
                    target_qs = Product.objects.select_related("default_supplier").order_by("sku")
                    if group:
                        target_qs = target_qs.filter(group__iexact=group)

                    target_products = list(target_qs)
                    if not target_products:
                        messages.error(request, "No se encontraron productos para aplicar los cambios.")
                        return redirect("inventory_product_costs")

                    updated = 0
                    for product in target_products:
                        update_fields = []
                        if group and product.group != group:
                            product.group = group
                            update_fields.append("group")
                        if supplier and product.default_supplier_id != supplier.id:
                            product.default_supplier = supplier
                            update_fields.append("default_supplier")
                        if cost_percent is not None:
                            multiplier = Decimal("1.00") + (cost_percent / Decimal("100.00"))
                            product.avg_cost = (product.avg_cost or Decimal("0.00")) * multiplier
                            product.avg_cost = product.avg_cost.quantize(Decimal("0.01"))
                            update_fields.append("avg_cost")
                        if margin_consumer is not None:
                            product.margin_consumer = margin_consumer
                            update_fields.append("margin_consumer")
                        if margin_barber is not None:
                            product.margin_barber = margin_barber
                            update_fields.append("margin_barber")
                        if margin_distributor is not None:
                            product.margin_distributor = margin_distributor
                            update_fields.append("margin_distributor")
                        if update_fields:
                            product.save(update_fields=update_fields)
                            updated += 1
                        if supplier:
                            SupplierProduct.objects.update_or_create(
                                supplier=supplier,
                                product=product,
                                defaults={
                                    "last_cost": product.avg_cost,
                                    "last_purchase_at": timezone.now(),
                                },
                            )
                    messages.success(request, f"Actualización masiva aplicada a {updated} productos.")
                    return redirect("inventory_product_costs")
        elif action == "update_cost_by_product":
            product_id = request.POST.get("product_id")
            avg_cost_raw = request.POST.get("avg_cost")
            if not product_id:
                messages.error(request, "Seleccioná un producto.")
                return redirect("inventory_product_costs")
            product = Product.objects.filter(id=product_id).first()
            if not product:
                messages.error(request, "Producto no encontrado.")
                return redirect("inventory_product_costs")
            avg_cost = _parse_decimal(avg_cost_raw)
            product.avg_cost = avg_cost
            product.save(update_fields=["avg_cost"])
            messages.success(request, "Costo actualizado.")
            return redirect("inventory_product_costs")
        else:
            formset = ProductCostFormSet(request.POST)
            if formset.is_valid():
                has_errors = False
                for form in formset:
                    product = product_map.get(form.cleaned_data["product_id"])
                    if not product:
                        continue
                    name = (form.cleaned_data.get("name") or "").strip()
                    group = (form.cleaned_data.get("group") or "").strip()
                    if not name:
                        form.add_error("name", "Nombre requerido.")
                        has_errors = True
                if has_errors:
                    messages.error(request, "Revisá el SKU o nombre del producto.")
                    return render(
                        request,
                        "inventory/cost_list.html",
                        {"formset": formset, "product_form": product_form},
                    )

                for form in formset:
                    product = product_map.get(form.cleaned_data["product_id"])
                    if not product:
                        continue
                    avg_cost = form.cleaned_data["avg_cost"]
                    supplier = form.cleaned_data.get("supplier")
                    name = form.cleaned_data.get("name") or product.name
                    group = (form.cleaned_data.get("group") or "").strip()
                    vat_percent = form.cleaned_data.get("vat_percent")
                    if vat_percent is None:
                        vat_percent = Decimal("0.00")
                    update_fields = []
                    if product.name != name:
                        product.name = name
                        update_fields.append("name")
                    if product.group != group:
                        product.group = group
                        update_fields.append("group")
                    if product.avg_cost != avg_cost:
                        product.avg_cost = avg_cost
                        update_fields.append("avg_cost")
                    if product.vat_percent != vat_percent:
                        product.vat_percent = vat_percent
                        update_fields.append("vat_percent")
                    if supplier and product.default_supplier_id != supplier.id:
                        product.default_supplier = supplier
                        update_fields.append("default_supplier")
                    if update_fields:
                        product.save(update_fields=update_fields)
                    if supplier:
                        SupplierProduct.objects.update_or_create(
                            supplier=supplier,
                            product=product,
                            defaults={"last_cost": avg_cost, "last_purchase_at": timezone.now()},
                        )
                messages.success(request, "Costos y proveedores actualizados.")
                return redirect("inventory_product_costs")
            messages.error(request, "Revisá los costos ingresados.")
    return render(
        request,
        "inventory/cost_list.html",
        {
            "formset": formset,
            "product_form": product_form,
            "bulk_form": bulk_form,
            "group_options": group_options,
            "products": products,
        },
    )


@login_required
def product_info(request):
    product_id = request.GET.get("product_id")
    if not product_id:
        return JsonResponse({"ok": False, "error": "missing_product_id"}, status=400)
    product = _products_with_last_cost_queryset().select_related("default_supplier").filter(id=product_id).first()
    if not product:
        return JsonResponse({"ok": False, "error": "product_not_found"}, status=404)
    return JsonResponse(
        {
            "ok": True,
            "default_supplier_id": product.default_supplier_id,
            "avg_cost": f"{(product.avg_cost or Decimal('0.00')):.2f}",
            "vat_percent": f"{(product.vat_percent or Decimal('0.00')):.2f}",
            "cost_with_vat": f"{product.cost_with_vat():.2f}",
        }
    )


@login_required
def product_search(request):
    term = (request.GET.get("q") or "").strip()
    if not term:
        return JsonResponse({"ok": True, "results": []})
    qs = _products_with_last_cost_queryset().filter(
        Q(sku__icontains=term) | Q(name__icontains=term) | Q(group__icontains=term)
    )
    results = []
    for product in qs.order_by("sku")[:20]:
        results.append(
            {
                "id": product.id,
                "label": _product_label_with_last_cost(product),
                "default_supplier_id": product.default_supplier_id,
                "avg_cost": f"{(product.avg_cost or Decimal('0.00')):.2f}",
                "vat_percent": f"{(product.vat_percent or Decimal('0.00')):.2f}",
                "cost_with_vat": f"{product.cost_with_vat():.2f}",
            }
        )
    return JsonResponse({"ok": True, "results": results})


def _col_letter(idx: int) -> str:
    result = ""
    while idx:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _build_xlsx(
    headers: list[str],
    rows: list[list[str | Decimal]],
    *,
    blue_cols: set[int] | None = None,
    number_cols: set[int] | None = None,
) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Root relations
        zf.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
            '<Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
            '<Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
            "</Types>",
        )
        zf.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
            '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "docProps/core.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
            'xmlns:dc="http://purl.org/dc/elements/1.1/" '
            'xmlns:dcterms="http://purl.org/dc/terms/" '
            'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
            'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
            "<dc:title>Precios</dc:title>"
            "</cp:coreProperties>",
        )
        zf.writestr(
            "docProps/app.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
            'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
            "<Application>Django</Application>"
            "</Properties>",
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
            "</Relationships>",
        )
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Precios" sheetId="1" r:id="rId1"/></sheets>'
            "</workbook>",
        )
        zf.writestr(
            "xl/styles.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<fonts count="3">'
            '<font><sz val="11"/><color theme="1"/><name val="Calibri"/></font>'
            '<font><b/><sz val="11"/><color theme="1"/><name val="Calibri"/></font>'
            '<font><sz val="11"/><color rgb="FF1F4EAD"/><name val="Calibri"/></font>'
            "</fonts>"
            '<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
            '<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
            '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
            '<cellXfs count="4">'
            '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
            '<xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
            '<xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1"/>'
            '<xf numFmtId="2" fontId="0" fillId="0" borderId="0" xfId="0" applyNumberFormat="1"/>'
            "</cellXfs>"
            '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
            "</styleSheet>",
        )

        cols = len(headers)
        rows_count = len(rows) + 1  # header row
        dimension = f"A1:{_col_letter(cols)}{rows_count}"

        blue_cols = blue_cols or set()
        number_cols = number_cols or set()

        def cell_xml(value, col_idx, row_idx, is_header=False):
            col = _col_letter(col_idx)
            ref = f"{col}{row_idx}"
            if is_header:
                return f'<c r="{ref}" t="inlineStr" s="1"><is><t>{escape(str(value))}</t></is></c>'
            if isinstance(value, str):
                style = "2" if col_idx in blue_cols else "0"
                return f'<c r="{ref}" t="inlineStr" s="{style}"><is><t>{escape(str(value))}</t></is></c>'
            style = "3" if col_idx in number_cols else "0"
            return f'<c r="{ref}" s="{style}"><v>{value}</v></c>'

        max_lengths = [len(str(h)) for h in headers]
        for row in rows:
            for idx, val in enumerate(row):
                if idx >= len(max_lengths):
                    continue
                if isinstance(val, Decimal):
                    text = f"{val:.2f}" if (idx + 1) in number_cols else str(val)
                else:
                    text = str(val)
                if len(text) > max_lengths[idx]:
                    max_lengths[idx] = len(text)
        col_widths = []
        for length in max_lengths:
            col_widths.append(min(60, max(8, round(length * 1.1 + 2, 2))))

        rows_xml = []
        header_cells = "".join(cell_xml(h, i + 1, 1, is_header=True) for i, h in enumerate(headers))
        rows_xml.append(f'<row r="1">{header_cells}</row>')
        for ridx, row in enumerate(rows, start=2):
            cells = "".join(cell_xml(val, cidx + 1, ridx) for cidx, val in enumerate(row))
            rows_xml.append(f'<row r="{ridx}">{cells}</row>')

        cols_xml = "".join(
            f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>'
            for idx, width in enumerate(col_widths, start=1)
        )
        sheet_xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
            f' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            f'<dimension ref="{dimension}"/>'
            f"<cols>{cols_xml}</cols>"
            "<sheetData>"
            f'{"".join(rows_xml)}'
            "</sheetData>"
            "</worksheet>"
        )
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return buf.getvalue()


@login_required
def product_prices_download(request, audience: str):
    products = Product.objects.order_by("sku")
    groups_raw = request.GET.get("groups", "")
    if groups_raw:
        groups = [g.strip() for g in groups_raw.split(",") if g.strip()]
        if groups:
            products = products.filter(group__in=groups)
    headers = ["Marca", "Producto", "Precio"]
    price_attr_map = {
        "consumer": "consumer_price",
        "barber": "barber_price",
        "distributor": "distributor_price",
    }
    if audience not in price_attr_map:
        return redirect("inventory_product_prices")

    attr = price_attr_map[audience]
    rows = [[p.group or "", p.name, getattr(p, attr)] for p in products]
    xlsx_bytes = _build_xlsx(headers, rows, blue_cols={1}, number_cols={3})
    audience_label_map = {
        "consumer": "consumidor",
        "barber": "peluqueria",
        "distributor": "distribuidor",
    }
    audience_label = audience_label_map.get(audience, audience)
    if groups_raw:
        if len(groups) == 1:
            brand_slug = slugify(groups[0])
        else:
            brand_slug = "varias_marcas"
        filename = f"{brand_slug}_{audience_label}.xlsx"
    else:
        filename = f"precios_{audience_label}.xlsx"
    response = HttpResponse(
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _decimal_or_zero(value: str | None) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0.00")
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _parse_decimal(value: str | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    cleaned = str(value).strip().replace(".", "").replace(",", ".")
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if cleaned == "":
        return Decimal("0.00")
    try:
        return Decimal(cleaned).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _abbr(text: str | None, length: int) -> str:
    if not text:
        return "X" * length
    cleaned = re.sub(r"[^A-Za-z0-9]", "", text).upper()
    return cleaned[:length].ljust(length, "X")


def _sku_prefix(group: str, description: str) -> str:
    words = [w for w in re.split(r"\s+", (description or "").strip()) if w]
    word1 = _abbr(words[0], 3) if len(words) > 0 else "XXX"
    word2 = _abbr(words[1], 3) if len(words) > 1 else "XXX"
    return f"{_abbr(group, 4)}{word1}{word2}"


def _normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    without_accents = "".join(c for c in normalized if not unicodedata.combining(c))
    return without_accents.strip().lower()


def _read_costs_xlsx_rows(upload) -> tuple[list[tuple[str, str, Decimal, str]], str | None]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], "Falta la dependencia openpyxl. Instalá openpyxl en el entorno."

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active
    except Exception:
        return [], "No se pudo leer el archivo XLSX."

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    header_map = {_normalize_header(h): idx for idx, h in enumerate(headers)}
    desc_keys = ["descripcion", "descripción", "producto", "descripcion producto", "nombre"]
    cost_keys = ["precio venta", "precio", "costo", "costo unitario", "precio costo", "precio venta unitario"]
    group_keys = ["grupo", "marca", "categoria", "categoría"]
    sku_keys = ["sku", "codigo", "código", "cod", "codigo sku"]

    def _pick_index(keys: list[str]) -> int | None:
        for key in keys:
            normalized = _normalize_header(key)
            if normalized in header_map:
                return header_map[normalized]
        return None

    desc_idx = _pick_index(desc_keys)
    cost_idx = _pick_index(cost_keys)
    group_idx = _pick_index(group_keys)
    sku_idx = _pick_index(sku_keys)

    if desc_idx is None or cost_idx is None:
        return [], "Faltan columnas obligatorias: Descripción/Producto y Precio/Costo."

    rows: list[tuple[str, str, Decimal, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        group = str(row[group_idx] or "").strip() if group_idx is not None else ""
        description = str(row[desc_idx] or "").strip()
        cost = _parse_decimal(row[cost_idx])
        sku = str(row[sku_idx] or "").strip() if sku_idx is not None else ""
        if not description:
            continue
        rows.append((group, description, cost, sku))
    return rows, None


def _read_ml_sales_xlsx_rows(upload) -> tuple[list[dict], str | None]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return [], "Falta la dependencia openpyxl. Instalá openpyxl en el entorno."

    try:
        wb = load_workbook(upload, data_only=True)
        ws = wb.active
    except Exception:
        return [], "No se pudo leer el archivo XLSX."

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in header_cells
    ]
    header_map = {_normalize_header(h): idx for idx, h in enumerate(headers)}

    def _pick_index(keys: list[str]) -> int | None:
        for key in keys:
            normalized = _normalize_header(key)
            if normalized in header_map:
                return header_map[normalized]
        return None

    date_idx = _pick_index(["fecha", "fecha venta"])
    order_idx = _pick_index(
        [
            "comprobante",
            "nro comprobante",
            "número comprobante",
            "numero comprobante",
            "orden",
            "nro orden",
            "número orden",
            "numero orden",
            "pedido",
            "nro pedido",
            "número pedido",
            "numero pedido",
            "order id",
            "id pedido",
        ]
    )
    product_idx = _pick_index(["producto", "publicacion", "publicación", "titulo", "título", "nombre"])
    qty_idx = _pick_index(["cantidad", "cant"])
    price_total_idx = _pick_index(["precio bruto venta", "precio bruto", "precio"])
    commission_idx = _pick_index(["comision", "comisión"])
    tax_idx = _pick_index(["impuestos", "impuesto", "iibb"])

    required = [date_idx, product_idx, qty_idx, price_total_idx, commission_idx, tax_idx]
    if any(idx is None for idx in required):
        return [], "Faltan columnas obligatorias: Fecha, Producto, Cantidad, Precio Bruto Venta, Comision, Impuestos."

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = str(row[product_idx] or "").strip() if product_idx is not None else ""
        if not title:
            continue
        qty = _parse_decimal(row[qty_idx])
        if qty <= 0:
            continue
        price_total = _parse_decimal(row[price_total_idx])
        commission = _parse_decimal(row[commission_idx])
        taxes = _parse_decimal(row[tax_idx])
        created_at = row[date_idx] if date_idx is not None else None
        order_id = str(row[order_idx] or "").strip() if order_idx is not None else ""
        rows.append(
            {
                "title": title,
                "quantity": qty,
                "price_total": price_total,
                "commission": commission,
                "taxes": taxes,
                "created_at": created_at,
                "order_id": order_id,
            }
        )
    return rows, None



def _process_costs_xlsx(
    upload,
    *,
    group_override: str | None = None,
) -> tuple[int, int, int] | str:
    rows, error = _read_costs_xlsx_rows(upload)
    if error:
        return error

    created = 0
    updated = 0
    skipped = 0
    prefix_counters: dict[str, int] = {}

    effective_group = (group_override or "").strip()

    for group, description, cost, sku in rows:
        if effective_group:
            group = effective_group

        product = None
        if sku:
            product = Product.objects.filter(sku__iexact=sku).first()
        if not product:
            product = Product.objects.filter(name=description, group=group).first()
        if product:
            product.avg_cost = cost
            product.save(update_fields=["avg_cost"])
            updated += 1
            continue

        if effective_group:
            skipped += 1
            continue

        prefix = _sku_prefix(group, description)
        if prefix not in prefix_counters:
            existing = (
                Product.objects.filter(sku__startswith=prefix)
                .values_list("sku", flat=True)
            )
            max_suffix = 0
            for sku in existing:
                suffix = sku[len(prefix):]
                if suffix.isdigit():
                    max_suffix = max(max_suffix, int(suffix))
            prefix_counters[prefix] = max_suffix

        prefix_counters[prefix] += 1
        sku = f"{prefix}{prefix_counters[prefix]:04d}"

        Product.objects.create(
            sku=sku,
            name=description,
            group=group,
            avg_cost=cost,
        )
        created += 1

    return created, updated, skipped


@login_required
@require_http_methods(["GET", "POST"])
def import_products(request):
    """
    Importa productos desde un CSV exportado de Google Sheets.
    Encabezados esperados (case-insensitive):
    SKU, Grupo, Nombre, Costo unitario, IVA, Margen consumidor, Margen barber, Margen distribuidor.
    Columnas faltantes se llenan con 0.
    """
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Subí un archivo CSV.")
            return redirect("inventory_import_products")

        import csv

        try:
            decoded = upload.read().decode("utf-8-sig").splitlines()
            reader = csv.DictReader(decoded)
        except Exception:
            messages.error(request, "No se pudo leer el CSV. Verificá el formato.")
            return redirect("inventory_import_products")

        required_cols = ["sku", "nombre", "costo unitario"]
        normalized_fieldnames = [name.strip().lower() for name in (reader.fieldnames or [])]
        missing = [col for col in required_cols if col not in normalized_fieldnames]
        if missing:
            messages.error(request, f"Faltan columnas obligatorias: {', '.join(missing)}")
            return redirect("inventory_import_products")

        created = 0
        updated = 0
        for row in reader:
            data = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            sku = data.get("sku")
            if not sku:
                continue
            product, was_created = Product.objects.get_or_create(
                sku=sku,
                defaults={
                    "name": data.get("nombre", ""),
                    "group": data.get("grupo", ""),
                    "avg_cost": _decimal_or_zero(data.get("costo unitario")),
                    "vat_percent": _decimal_or_zero(data.get("iva")),
                    "margin_consumer": _decimal_or_zero(data.get("margen consumidor")),
                    "margin_barber": _decimal_or_zero(data.get("margen barber")),
                    "margin_distributor": _decimal_or_zero(data.get("margen distribuidor")),
                },
            )
            if not was_created:
                product.name = data.get("nombre", product.name)
                product.group = data.get("grupo", product.group)
                product.avg_cost = _decimal_or_zero(data.get("costo unitario"))
                product.vat_percent = _decimal_or_zero(data.get("iva"))
                product.margin_consumer = _decimal_or_zero(data.get("margen consumidor"))
                product.margin_barber = _decimal_or_zero(data.get("margen barber"))
                product.margin_distributor = _decimal_or_zero(data.get("margen distribuidor"))
                product.save(
                    update_fields=[
                        "name",
                        "group",
                        "avg_cost",
                        "vat_percent",
                        "margin_consumer",
                        "margin_barber",
                        "margin_distributor",
                    ]
                )
                updated += 1
            else:
                created += 1

        messages.success(request, f"Importación completa. Nuevos: {created}, Actualizados: {updated}.")
        return redirect("inventory_product_prices")

    return render(request, "inventory/product_import.html", {"title": "Importar productos"})


@login_required
@require_http_methods(["GET", "POST"])
def import_costs_xlsx(request):
    """
    Importa costos desde un Excel (.xlsx) con columnas:
    Grupo, Descripción, Precio Venta.
    """
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Subí un archivo XLSX.")
            return redirect("inventory_import_costs")

        group_override = (request.POST.get("group_override") or "").strip()
        result = _process_costs_xlsx(upload, group_override=group_override or None)
        if isinstance(result, str):
            messages.error(request, result)
            return redirect("inventory_import_costs")
        created, updated, skipped = result
        messages.success(
            request,
            f"Importación completa. Nuevos: {created}, Actualizados: {updated}, Omitidos: {skipped}.",
        )
        return redirect("inventory_product_prices")

    group_options = (
        Product.objects.exclude(group="")
        .order_by("group")
        .values_list("group", flat=True)
        .distinct()
    )
    return render(
        request,
        "inventory/cost_import.html",
        {"title": "Importar costos", "group_options": group_options},
    )


@login_required
def product_delete(request, pk: int):
    product = get_object_or_404(Product, pk=pk)
    try:
        product.delete()
        messages.success(request, "Producto eliminado.")
    except ProtectedError:
        messages.error(
            request,
            "No se puede eliminar el producto porque está usado en ventas o movimientos. Podés desactivarlo retirando stock o duplicarlo.",
        )
    return redirect("inventory_product_prices")


@login_required
@require_http_methods(["GET"])
def mercadolibre_connect(request):
    if not settings.ML_CLIENT_ID or not settings.ML_CLIENT_SECRET or not settings.ML_REDIRECT_URI:
        messages.error(request, "Faltan credenciales de MercadoLibre en las variables de entorno.")
        return redirect("inventory_mercadolibre_dashboard")
    state = secrets.token_urlsafe(16)
    request.session["ml_state"] = state
    return redirect(ml.get_authorize_url(state))


@login_required
@require_http_methods(["GET"])
def mercadolibre_callback(request):
    code = request.GET.get("code")
    state = request.GET.get("state")
    error = request.GET.get("error")
    error_description = request.GET.get("error_description")
    if error:
        return render(
            request,
            "inventory/mercadolibre_callback.html",
            {
                "code": code,
                "state": state,
                "error": error,
                "error_description": error_description,
            },
        )
    if not code:
        messages.error(request, "No se recibió el código de autorización.")
        return redirect("inventory_mercadolibre_dashboard")
    expected_state = request.session.pop("ml_state", None)
    if expected_state and state != expected_state:
        messages.error(request, "State inválido en el callback de MercadoLibre.")
        return redirect("inventory_mercadolibre_dashboard")
    token_data = ml.exchange_code_for_token(code)
    if token_data.get("error"):
        messages.error(
            request,
            f"No se pudo conectar con MercadoLibre. {token_data.get('error_description', token_data.get('error'))}",
        )
        return redirect("inventory_mercadolibre_dashboard")
    access_token = token_data.get("access_token")
    refresh_token = token_data.get("refresh_token", "")
    expires_in = int(token_data.get("expires_in", 0) or 0)
    if not access_token:
        messages.error(request, "No se pudo completar la conexión con MercadoLibre.")
        return redirect("inventory_mercadolibre_dashboard")

    connection, _ = MercadoLibreConnection.objects.get_or_create(user=request.user)
    connection.access_token = access_token
    connection.refresh_token = refresh_token
    connection.expires_at = timezone.now() + timedelta(seconds=expires_in) if expires_in else None
    profile = ml.get_user_profile(access_token)
    connection.ml_user_id = str(profile.get("id", "") or "")
    connection.nickname = profile.get("nickname", "") or ""
    connection.save(update_fields=["access_token", "refresh_token", "expires_at", "ml_user_id", "nickname"])

    messages.success(request, "MercadoLibre conectado correctamente.")
    return redirect("inventory_mercadolibre_dashboard")


@login_required
@require_http_methods(["GET", "POST"])
def mercadolibre_dashboard(request):
    missing_credentials = not settings.ML_CLIENT_ID or not settings.ML_CLIENT_SECRET or not settings.ML_REDIRECT_URI
    try:
        connection = MercadoLibreConnection.objects.filter(user=request.user).first()
        items_qs = MercadoLibreItem.objects.select_related("product")
    except OperationalError:
        messages.error(request, "Faltan tablas de MercadoLibre. Ejecutá migrate y recargá.")
        connection = None
        items_qs = MercadoLibreItem.objects.none()
    metrics = {}
    if connection and connection.last_metrics:
        try:
            metrics = json.loads(connection.last_metrics)
        except json.JSONDecodeError:
            metrics = {}
    search_query = (request.GET.get("q") or "").strip()
    if search_query:
        items_qs = items_qs.filter(title__icontains=search_query)
    items_qs = items_qs.order_by("-available_quantity", "title")
    paginator = Paginator(items_qs, 50)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    items = page_obj.object_list

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "sync":
            if not connection or not connection.access_token:
                messages.error(request, "Primero conectá la cuenta de MercadoLibre.")
            else:
                result = ml.sync_items_and_stock(connection, request.user)
                metrics = result.metrics
                if metrics.get("error") == "unauthorized":
                    messages.error(
                        request,
                        "MercadoLibre rechazó el token. Volvé a conectar la cuenta para actualizar el acceso.",
                    )
                else:
                    notice = (
                        f"Sync OK. Items: {result.total_items}, Matcheados: {result.matched}, "
                        f"Sin match: {result.unmatched}, Stock actualizado: {result.updated_stock}."
                    )
                    if metrics.get("truncated"):
                        notice += " (Sync limitado por configuración)"
                    messages.success(request, notice)
        elif action == "sync_full":
            if not connection or not connection.access_token:
                messages.error(request, "Primero conectá la cuenta de MercadoLibre.")
            else:
                result = ml.sync_items_and_stock(connection, request.user, ignore_env_limit=True)
                metrics = result.metrics
                if metrics.get("error") == "unauthorized":
                    messages.error(
                        request,
                        "MercadoLibre rechazó el token. Volvé a conectar la cuenta para actualizar el acceso.",
                    )
                else:
                    notice = (
                        f"Sync completo OK. Items: {result.total_items}, Matcheados: {result.matched}, "
                        f"Sin match: {result.unmatched}, Stock actualizado: {result.updated_stock}."
                    )
                    messages.success(request, notice)
        elif action == "sync_orders":
            if not connection or not connection.access_token:
                messages.error(request, "Primero conectá la cuenta de MercadoLibre.")
            else:
                days_env = os.environ.get("ML_ORDERS_DAYS", "")
                days = int(days_env) if days_env.isdigit() else 30
                result = ml.sync_recent_orders(connection, request.user, days=days)
                if result.get("reasons", {}).get("unauthorized"):
                    messages.error(
                        request,
                        "MercadoLibre rechazó el token. Volvé a conectar la cuenta para actualizar el acceso.",
                    )
                else:
                    reason_parts = []
                    for key, count in result.get("reasons", {}).items():
                        reason_parts.append(f"{key}: {count}")
                    reason_text = f" ({', '.join(reason_parts)})" if reason_parts else ""
                    messages.success(
                        request,
                        "Sync ventas OK. Revisadas: "
                        f"{result['total']}, nuevas: {result['created']}, "
                        f"actualizadas: {result.get('updated', 0)}.{reason_text}",
                    )
        elif action == "sync_item":
            item_id = (request.POST.get("ml_item_id") or "").strip()
            if not item_id:
                messages.error(request, "Ingresá el ID de la publicación.")
            elif not connection or not connection.access_token:
                messages.error(request, "Primero conectá la cuenta de MercadoLibre.")
            else:
                if item_id.isdigit():
                    site_id = (getattr(settings, "ML_SITE_ID", "") or "MLA").upper()
                    item_id = f"{site_id}{item_id}"
                try:
                    item = ml.get_item(item_id, connection.access_token)
                    title = item.get("title", "") or ""
                    available = int(item.get("available_quantity", 0) or 0)
                    status = item.get("status", "") or ""
                    shipping = item.get("shipping") or {}
                    logistic_type = item.get("logistic_type", "") or shipping.get("logistic_type", "") or ""
                    permalink = item.get("permalink", "") or ""
                    existing = MercadoLibreItem.objects.filter(item_id=item_id).first()
                    product = existing.product if existing else None
                    matched_name = existing.matched_name if existing else ""
                    MercadoLibreItem.objects.update_or_create(
                        item_id=item_id,
                        defaults={
                            "title": title,
                            "available_quantity": available,
                            "status": status,
                            "logistic_type": logistic_type,
                            "permalink": permalink,
                            "product": product,
                            "matched_name": matched_name,
                        },
                    )
                    messages.success(request, "Publicación sincronizada.")
                except Exception as exc:
                    messages.error(request, f"No se pudo sincronizar: {exc}")
        elif action == "link_item":
            item_id = request.POST.get("item_id")
            product_id = request.POST.get("product_id")
            ml_item = MercadoLibreItem.objects.filter(id=item_id).first()
            if not ml_item:
                messages.error(request, "No se encontró la publicación.")
            else:
                if product_id:
                    product = Product.objects.filter(id=product_id).first()
                    if not product:
                        messages.error(request, "Producto no encontrado.")
                    else:
                        ml_item.product = product
                        ml_item.matched_name = product.name
                        ml_item.save(update_fields=["product", "matched_name"])
                        messages.success(request, "Match actualizado.")
                else:
                    ml_item.product = None
                    ml_item.matched_name = ""
                    ml_item.save(update_fields=["product", "matched_name"])
                    messages.success(request, "Match eliminado.")

    products = Product.objects.order_by("name")
    recent_cutoff = timezone.now() - timedelta(days=30)
    return render(
        request,
        "inventory/mercadolibre_dashboard.html",
        {
            "connection": connection,
            "items": items,
            "metrics": metrics,
            "missing_credentials": missing_credentials,
            "page_obj": page_obj,
            "search_query": search_query,
            "recent_cutoff": recent_cutoff,
            "products": products,
        },
    )


@csrf_exempt
@require_http_methods(["GET", "POST"])
def mercadolibre_webhook(request):
    if request.method == "GET":
        return HttpResponse("OK")
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return HttpResponse("Invalid JSON", status=400)
    notification = MercadoLibreNotification.objects.create(
        topic=payload.get("topic", "") or "",
        resource=payload.get("resource", "") or "",
        ml_user_id=str(payload.get("user_id", "") or ""),
        application_id=str(payload.get("application_id", "") or ""),
        raw_payload=request.body.decode("utf-8"),
    )
    try:
        if notification.topic == "orders":
            resource = notification.resource or ""
            parts = resource.strip("/").split("/")
            order_id = ""
            if "orders" in parts:
                idx = parts.index("orders")
                if idx + 1 < len(parts):
                    order_id = parts[idx + 1]
            if order_id:
                connection = MercadoLibreConnection.objects.filter(ml_user_id=notification.ml_user_id).first()
                if connection:
                    User = get_user_model()
                    sync_user = (
                        User.objects.filter(is_superuser=True).order_by("id").first()
                        or User.objects.order_by("id").first()
                    )
                    if sync_user:
                        ml.sync_order(connection, order_id, sync_user)
    except Exception:
        pass
    return HttpResponse(f"OK:{notification.id}")


class CustomerForm(forms.ModelForm):
    email = forms.CharField(label="Teléfono", required=False)

    class Meta:
        model = Customer
        fields = ["name", "email", "audience"]
        labels = {"name": "Nombre", "email": "Teléfono", "audience": "Tipo"}


class CustomerDiscountForm(forms.ModelForm):
    class Meta:
        model = CustomerProductDiscount
        fields = ["customer", "product", "discount_percent"]
        labels = {
            "customer": "Cliente",
            "product": "Producto",
            "discount_percent": "Descuento %",
        }

    def validate_unique(self):
        return


class CustomerGroupDiscountForm(forms.ModelForm):
    class Meta:
        model = CustomerGroupDiscount
        fields = ["customer", "group", "discount_percent"]
        labels = {
            "customer": "Cliente",
            "group": "Marca / Grupo",
            "discount_percent": "Descuento %",
        }

    def validate_unique(self):
        return


class CustomerPaymentForm(forms.ModelForm):
    class Meta:
        model = CustomerPayment
        fields = ["sale", "amount", "method", "kind", "paid_at", "notes"]
        labels = {
            "sale": "Pedido",
            "amount": "Monto",
            "method": "Método",
            "kind": "Tipo",
            "paid_at": "Fecha",
            "notes": "Notas",
        }
        widgets = {"paid_at": forms.DateInput(attrs={"type": "date"})}

    def __init__(self, *args, **kwargs):
        customer = kwargs.pop("customer", None)
        super().__init__(*args, **kwargs)
        self.fields["amount"].min_value = Decimal("0.01")
        self.fields["sale"].required = False
        self.fields["notes"].required = False
        if customer is not None:
            self.fields["sale"].queryset = Sale.objects.filter(customer=customer).order_by("-created_at")
        self.fields["sale"].empty_label = "Cuenta corriente"


class TaxExpenseForm(forms.ModelForm):
    class Meta:
        model = TaxExpense
        fields = ["description", "amount", "paid_at"]
        labels = {"description": "Descripción", "amount": "Monto", "paid_at": "Fecha"}
        widgets = {"paid_at": forms.DateInput(attrs={"type": "date"})}


@login_required
def customers_view(request):
    customer_form = CustomerForm()
    discount_form = CustomerDiscountForm()
    group_discount_form = CustomerGroupDiscountForm()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_customer":
            customer_form = CustomerForm(request.POST)
            if customer_form.is_valid():
                customer_form.save()
                messages.success(request, "Cliente creado.")
                return redirect("inventory_customers")
            else:
                messages.error(request, "Revisá los datos del cliente.")
        elif action == "create_discount":
            discount_form = CustomerDiscountForm(request.POST)
            if discount_form.is_valid():
                customer = discount_form.cleaned_data["customer"]
                product = discount_form.cleaned_data["product"]
                discount = discount_form.cleaned_data["discount_percent"]
                CustomerProductDiscount.objects.update_or_create(
                    customer=customer,
                    product=product,
                    defaults={"discount_percent": discount},
                )
                messages.success(request, "Descuento asignado.")
                return redirect("inventory_customers")
            else:
                messages.error(request, "Revisá los datos del descuento.")
        elif action == "create_group_discount":
            group_discount_form = CustomerGroupDiscountForm(request.POST)
            if group_discount_form.is_valid():
                customer = group_discount_form.cleaned_data["customer"]
                group = (group_discount_form.cleaned_data["group"] or "").strip()
                discount = group_discount_form.cleaned_data["discount_percent"]
                CustomerGroupDiscount.objects.update_or_create(
                    customer=customer,
                    group=group,
                    defaults={"discount_percent": discount},
                )
                messages.success(request, "Descuento por grupo asignado.")
                return redirect("inventory_customers")
            messages.error(request, "Revisá los datos del descuento por grupo.")
        elif action == "update_customer_audience":
            customer_id = request.POST.get("customer_id")
            audience = request.POST.get("audience")
            valid_audiences = {choice[0] for choice in Customer.Audience.choices}
            if customer_id and audience in valid_audiences:
                Customer.objects.filter(id=customer_id).update(audience=audience)
                messages.success(request, "Tipo de cliente actualizado.")
                return redirect("inventory_customers")
            messages.error(request, "Revisá el tipo de cliente.")
        elif action == "update_customer_phone":
            customer_id = request.POST.get("customer_id")
            phone = (request.POST.get("phone") or "").strip()
            if customer_id:
                Customer.objects.filter(id=customer_id).update(email=phone)
                messages.success(request, "Teléfono actualizado.")
                return redirect("inventory_customers")
            messages.error(request, "No se pudo actualizar el teléfono.")

    customers = Customer.objects.prefetch_related("discounts__product", "group_discounts").order_by("name")
    sales_totals = {
        row["customer_id"]: row["total"] or Decimal("0.00")
        for row in Sale.objects.filter(customer__isnull=False)
        .values("customer_id")
        .annotate(total=Sum("total"))
    }
    payments_totals = {
        row["customer_id"]: row["total"] or Decimal("0.00")
        for row in CustomerPayment.objects.filter(kind=CustomerPayment.Kind.PAYMENT)
        .values("customer_id")
        .annotate(total=Sum("amount"))
    }
    refunds_totals = {
        row["customer_id"]: row["total"] or Decimal("0.00")
        for row in CustomerPayment.objects.filter(kind=CustomerPayment.Kind.REFUND)
        .values("customer_id")
        .annotate(total=Sum("amount"))
    }
    debtors = []
    total_debt = Decimal("0.00")
    for customer in customers:
        sales_total = sales_totals.get(customer.id, Decimal("0.00"))
        payments_total = payments_totals.get(customer.id, Decimal("0.00"))
        refunds_total = refunds_totals.get(customer.id, Decimal("0.00"))
        balance = sales_total - payments_total + refunds_total
        if balance > 0:
            debtors.append({"customer": customer, "balance": balance})
            total_debt += balance
    debtors.sort(key=lambda item: item["balance"], reverse=True)
    debtors = debtors[:8]
    group_options = list(
        Product.objects.exclude(group__exact="")
        .values_list("group", flat=True)
        .distinct()
        .order_by("group")
    )
    return render(
        request,
        "inventory/customers.html",
        {
            "customer_form": customer_form,
            "discount_form": discount_form,
            "group_discount_form": group_discount_form,
            "customers": customers,
            "audience_choices": Customer.Audience.choices,
            "group_options": group_options,
            "total_debt": total_debt,
            "debtors": debtors,
        },
    )


@login_required
def customer_history_view(request, customer_id):
    customer = get_object_or_404(Customer, id=customer_id)
    payment_form = CustomerPaymentForm(customer=customer)

    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "add_payment":
            payment_form = CustomerPaymentForm(request.POST, customer=customer)
            if payment_form.is_valid():
                payment = payment_form.save(commit=False)
                payment.customer = customer
                payment.save()
                messages.success(request, "Pago registrado.")
                return redirect("inventory_customer_history", customer_id=customer.id)
            messages.error(request, "Revisá los datos del pago.")

    sales = list(
        Sale.objects.filter(customer=customer)
        .select_related("warehouse")
        .order_by("-created_at", "-id")
    )
    payments = list(
        CustomerPayment.objects.filter(customer=customer)
        .select_related("sale")
        .order_by("-paid_at", "-id")
    )

    payment_by_sale = {}
    for payment in payments:
        if not payment.sale_id:
            continue
        sale_id = payment.sale_id
        info = payment_by_sale.setdefault(
            sale_id,
            {"paid_total": Decimal("0.00"), "methods": []},
        )
        signed = payment.amount if payment.kind == CustomerPayment.Kind.PAYMENT else -payment.amount
        info["paid_total"] += signed
        info["methods"].append(payment.get_method_display())

    sales_rows = []
    for sale in sales:
        paid_info = payment_by_sale.get(sale.id) or {"paid_total": Decimal("0.00"), "methods": []}
        paid_total = paid_info["paid_total"]
        balance = sale.total - paid_total
        methods = ", ".join(dict.fromkeys([m for m in paid_info["methods"] if m])) or "-"
        sales_rows.append(
            {
                "sale": sale,
                "paid_total": paid_total,
                "methods": methods,
                "balance": balance,
            }
        )

    ledger_entries = []
    for sale in sales:
        sale_date = sale.created_at
        if timezone.is_naive(sale_date):
            sale_date = timezone.make_aware(sale_date, timezone.get_current_timezone())
        ledger_entries.append(
            {
                "date": sale_date,
                "date_display": sale.created_at,
                "kind": "SALE",
                "label": sale.ml_order_id or sale.invoice_number,
                "detail": f"Venta ({sale.warehouse.name})",
                "debit": sale.total,
                "credit": Decimal("0.00"),
            }
        )
    for payment in payments:
        is_payment = payment.kind == CustomerPayment.Kind.PAYMENT
        debit = Decimal("0.00") if is_payment else payment.amount
        credit = payment.amount if is_payment else Decimal("0.00")
        label = payment.get_method_display()
        detail_parts = [label]
        if payment.sale_id:
            detail_parts.append(payment.sale.ml_order_id or payment.sale.invoice_number)
        if payment.notes:
            detail_parts.append(payment.notes)
        payment_date = datetime.combine(payment.paid_at, time.min)
        payment_date = timezone.make_aware(payment_date, timezone.get_current_timezone())
        ledger_entries.append(
            {
                "date": payment_date,
                "date_display": payment.paid_at,
                "kind": "PAYMENT" if is_payment else "REFUND",
                "label": "Pago" if is_payment else "Devolución/Ajuste",
                "detail": " · ".join([part for part in detail_parts if part]),
                "debit": debit,
                "credit": credit,
            }
        )

    ledger_entries.sort(key=lambda item: item["date"])
    balance = Decimal("0.00")
    for entry in ledger_entries:
        balance += entry["debit"]
        balance -= entry["credit"]
        entry["balance"] = balance

    total_sales = sum((sale.total for sale in sales), Decimal("0.00"))
    total_payments = sum(
        (payment.amount for payment in payments if payment.kind == CustomerPayment.Kind.PAYMENT),
        Decimal("0.00"),
    )
    total_refunds = sum(
        (payment.amount for payment in payments if payment.kind == CustomerPayment.Kind.REFUND),
        Decimal("0.00"),
    )
    current_balance = total_sales - total_payments + total_refunds

    return render(
        request,
        "inventory/customer_history.html",
        {
            "customer": customer,
            "sales_rows": sales_rows,
            "payments": payments,
            "payment_form": payment_form,
            "ledger_entries": ledger_entries,
            "total_sales": total_sales,
            "total_payments": total_payments,
            "total_refunds": total_refunds,
            "current_balance": current_balance,
        },
    )


@login_required
def taxes_view(request):
    tax_form = TaxExpenseForm()
    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "delete_tax":
            tax_id = request.POST.get("tax_id")
            deleted, _ = TaxExpense.objects.filter(id=tax_id).delete()
            if deleted:
                messages.success(request, "Impuesto eliminado.")
            else:
                messages.error(request, "No se encontró el impuesto.")
            return redirect("inventory_taxes")

        tax_form = TaxExpenseForm(request.POST)
        if tax_form.is_valid():
            tax_form.save()
            messages.success(request, "Impuesto registrado.")
            return redirect("inventory_taxes")
        messages.error(request, "Revisá los datos del impuesto.")

    taxes = TaxExpense.objects.order_by("-paid_at", "-id")
    return render(
        request,
        "inventory/taxes.html",
        {"tax_form": tax_form, "taxes": taxes},
    )
