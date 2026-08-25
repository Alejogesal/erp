"""Proveedor principal por marca (BrandSupplier): la elección del usuario manda
sobre el 'más barato' automático, con fallback si no hay precio."""

from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from inventory import services
from inventory.middleware import _current_user
from inventory.models import BrandSupplier, Product, Supplier, SupplierProduct


def _reset_current_user():
    # El middleware guarda el usuario en un threadlocal que persiste entre tests;
    # lo limpiamos para que los AuditLog de saves fuera de un request no apunten a
    # un usuario de otro test (que ya no existe en esta DB).
    _current_user.user = None


class BrandSupplierServiceTests(TestCase):
    def setUp(self):
        _reset_current_user()
        self.a = Supplier.objects.create(name="Barato A")
        self.b = Supplier.objects.create(name="Caro B")
        self.product = Product.objects.create(name="Prod", group="Marca1", margin_consumer=Decimal("0.00"))
        # A es más barato (50) que B (80). IVA 0 => last_cost == cost_net.
        SupplierProduct.objects.create(supplier=self.a, product=self.product, last_cost=Decimal("50"))
        SupplierProduct.objects.create(supplier=self.b, product=self.product, last_cost=Decimal("80"))

    def test_without_brand_pin_uses_cheapest(self):
        services.sync_principal_to_cheapest(self.product)
        self.product.refresh_from_db()
        self.assertEqual(self.product.default_supplier_id, self.a.id)
        self.assertEqual(self.product.avg_cost, Decimal("50.00"))

    def test_brand_pin_overrides_cheapest(self):
        BrandSupplier.objects.create(group="Marca1", supplier=self.b)
        services.sync_principal_to_cheapest(self.product)
        self.product.refresh_from_db()
        # Aunque A es más barato, manda el proveedor elegido para la marca (B).
        self.assertEqual(self.product.default_supplier_id, self.b.id)
        self.assertEqual(self.product.avg_cost, Decimal("80.00"))

    def test_brand_pin_case_insensitive(self):
        BrandSupplier.objects.create(group="MARCA1", supplier=self.b)
        services.sync_principal_to_cheapest(self.product)
        self.product.refresh_from_db()
        self.assertEqual(self.product.default_supplier_id, self.b.id)

    def test_falls_back_to_cheapest_when_chosen_not_linked(self):
        # Marca apunta a un proveedor al que el producto NO está vinculado.
        other = Supplier.objects.create(name="Sin vínculo C")
        BrandSupplier.objects.create(group="Marca1", supplier=other)
        services.sync_principal_to_cheapest(self.product)
        self.product.refresh_from_db()
        self.assertEqual(self.product.default_supplier_id, self.a.id)

    def test_falls_back_when_chosen_has_no_price(self):
        c = Supplier.objects.create(name="Precio 0 C")
        SupplierProduct.objects.create(supplier=c, product=self.product, last_cost=Decimal("0"))
        BrandSupplier.objects.create(group="Marca1", supplier=c)
        services.sync_principal_to_cheapest(self.product)
        self.product.refresh_from_db()
        self.assertEqual(self.product.default_supplier_id, self.a.id)


class BrandSupplierViewTests(TestCase):
    def setUp(self):
        _reset_current_user()
        self.user = get_user_model().objects.create_user(username="u", password="x")
        self.client.force_login(self.user)
        self.a = Supplier.objects.create(name="Barato A")
        self.b = Supplier.objects.create(name="Caro B")
        self.product = Product.objects.create(name="Prod", group="Marca1", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(supplier=self.a, product=self.product, last_cost=Decimal("50"))
        SupplierProduct.objects.create(supplier=self.b, product=self.product, last_cost=Decimal("80"))
        services.sync_principal_to_cheapest(self.product)  # arranca en A

    def test_set_brand_supplier_reassigns_products(self):
        resp = self.client.post(reverse("inventory_suppliers"), {
            "action": "set_brand_supplier", "group": "Marca1", "supplier": self.b.id,
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(BrandSupplier.objects.filter(group="Marca1", supplier=self.b).exists())
        self.product.refresh_from_db()
        self.assertEqual(self.product.default_supplier_id, self.b.id)
        self.assertEqual(self.product.avg_cost, Decimal("80.00"))

    def test_remove_brand_supplier_reverts_to_cheapest(self):
        BrandSupplier.objects.create(group="Marca1", supplier=self.b)
        services.sync_principal_to_cheapest(self.product)
        resp = self.client.post(reverse("inventory_suppliers"), {
            "action": "remove_brand_supplier", "group": "Marca1",
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(BrandSupplier.objects.filter(group="Marca1").exists())
        self.product.refresh_from_db()
        self.assertEqual(self.product.default_supplier_id, self.a.id)


class BrandSupplierDownloadTests(TestCase):
    def setUp(self):
        _reset_current_user()
        self.user = get_user_model().objects.create_user(username="u", password="x")
        self.client.force_login(self.user)
        self.a = Supplier.objects.create(name="Prov A")
        self.b = Supplier.objects.create(name="Prov B")

    def test_download_uses_explicit_brand_supplier(self):
        # Dos productos de "Marca1": uno cuyo principal (por más barato) es A, otro B.
        p1 = Product.objects.create(name="Uno", group="Marca1", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(supplier=self.a, product=p1, last_cost=Decimal("10"))
        p2 = Product.objects.create(name="Dos", group="Marca1", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(supplier=self.b, product=p2, last_cost=Decimal("20"))
        services.sync_principal_to_cheapest(p1)
        services.sync_principal_to_cheapest(p2)
        # Sin BrandSupplier, el deducido (más frecuente) sería ambiguo (1 y 1).
        # Fijamos explícitamente B como principal de la marca.
        BrandSupplier.objects.create(group="Marca1", supplier=self.b)
        # Reasignar: ahora ambos deberían intentar B; p1 no está vinculado a B => cae a A.
        services.sync_principal_to_cheapest(p1)
        services.sync_principal_to_cheapest(p2)

        url = reverse("inventory_product_prices_download", args=["consumer"])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        ws = load_workbook(BytesIO(resp.content)).active
        nombres = [r[1] for r in ws.iter_rows(min_row=2, values_only=True)]
        # Solo el producto cuyo principal es B (el elegido para la marca) aparece.
        self.assertIn("Dos", nombres)
        self.assertNotIn("Uno", nombres)


class BulkBrandSupplierTests(TestCase):
    def setUp(self):
        _reset_current_user()
        self.user = get_user_model().objects.create_user(username="b", password="x")
        self.client.force_login(self.user)
        self.a = Supplier.objects.create(name="Prov A")
        self.b = Supplier.objects.create(name="Prov B")
        # Marca1: dos productos vinculados a A y B; A más barato.
        self.p1 = Product.objects.create(name="Uno", group="Marca1", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(supplier=self.a, product=self.p1, last_cost=Decimal("50"))
        SupplierProduct.objects.create(supplier=self.b, product=self.p1, last_cost=Decimal("80"))
        services.sync_principal_to_cheapest(self.p1)  # A
        # Marca2: un producto solo con A.
        self.p2 = Product.objects.create(name="Dos", group="Marca2", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(supplier=self.a, product=self.p2, last_cost=Decimal("30"))
        services.sync_principal_to_cheapest(self.p2)

    def test_bulk_assigns_and_reassigns(self):
        # Asigna Marca1 -> B (manda sobre A), Marca2 -> sin asignar (vacío).
        resp = self.client.post(reverse("inventory_brand_suppliers"), {
            "action": "bulk_set_brand_suppliers",
            "bg_group": ["Marca1", "Marca2"],
            "bg_supplier": [str(self.b.id), ""],
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(BrandSupplier.objects.filter(group="Marca1", supplier=self.b).exists())
        self.assertFalse(BrandSupplier.objects.filter(group="Marca2").exists())
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.default_supplier_id, self.b.id)  # reasignado a B
        self.assertEqual(self.p1.avg_cost, Decimal("80.00"))

    def test_bulk_removes_existing_when_emptied(self):
        BrandSupplier.objects.create(group="Marca1", supplier=self.b)
        services.sync_principal_to_cheapest(self.p1)  # B
        resp = self.client.post(reverse("inventory_brand_suppliers"), {
            "action": "bulk_set_brand_suppliers",
            "bg_group": ["Marca1"],
            "bg_supplier": [""],  # vaciar => quitar asignación
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(BrandSupplier.objects.filter(group="Marca1").exists())
        self.p1.refresh_from_db()
        self.assertEqual(self.p1.default_supplier_id, self.a.id)  # vuelve al más barato

    def test_brand_suppliers_page_loads(self):
        resp = self.client.get(reverse("inventory_brand_suppliers"))
        self.assertEqual(resp.status_code, 200)
        # La marca aparece en la tabla.
        self.assertContains(resp, "Marca1")


class BrandListNamingTests(TestCase):
    """La lista propia sale como la escribe el proveedor principal de la marca, y
    nunca trae productos que ese proveedor no tiene."""

    def setUp(self):
        _reset_current_user()
        self.user = get_user_model().objects.create_user(username="u2", password="x")
        self.client.force_login(self.user)
        self.aris = Supplier.objects.create(name="Aris Norma")
        self.glm = Supplier.objects.create(name="GLM Distribuidora")

    def _rows(self):
        url = reverse("inventory_product_prices_download", args=["consumer"])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        ws = load_workbook(BytesIO(resp.content)).active
        return list(ws.iter_rows(min_row=2, values_only=True))

    def test_name_comes_from_brand_principal_not_from_product(self):
        # Producto creado al importar la lista de GLM: Product.name quedó con SU texto.
        p = Product.objects.create(
            name="Fidelite balsam caviar x900", group="FIDELITE", margin_consumer=Decimal("0.00")
        )
        SupplierProduct.objects.create(
            supplier=self.glm, product=p, last_cost=Decimal("100"),
            supplier_name="Fidelite balsam caviar x900",
        )
        SupplierProduct.objects.create(
            supplier=self.aris, product=p, last_cost=Decimal("120"),
            supplier_name="BALSAMO CAVIAR 900ML FIDELITE",
        )
        BrandSupplier.objects.create(group="FIDELITE", supplier=self.aris)
        services.sync_principal_to_cheapest(p)

        nombres = [r[1] for r in self._rows()]
        self.assertEqual(nombres, ["BALSAMO CAVIAR 900ML FIDELITE"])

    def test_falls_back_to_product_name_when_principal_list_not_imported(self):
        p = Product.objects.create(name="Fidelite shampoo x300", group="FIDELITE", margin_consumer=Decimal("0.00"))
        # Vínculo con Aris cargado a mano (sin lista importada): no hay supplier_name.
        SupplierProduct.objects.create(supplier=self.aris, product=p, last_cost=Decimal("90"))
        BrandSupplier.objects.create(group="FIDELITE", supplier=self.aris)
        services.sync_principal_to_cheapest(p)

        self.assertEqual([r[1] for r in self._rows()], ["Fidelite shampoo x300"])

    def test_product_only_the_other_supplier_has_is_ignored(self):
        solo_glm = Product.objects.create(name="Fidelite crema GLM", group="FIDELITE", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(
            supplier=self.glm, product=solo_glm, last_cost=Decimal("70"), supplier_name="Fidelite crema GLM",
        )
        ambos = Product.objects.create(name="Fidelite mascara", group="FIDELITE", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(supplier=self.glm, product=ambos, last_cost=Decimal("50"))
        SupplierProduct.objects.create(
            supplier=self.aris, product=ambos, last_cost=Decimal("60"), supplier_name="MASCARA FIDELITE",
        )
        BrandSupplier.objects.create(group="FIDELITE", supplier=self.aris)
        for p in (solo_glm, ambos):
            services.sync_principal_to_cheapest(p)

        self.assertEqual([r[1] for r in self._rows()], ["MASCARA FIDELITE"])

    def test_ignores_other_supplier_even_if_default_supplier_is_stale(self):
        # default_supplier desincronizado (apunta al principal de la marca aunque el
        # producto no está en su lista): igual tiene que quedar afuera.
        solo_glm = Product.objects.create(
            name="Fidelite solo GLM", group="FIDELITE", margin_consumer=Decimal("0.00"),
            default_supplier=self.aris,
        )
        SupplierProduct.objects.create(supplier=self.glm, product=solo_glm, last_cost=Decimal("70"))
        BrandSupplier.objects.create(group="FIDELITE", supplier=self.aris)

        self.assertEqual(self._rows(), [])

    def test_brand_matching_ignores_case_and_spaces(self):
        p = Product.objects.create(name="Interno", group=" Fidelite ", margin_consumer=Decimal("0.00"))
        SupplierProduct.objects.create(
            supplier=self.aris, product=p, last_cost=Decimal("80"), supplier_name="COMO LO ESCRIBE ARIS",
        )
        SupplierProduct.objects.create(supplier=self.glm, product=p, last_cost=Decimal("10"))
        BrandSupplier.objects.create(group="FIDELITE", supplier=self.aris)

        rows = self._rows()
        self.assertEqual([r[1] for r in rows], ["COMO LO ESCRIBE ARIS"])
        # La marca sale con un texto único (el de BrandSupplier), sin los espacios
        # del group del producto.
        self.assertEqual([r[0] for r in rows], ["FIDELITE"])


class PriceListPreviewTests(TestCase):
    """La pantalla "Lista de precios" muestra lo mismo que el Excel."""

    def setUp(self):
        _reset_current_user()
        self.user = get_user_model().objects.create_user(username="u3", password="x")
        self.client.force_login(self.user)
        self.aris = Supplier.objects.create(name="Aris Norma")
        self.glm = Supplier.objects.create(name="GLM Distribuidora")

        self.ambos = Product.objects.create(
            name="Fidelite mascara", group="FIDELITE", margin_consumer=Decimal("0.00")
        )
        SupplierProduct.objects.create(supplier=self.glm, product=self.ambos, last_cost=Decimal("50"))
        SupplierProduct.objects.create(
            supplier=self.aris, product=self.ambos, last_cost=Decimal("60"),
            supplier_name="MASCARA FIDELITE",
        )
        self.solo_glm = Product.objects.create(
            name="Fidelite crema GLM", group="FIDELITE", margin_consumer=Decimal("0.00")
        )
        SupplierProduct.objects.create(supplier=self.glm, product=self.solo_glm, last_cost=Decimal("70"))
        BrandSupplier.objects.create(group="FIDELITE", supplier=self.aris)
        for p in (self.ambos, self.solo_glm):
            services.sync_principal_to_cheapest(p)

    def _entries(self, url):
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        return resp.context["entries"]

    def test_preview_matches_download(self):
        entries = self._entries(reverse("inventory_product_prices"))
        self.assertEqual([(e["brand"], e["name"]) for e in entries], [("FIDELITE", "MASCARA FIDELITE")])

        resp = self.client.get(reverse("inventory_product_prices_download", args=["consumer"]))
        ws = load_workbook(BytesIO(resp.content)).active
        excel = [(r[0], r[1]) for r in ws.iter_rows(min_row=2, values_only=True)]
        self.assertEqual([(e["brand"], e["name"]) for e in entries], excel)

    def test_todos_shows_excluded_products_flagged(self):
        entries = self._entries(reverse("inventory_product_prices") + "?todos=1")
        by_name = {e["name"]: e["in_list"] for e in entries}
        self.assertEqual(by_name, {"MASCARA FIDELITE": True, "Fidelite crema GLM": False})


class SuppliersPageLoadTests(TestCase):
    """La página de proveedores renderiza los vínculos con sus productos."""

    def setUp(self):
        _reset_current_user()
        self.user = get_user_model().objects.create_user(username="u4", password="x")
        self.client.force_login(self.user)

    def test_page_lists_supplier_products(self):
        supplier = Supplier.objects.create(name="Aris Norma")
        for i in range(3):
            product = Product.objects.create(name=f"Prod {i}", group="Marca1", sku=f"SKU{i}")
            SupplierProduct.objects.create(
                supplier=supplier, product=product, last_cost=Decimal("10"),
                supplier_name=f"NOMBRE ARIS {i}",
            )
        resp = self.client.get(reverse("inventory_suppliers"))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Prod 1")
