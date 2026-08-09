"""Descarga de lista de precios: solo productos del proveedor principal de cada
marca, y marcas en orden alfabético."""

from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from openpyxl import load_workbook

from inventory.models import Product, Supplier, SupplierProduct


class PriceListDownloadTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username="u", password="x")
        self.client.force_login(self.user)
        self.prov_a = Supplier.objects.create(name="Proveedor A")
        self.prov_b = Supplier.objects.create(name="Proveedor B")

    def _product(self, name, group, principal, cost):
        p = Product.objects.create(
            name=name, group=group, avg_cost=Decimal(cost),
            margin_consumer=Decimal("0.00"), default_supplier=principal,
        )
        SupplierProduct.objects.create(supplier=principal, product=p, last_cost=Decimal(cost))
        return p

    def _download_rows(self):
        url = reverse("inventory_product_prices_download", args=["consumer"])
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        return rows[0], rows[1:]  # header, data

    def test_only_principal_supplier_of_brand_and_alphabetical(self):
        # Marca "Zeta": principal es Proveedor A (2 productos).
        self._product("Zeta Uno", "Zeta", self.prov_a, "100")
        self._product("Zeta Dos", "Zeta", self.prov_a, "120")
        # Un producto de la misma marca cuyo principal es B (secundario/duplicado):
        # NO debe aparecer.
        self._product("Zeta Duplicado", "Zeta", self.prov_b, "999")
        # Marca "Alfa": principal Proveedor B.
        self._product("Alfa Uno", "Alfa", self.prov_b, "50")

        header, data = self._download_rows()
        self.assertEqual(header, ("Marca", "Producto", "Precio"))
        marcas = [r[0] for r in data]
        nombres = [r[1] for r in data]
        # Orden alfabético por marca: Alfa antes que Zeta.
        self.assertEqual(marcas, ["Alfa", "Zeta", "Zeta"])
        # El duplicado (proveedor B en marca cuyo principal es A) quedó afuera.
        self.assertNotIn("Zeta Duplicado", nombres)
        self.assertIn("Zeta Uno", nombres)
        self.assertIn("Alfa Uno", nombres)

    def test_product_without_brand_is_excluded(self):
        self._product("Sin Marca", "", self.prov_a, "100")
        self._product("Con Marca", "Beta", self.prov_a, "100")
        _, data = self._download_rows()
        nombres = [r[1] for r in data]
        self.assertEqual(nombres, ["Con Marca"])

    def test_kits_are_always_included(self):
        self._product("Base", "Gamma", self.prov_a, "100")
        Product.objects.create(name="Kit Especial", group="Gamma", is_kit=True, margin_consumer=Decimal("0.00"))
        _, data = self._download_rows()
        nombres = [r[1] for r in data]
        self.assertIn("Kit Especial", nombres)
